from __future__ import annotations

from pathlib import Path

import numpy as np
from openpyxl import load_workbook
from openpyxl.utils.cell import range_boundaries

from src.utils.charts_common import close_figure


def _text_color_for_bg_rgba(rgba) -> str:
    r, g, b = float(rgba[0]), float(rgba[1]), float(rgba[2])
    lum = 0.2126 * r + 0.7152 * g + 0.0722 * b
    return "#ffffff" if lum < 0.50 else "#2f2f2f"


def _wrap_words(text: str, *, max_line_len: int = 16) -> str:
    """Wrap a label to avoid overlapping the chart area."""

    s = (text or "").strip()
    if not s:
        return ""
    words = s.split()
    lines: list[str] = []
    cur: list[str] = []
    cur_len = 0
    for w in words:
        w_len = len(w)
        if not cur:
            cur = [w]
            cur_len = w_len
            continue

        if cur_len + 1 + w_len <= max_line_len:
            cur.append(w)
            cur_len += 1 + w_len
        else:
            lines.append(" ".join(cur))
            cur = [w]
            cur_len = w_len

    if cur:
        lines.append(" ".join(cur))
    return "\n".join(lines)


def _iter_linear_cells(ws, a1_range: str):
    min_col, min_row, max_col, max_row = range_boundaries(a1_range)

    if min_row != max_row and min_col != max_col:
        raise ValueError(f"Range precisa ser linear (1 linha ou 1 coluna): {a1_range}")

    if min_row == max_row:
        for c in range(min_col, max_col + 1):
            yield ws.cell(row=min_row, column=c)
        return

    for r in range(min_row, max_row + 1):
        yield ws.cell(row=r, column=min_col)


def _read_linear_labels(ws, a1_range: str) -> list[str]:
    labels: list[str] = []
    for idx, cell in enumerate(_iter_linear_cells(ws, a1_range), start=1):
        raw = cell.value
        labels.append(("" if raw is None else str(raw)).strip() or str(idx))
    return labels


def _read_linear_numeric_values(ws, a1_range: str) -> list[float]:
    values: list[float] = []
    for cell in _iter_linear_cells(ws, a1_range):
        raw = cell.value
        if raw is None or (isinstance(raw, str) and raw.strip() == ""):
            values.append(0.0)
            continue
        try:
            values.append(float(raw))
        except Exception as exc:
            raise ValueError(f"Valor não numérico em {ws.title}!{cell.coordinate}: {raw!r}") from exc
    return values


def _read_stacked_rows(
    ws,
    *,
    xlabels_range: str,
    series_specs: list[tuple[str, str]],
) -> tuple[list[str], list[str], np.ndarray]:
    xlabels = _read_linear_labels(ws, xlabels_range)
    point_count = len(xlabels)
    if point_count == 0:
        raise ValueError(f"Sem pontos para gerar gráfico em {ws.title}!{xlabels_range}")

    series_names: list[str] = []
    series_columns: list[list[float]] = []
    for series_name, values_range in series_specs:
        series_values = _read_linear_numeric_values(ws, values_range)
        if len(series_values) != point_count:
            raise ValueError(
                f"Quantidade de labels em {xlabels_range} difere de {values_range}: "
                f"{point_count} != {len(series_values)}"
            )
        series_names.append(series_name)
        series_columns.append(series_values)

    values = np.column_stack(series_columns)
    return xlabels, series_names, values


def _plot_stacked_vertical(
    *,
    xlabels: list[str],
    series_names: list[str],
    values: np.ndarray,
    output_path: Path,
    colors: list[str],
    font_scale: float = 1.3,
    show_delta_bracket: bool = True,
    show_delta_pct: bool = True,
    show_segment_pct: bool = True,
    bold_last_bar: bool = True,
    bold_text: bool = True,
    inline_left_legend: bool = True,
) -> None:
    import matplotlib.pyplot as plt
    from matplotlib.colors import to_rgba

    n = len(xlabels)
    if n == 0:
        raise ValueError("Sem barras para plotar")

    m = len(series_names)
    if values.shape != (n, m):
        raise ValueError(f"values shape inválido: {values.shape}, esperado {(n, m)}")

    if not colors:
        colors = ["#123a7a", "#8d98a6", "#2f2f2f"]

    fig, ax = plt.subplots(figsize=(10, 4.8), dpi=240)
    fig.patch.set_alpha(0)
    ax.set_facecolor("none")

    x = np.arange(n, dtype=float)
    bar_width = 0.62

    bottom = np.zeros(n, dtype=float)
    containers = []
    for j in range(m):
        y = np.asarray(values[:, j], dtype=float)
        cont = ax.bar(
            x,
            y,
            bottom=bottom,
            width=bar_width,
            color=colors[j % len(colors)],
            edgecolor="none",
            zorder=2,
        )
        containers.append(cont)
        bottom = bottom + np.nan_to_num(y, nan=0.0)

    totals = bottom

    # Expand x-limits to make room for inline legend on the left.
    if inline_left_legend:
        ax.set_xlim(float(x.min()) - 1.85, float(x.max()) + 0.75)

    # Segment labels: show value + % of bar total, centered inside each segment.
    for j, cont in enumerate(containers):
        face = colors[j % len(colors)]
        rgba = to_rgba(face)
        txt_color = _text_color_for_bg_rgba(rgba)

        for i, rect in enumerate(cont.patches):
            seg_val = float(values[i, j])
            if not np.isfinite(seg_val) or abs(seg_val) < 1e-12:
                continue
            total = float(totals[i])
            if not np.isfinite(total) or abs(total) < 1e-12:
                pct_txt = ""
            else:
                pct = (seg_val / total) * 100.0
                pct_txt = f"{pct:.0f}%".replace(".", ",")

            val_txt = f"{seg_val:,.0f}".replace(",", ".")
            label = val_txt
            if show_segment_pct and pct_txt:
                label = f"{val_txt}\n({pct_txt})"

            ax.text(
                rect.get_x() + rect.get_width() / 2,
                rect.get_y() + rect.get_height() / 2,
                label,
                ha="center",
                va="center",
                fontsize=8.8 * float(font_scale),
                color=txt_color,
                fontweight="bold" if (bold_text and bold_last_bar and i == n - 1) else "normal",
                zorder=5,
                clip_on=False,
            )

    # Totals on top
    for i, total in enumerate(totals):
        if not np.isfinite(total):
            continue
        ax.text(
            float(x[i]),
            float(total),
            f"{float(total):,.0f}".replace(",", "."),
            ha="center",
            va="bottom",
            fontsize=10.0 * float(font_scale),
            fontweight="bold" if (bold_text and bold_last_bar and i == n - 1) else "normal",
            color="#2f2f2f",
            zorder=4,
            clip_on=False,
        )

    # Inline legend on the left: color swatch + series name, vertically aligned with a reference bar.
    if inline_left_legend and m:
        # Prefer the last bar (usually the most relevant). If a segment is zero there,
        # fall back to the first bar where that segment is non-zero.
        ref_i = n - 1
        # Keep legend outside the bar area, but close to it.
        # Text wrapping handles long labels.
        x_swatch = float(x.min()) - 1.55

        for j, name in enumerate(series_names):
            seg_val = float(values[ref_i, j])
            i_for_pos = ref_i
            if not np.isfinite(seg_val) or abs(seg_val) < 1e-12:
                for ii in range(n):
                    vv = float(values[ii, j])
                    if np.isfinite(vv) and abs(vv) > 1e-12:
                        seg_val = vv
                        i_for_pos = ii
                        break

            # Compute the center of this segment within the stacked bar at i_for_pos.
            y0 = float(np.nansum(values[i_for_pos, :j]))
            yc = y0 + float(seg_val) / 2.0

            color = colors[j % len(colors)]
            ax.scatter(
                [x_swatch],
                [yc],
                s=140.0 * float(font_scale),
                marker="s",
                color=color,
                edgecolors="none",
                zorder=6,
                clip_on=False,
            )
            wrapped = _wrap_words(str(name), max_line_len=16)
            ax.text(
                x_swatch + 0.16,
                yc,
                wrapped,
                ha="left",
                va="center",
                fontsize=10.0 * float(font_scale),
                color="#2f2f2f",
                zorder=6,
                clip_on=False,
            )

    # Brackets / deltas between totals
    delta_label_top: float | None = None
    if show_delta_pct and n >= 2:
        abs_max = float(np.nanmax(np.abs(totals))) if np.isfinite(np.nanmax(np.abs(totals))) else 0.0
        offset_y = max(abs_max * 0.06, 0.5)
        bracket_h = max(abs_max * 0.03, 0.5)

        for level, i in enumerate(range(1, n)):
            prev = float(totals[i - 1])
            curr = float(totals[i])
            if not np.isfinite(prev) or not np.isfinite(curr) or prev == 0:
                continue
            pct = (curr / prev - 1.0) * 100.0
            label = f"{pct:+.1f}%".replace(".", ",")

            x1 = float(x[i - 1])
            x2 = float(x[i])
            top = max(prev, curr)
            y_anchor = top + offset_y + level * (bracket_h + offset_y * 0.9)

            if show_delta_bracket:
                ax.plot(
                    [x1, x1, x2, x2],
                    [y_anchor, y_anchor + bracket_h, y_anchor + bracket_h, y_anchor],
                    color="#2f2f2f",
                    linewidth=1.2,
                    solid_capstyle="round",
                    zorder=3,
                )
                text_y = y_anchor + bracket_h + offset_y * 0.25
            else:
                text_y = y_anchor

            ax.text(
                (x1 + x2) / 2.0,
                text_y,
                label,
                ha="center",
                va="bottom",
                fontsize=9.0 * float(font_scale),
                color="#2f2f2f",
                zorder=4,
                clip_on=False,
            )
            delta_label_top = text_y if delta_label_top is None else max(delta_label_top, text_y)

        if delta_label_top is not None:
            cur_ymin, cur_ymax = ax.get_ylim()
            ax.set_ylim(cur_ymin, max(cur_ymax, delta_label_top + offset_y))

    ax.set_xticks(x)
    ax.set_xticklabels(xlabels, fontsize=10.0 * float(font_scale))

    ax.set_yticks([])
    for s in ("left", "right", "top"):
        ax.spines[s].set_visible(False)
    ax.spines["bottom"].set_visible(True)
    ax.tick_params(axis="y", left=False, labelleft=False)
    ax.grid(False)
    ax.margins(x=0.05, y=0.12)

    fig.tight_layout(pad=0.2)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    fig.savefig(output_path, dpi=450, transparent=True, bbox_inches="tight", pad_inches=0.05)
    close_figure(fig)


def generate_slide8_charts(
    *,
    xlsx_path: Path,
    output_dir: Path,
) -> list[Path]:
    """Slide 8: gera 6 gráficos a partir das fontes originais em DRE Saida 2 e Tabelas."""

    output_dir.mkdir(parents=True, exist_ok=True)

    wb = load_workbook(filename=xlsx_path, data_only=True)
    dre_sheet_name = "DRE Saida 2"
    tabelas_sheet_name = "Tabelas"
    if dre_sheet_name not in wb.sheetnames:
        raise ValueError(f"Aba não encontrada: {dre_sheet_name!r}. Disponíveis: {wb.sheetnames}")
    if tabelas_sheet_name not in wb.sheetnames:
        raise ValueError(f"Aba não encontrada: {tabelas_sheet_name!r}. Disponíveis: {wb.sheetnames}")
    ws_dre = wb[dre_sheet_name]
    ws_tabelas = wb[tabelas_sheet_name]

    generated: list[Path] = []

    # Slide 8: dois tons de azul (escuro + claro)
    palette = ["#0B2E6B", "#5B8FF9"]

    # Optional map by series name (preserves semantic colors even if stacking order changes)
    color_by_key = {
        "margem financeira bruta": palette[0],
        "margem financeira bruta total": palette[0],
        "mercado": palette[0],
        "clientes": palette[1],
        "seguros": palette[0],
        "servicos e tarifas": palette[1],
        "serviços e tarifas": palette[1],
        "servicos e seguros": palette[1],
        "serviços e seguros": palette[1],
    }

    def _norm_key(s: str) -> str:
        return (s or "").strip().lower()

    def _reorder_for_stacking(series: list[str], vals: np.ndarray) -> tuple[list[str], np.ndarray, list[str]]:
        """Reorder stacked series so smaller totals go to the bottom.

        This improves readability and matches the earlier request (smaller 'Serviços...' at the bottom).
        """

        if not series or vals.size == 0:
            return series, vals, []

        # totals per series across all bars
        totals = np.nansum(vals.astype(float), axis=0)
        order = list(np.argsort(totals))  # ascending => smaller at bottom

        series2 = [series[i] for i in order]
        vals2 = vals[:, order]

        colors2: list[str] = []
        for j, name in enumerate(series2):
            key = _norm_key(name)
            colors2.append(color_by_key.get(key, palette[j % len(palette)]))

        return series2, vals2, colors2

    xlabels13, series13, values13 = _read_stacked_rows(
        ws_dre,
        xlabels_range="D3:F3",
        series_specs=[
            ("Margem Financeira Bruta", "D5:F5"),
            ("Serviços e Seguros", "D9:F9"),
        ],
    )
    xlabels14, series14, values14 = _read_stacked_rows(
        ws_dre,
        xlabels_range="G3:H3",
        series_specs=[
            ("Margem Financeira Bruta", "G5:H5"),
            ("Serviços e Seguros", "G9:H9"),
        ],
    )

    # 13) Trimestres
    series, values, colors = _reorder_for_stacking(series13, values13)
    out13 = output_dir / "13_slide8_trimestres.png"
    _plot_stacked_vertical(
        xlabels=xlabels13,
        series_names=series,
        values=values,
        output_path=out13,
        colors=colors,
        font_scale=1.3,
        show_delta_bracket=True,
        show_delta_pct=True,
        bold_last_bar=True,
        bold_text=True,
        inline_left_legend=True,
    )
    generated.append(out13)

    # 14) 9M
    series, values, colors = _reorder_for_stacking(series14, values14)
    out14 = output_dir / "14_slide8_9m.png"
    _plot_stacked_vertical(
        xlabels=xlabels14,
        series_names=series,
        values=values,
        output_path=out14,
        colors=colors,
        font_scale=1.3,
        show_delta_bracket=True,
        show_delta_pct=True,
        bold_last_bar=False,
        bold_text=False,
        inline_left_legend=True,
    )
    generated.append(out14)

    # 15/16) Margem Financeira Bruta Total (Trimestres + 9M)
    xlabels, series, values = _read_stacked_rows(
        ws_dre,
        xlabels_range="D3:F3",
        series_specs=[
            ("Clientes", "D6:F6"),
            ("Mercado", "D7:F7"),
        ],
    )
    series, values, colors = _reorder_for_stacking(series, values)
    out15 = output_dir / "15_margem_financeira_bruta_total_trimestres.png"
    _plot_stacked_vertical(
        xlabels=xlabels,
        series_names=series,
        values=values,
        output_path=out15,
        colors=colors,
        font_scale=1.3,
        show_delta_bracket=True,
        show_delta_pct=True,
        bold_last_bar=True,
        bold_text=True,
        inline_left_legend=True,
    )
    generated.append(out15)

    xlabels, series, values = _read_stacked_rows(
        ws_dre,
        xlabels_range="G3:H3",
        series_specs=[
            ("Clientes", "G6:H6"),
            ("Mercado", "G7:H7"),
        ],
    )
    series, values, colors = _reorder_for_stacking(series, values)
    out16 = output_dir / "16_margem_financeira_bruta_total_9m.png"
    _plot_stacked_vertical(
        xlabels=xlabels,
        series_names=series,
        values=values,
        output_path=out16,
        colors=colors,
        font_scale=1.3,
        show_delta_bracket=True,
        show_delta_pct=True,
        show_segment_pct=False,
        bold_last_bar=False,
        bold_text=False,
        inline_left_legend=True,
    )
    generated.append(out16)

    # 17/18) Receitas de Serviços e Corretagem (Trimestres + 9M)
    xlabels, series, values = _read_stacked_rows(
        ws_tabelas,
        xlabels_range="D16:F16",
        series_specs=[
            ("Seguros", "D19:F19"),
            ("Serviços e Seguros", "D28:F28"),
        ],
    )
    series, values, colors = _reorder_for_stacking(series, values)
    out17 = output_dir / "17_servicos_corretagem_trimestres.png"
    _plot_stacked_vertical(
        xlabels=xlabels,
        series_names=series,
        values=values,
        output_path=out17,
        colors=colors,
        font_scale=1.3,
        show_delta_bracket=True,
        show_delta_pct=True,
        show_segment_pct=False,
        bold_last_bar=True,
        bold_text=True,
        inline_left_legend=True,
    )
    generated.append(out17)

    xlabels, series, values = _read_stacked_rows(
        ws_tabelas,
        xlabels_range="G16:H16",
        series_specs=[
            ("Seguros", "G19:H19"),
            ("Serviços e Seguros", "G28:H28"),
        ],
    )
    series, values, colors = _reorder_for_stacking(series, values)
    out18 = output_dir / "18_servicos_corretagem_9m.png"
    _plot_stacked_vertical(
        xlabels=xlabels,
        series_names=series,
        values=values,
        output_path=out18,
        colors=colors,
        font_scale=1.3,
        show_delta_bracket=True,
        show_delta_pct=True,
        bold_last_bar=False,
        bold_text=False,
        inline_left_legend=True,
    )
    generated.append(out18)

    return generated


if __name__ == "__main__":
    xlsx = Path("testing.xlsx")
    out = Path(".")
    if xlsx.exists():
        files = generate_slide8_charts(xlsx_path=xlsx, output_dir=out)
        print(f"Gerados: {files}")
    else:
        print(f"Arquivo {xlsx} não encontrado")
