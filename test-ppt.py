from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path
from typing import List, Optional, Sequence, Tuple, Union

import matplotlib.pyplot as plt
import numpy as np
from openpyxl import load_workbook
from openpyxl.utils.cell import range_boundaries


@dataclass(frozen=True)
class ExcelBarChartSpec:
    file_path: Union[str, Path]
    sheet_name: str
    values_range: str
    xlabels_range: str
    ylabel_cell: Optional[str] = None
    title: Optional[str] = None
    highlight_last: bool = True
    output_path: Optional[Union[str, Path]] = None


def _read_range_row(ws, cell_range: str) -> List[object]:
    min_col, min_row, max_col, max_row = range_boundaries(cell_range)
    out: List[object] = []
    for r in range(min_row, max_row + 1):
        for c in range(min_col, max_col + 1):
            out.append(ws.cell(row=r, column=c).value)
    return out


def _to_float_list(values: List[object]) -> List[float]:
    out: List[float] = []
    for v in values:
        if v is None or (isinstance(v, str) and v.strip() == ""):
            out.append(0.0)
            continue
        try:
            out.append(float(v))
        except (TypeError, ValueError) as exc:
            raise ValueError(f"Valor não numérico no range: {v!r}") from exc
    return out


def plot_bar_from_excel(spec: ExcelBarChartSpec) -> Tuple[plt.Figure, plt.Axes]:
    file_path = Path(spec.file_path)
    if not file_path.exists():
        raise FileNotFoundError(f"Arquivo não encontrado: {file_path}")

    wb = load_workbook(filename=file_path, data_only=True)
    if spec.sheet_name not in wb.sheetnames:
        raise ValueError(f"Aba não encontrada: {spec.sheet_name!r}. Disponíveis: {wb.sheetnames}")

    ws = wb[spec.sheet_name]

    values = _to_float_list(_read_range_row(ws, spec.values_range))
    xlabels = ["" if v is None else str(v) for v in _read_range_row(ws, spec.xlabels_range)]

    if len(values) != len(xlabels):
        raise ValueError(
            f"Tamanhos diferentes: valores={len(values)} xlabels={len(xlabels)}"
        )

    ylabel = ""
    if spec.ylabel_cell:
        v = ws[spec.ylabel_cell].value
        ylabel = "" if v is None else str(v)

    fig, ax = plt.subplots(figsize=(10, 4.8))
    fig.patch.set_facecolor("white")
    ax.set_facecolor("white")

    colors = ["#8d98a6"] * len(values)
    if spec.highlight_last and colors:
        colors[-1] = "#123a7a"

    bars = ax.bar(np.arange(len(values)), values, color=colors, edgecolor="none")

    for rect, val in zip(bars, values):
        ax.text(
            rect.get_x() + rect.get_width() / 2,
            rect.get_height(),
            f"{val:,.0f}".replace(",", "."),
            ha="center",
            va="bottom",
            fontsize=9,
        )

    ax.set_xticks(np.arange(len(xlabels)))
    ax.set_xticklabels(xlabels, rotation=0, fontsize=9)
    ax.set_ylabel(ylabel)
    if spec.title:
        ax.set_title(spec.title)

    ax.yaxis.grid(True, linestyle="-", linewidth=0.6, alpha=0.25)
    ax.set_axisbelow(True)

    fig.tight_layout()

    if spec.output_path:
        out = Path(spec.output_path)
        out.parent.mkdir(parents=True, exist_ok=True)
        fig.savefig(out, dpi=220)

    return fig, ax


def plot_line_from_excel(
    file_path: Union[str, Path],
    sheet_name: str,
    values_range: str,
    xlabels_range: str,
    title: Optional[str] = None,
    output_path: Optional[Union[str, Path]] = None,
    fmt_as_percent: bool = True,
) -> Tuple[plt.Figure, plt.Axes]:
    file_path = Path(file_path)
    if not file_path.exists():
        raise FileNotFoundError(f"Arquivo não encontrado: {file_path}")

    wb = load_workbook(filename=file_path, data_only=True)
    if sheet_name not in wb.sheetnames:
        raise ValueError(f"Aba não encontrada: {sheet_name!r}. Disponíveis: {wb.sheetnames}")

    ws = wb[sheet_name]

    values = _to_float_list(_read_range_row(ws, values_range))
    xlabels = ["" if v is None else str(v) for v in _read_range_row(ws, xlabels_range)]
    if len(values) != len(xlabels):
        raise ValueError(
            f"Tamanhos diferentes: valores={len(values)} xlabels={len(xlabels)}"
        )

    x = np.arange(len(values))

    fig, ax = plt.subplots(figsize=(10, 4.2))
    ax.plot(x, values, linewidth=2.0, marker="o", markersize=4.5, color="#2f2f2f")

    for xi, yi in zip(x, values):
        label = f"{yi:.1f}%".replace(".", ",") if fmt_as_percent else str(yi)
        ax.text(xi, yi, label, ha="center", va="bottom", fontsize=9, color="#2f2f2f")

    ax.set_xticks(x)
    ax.set_xticklabels(xlabels, fontsize=9)

    if title:
        ax.set_title(title)

    ax.yaxis.grid(True, linestyle="-", linewidth=0.6, alpha=0.25)
    ax.set_axisbelow(True)

    fig.tight_layout()

    if output_path:
        out = Path(output_path)
        out.parent.mkdir(parents=True, exist_ok=True)
        fig.savefig(out, dpi=220)

    return fig, ax


if __name__ == "__main__":
    # 1) Gráfico barras: Trimestres (D:H)
    plot_bar_from_excel(
        ExcelBarChartSpec(
            file_path="seu_arquivo.xlsx",
            sheet_name="DRE Saida",
            values_range="D18:H18",
            xlabels_range="D3:H3",
            ylabel_cell="C18",
            title="Lucro Líquido - Trimestres",
            highlight_last=True,
            output_path="01_lucro_trimestres.png",
        )
    )
    plt.show()

    # 2) Gráfico barras: 9M (L:M)
    plot_bar_from_excel(
        ExcelBarChartSpec(
            file_path="seu_arquivo.xlsx",
            sheet_name="DRE Saida",
            values_range="L18:M18",
            xlabels_range="L3:M3",
            ylabel_cell="C18",
            title="Lucro Líquido - 9M",
            highlight_last=True,
            output_path="02_lucro_9m.png",
        )
    )
    plt.show()

    # 3) (Opcional) Linha ROE: Trimestres (ajuste o range do ROE no seu Excel)
    # plot_line_from_excel(
    #     file_path="seu_arquivo.xlsx",
    #     sheet_name="DRE Saida",
    #     values_range="D10:H10",
    #     xlabels_range="D3:H3",
    #     title="ROE - Trimestres",
    #     output_path="03_roe_trimestres.png",
    #     fmt_as_percent=True,
    # )
    # plt.show()

    # 4) (Opcional) Linha ROE: 9M (ajuste o range)
    # plot_line_from_excel(
    #     file_path="seu_arquivo.xlsx",
    #     sheet_name="DRE Saida",
    #     values_range="L10:M10",
    #     xlabels_range="L3:M3",
    #     title="ROE - 9M",
    #     output_path="04_roe_9m.png",
    #     fmt_as_percent=True,
    # )
    # plt.show()
