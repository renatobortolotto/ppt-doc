import unittest

import tempfile
from pathlib import Path

from matplotlib.colors import to_hex
from matplotlib.patches import Wedge
from openpyxl import Workbook

from src.utils.charts_common import (
    _spread_y_positions,
    ExcelBarChartSpec,
    close_figure,
    plot_bar_from_excel,
    plot_donut_chart,
    plot_line_from_excel,
    to_float_list,
)


class TestChartsCommon(unittest.TestCase):
    def test_spread_y_positions_enforces_min_gap(self):
        placed = _spread_y_positions(
            [0.20, 0.22, 0.25, 0.90],
            min_gap=0.15,
            lower=-1.0,
            upper=1.0,
        )
        self.assertEqual(len(placed), 4)
        self.assertGreaterEqual(placed[1] - placed[0], 0.15 - 1e-9)
        self.assertGreaterEqual(placed[2] - placed[1], 0.15 - 1e-9)
        self.assertGreaterEqual(placed[3] - placed[2], 0.15 - 1e-9)
        self.assertGreaterEqual(min(placed), -1.0 - 1e-9)
        self.assertLessEqual(max(placed), 1.0 + 1e-9)

    def test_to_float_list_parses_percent_strings(self):
        values = ["9%", " 9 % ", "9,5%", "(10%)"]
        out = to_float_list(values)
        self.assertEqual(out, [9.0, 9.0, 9.5, -10.0])

    def test_to_float_list_parses_ptbr_numbers(self):
        values = ["1.234,56", "0,09", "2.000", ""]
        out = to_float_list(values)
        self.assertEqual(out, [1234.56, 0.09, 2000.0, 0.0])

    def test_to_float_list_rejects_non_numeric(self):
        with self.assertRaises(ValueError):
            to_float_list(["N/A"])  # should still error

    def test_plot_line_percent_formatted_cells_scale_to_points(self):
        wb = Workbook()
        ws = wb.active
        ws.title = "S"

        # Values stored as fractions but formatted as percent
        ws["A1"].value = 0.09
        ws["A1"].number_format = "0%"
        ws["B1"].value = 0.12
        ws["B1"].number_format = "0%"

        ws["A2"].value = "T1"
        ws["B2"].value = "T2"

        with tempfile.TemporaryDirectory() as td:
            td_path = Path(td)
            xlsx_path = td_path / "t.xlsx"
            out_path = td_path / "out.png"
            wb.save(xlsx_path)

            fig, ax = plot_line_from_excel(
                file_path=xlsx_path,
                sheet_name="S",
                values_range="A1:B1",
                xlabels_range="A2:B2",
                output_path=out_path,
                fmt_as_percent=True,
                smooth=False,
                show_markers=False,
            )
            try:
                labels = [t.get_text() for t in ax.texts]
                self.assertEqual(labels, ["9,0%", "12,0%"])
            finally:
                close_figure(fig)

    def test_plot_bar_value_scale_and_comma_decimal_labels(self):
        wb = Workbook()
        ws = wb.active
        ws.title = "S"

        ws["A1"].value = 8820
        ws["B1"].value = 9100
        ws["A2"].value = "T1"
        ws["B2"].value = "T2"

        with tempfile.TemporaryDirectory() as td:
            td_path = Path(td)
            xlsx_path = td_path / "t.xlsx"
            out_path = td_path / "out.png"
            wb.save(xlsx_path)

            fig, ax = plot_bar_from_excel(
                ExcelBarChartSpec(
                    file_path=xlsx_path,
                    sheet_name="S",
                    values_range="A1:B1",
                    xlabels_range="A2:B2",
                    output_path=out_path,
                    value_decimals=1,
                    value_scale=0.001,
                    value_decimal_comma=True,
                )
            )
            try:
                labels = [t.get_text() for t in ax.texts]
                self.assertEqual(labels, ["8,8", "9,1"])
            finally:
                close_figure(fig)

    def test_plot_bar_respects_explicit_delta_pair_order(self):
        wb = Workbook()
        ws = wb.active
        ws.title = "S"

        ws["A1"].value = 100
        ws["B1"].value = 120
        ws["C1"].value = 150
        ws["A2"].value = "T1"
        ws["B2"].value = "T2"
        ws["C2"].value = "T3"

        with tempfile.TemporaryDirectory() as td:
            td_path = Path(td)
            xlsx_path = td_path / "t.xlsx"
            out_path = td_path / "out.png"
            wb.save(xlsx_path)

            fig, ax = plot_bar_from_excel(
                ExcelBarChartSpec(
                    file_path=xlsx_path,
                    sheet_name="S",
                    values_range="A1:C1",
                    xlabels_range="A2:C2",
                    output_path=out_path,
                    show_delta_pct=True,
                    show_delta_bracket=True,
                    delta_pairs=((0, 2), (1, 2)),
                )
            )
            try:
                delta_labels = [text.get_text() for text in ax.texts if "%" in text.get_text()]
                self.assertEqual(delta_labels, ["+50,0%", "+25,0%"])
            finally:
                close_figure(fig)

    def test_plot_bar_can_shift_delta_label_left_with_explicit_fraction(self):
        wb = Workbook()
        ws = wb.active
        ws.title = "S"

        ws["A1"].value = 100
        ws["B1"].value = 120
        ws["C1"].value = 150
        ws["A2"].value = "T1"
        ws["B2"].value = "T2"
        ws["C2"].value = "T3"

        with tempfile.TemporaryDirectory() as td:
            td_path = Path(td)
            xlsx_path = td_path / "t.xlsx"
            out_path = td_path / "out.png"
            wb.save(xlsx_path)

            fig, ax = plot_bar_from_excel(
                ExcelBarChartSpec(
                    file_path=xlsx_path,
                    sheet_name="S",
                    values_range="A1:C1",
                    xlabels_range="A2:C2",
                    output_path=out_path,
                    show_delta_pct=True,
                    show_delta_bracket=True,
                    delta_pairs=((0, 2), (1, 2)),
                    delta_label_x_fractions=(0.30, 0.50),
                )
            )
            try:
                delta_texts = [text for text in ax.texts if "%" in text.get_text()]
                self.assertEqual([text.get_text() for text in delta_texts], ["+50,0%", "+25,0%"])

                first_x = delta_texts[0].get_position()[0]
                second_x = delta_texts[1].get_position()[0]

                bar_centers = [
                    patch.get_x() + patch.get_width() / 2.0
                    for patch in ax.patches
                ]
                first_midpoint = (bar_centers[0] + bar_centers[2]) / 2.0
                second_midpoint = (bar_centers[1] + bar_centers[2]) / 2.0

                self.assertLess(first_x, first_midpoint)
                self.assertAlmostEqual(second_x, second_midpoint, places=6)
            finally:
                close_figure(fig)

    def test_plot_bar_gap_scale_reduces_only_the_space_between_bars(self):
        wb = Workbook()
        ws = wb.active
        ws.title = "S"

        ws["A1"].value = 100
        ws["B1"].value = 120
        ws["C1"].value = 150
        ws["A2"].value = "T1"
        ws["B2"].value = "T2"
        ws["C2"].value = "T3"

        with tempfile.TemporaryDirectory() as td:
            td_path = Path(td)
            xlsx_path = td_path / "t.xlsx"
            out_default = td_path / "out-default.png"
            out_tight = td_path / "out-tight.png"
            wb.save(xlsx_path)

            fig_default, ax_default = plot_bar_from_excel(
                ExcelBarChartSpec(
                    file_path=xlsx_path,
                    sheet_name="S",
                    values_range="A1:C1",
                    xlabels_range="A2:C2",
                    output_path=out_default,
                    bar_width_scale=0.5,
                )
            )
            fig_tight, ax_tight = plot_bar_from_excel(
                ExcelBarChartSpec(
                    file_path=xlsx_path,
                    sheet_name="S",
                    values_range="A1:C1",
                    xlabels_range="A2:C2",
                    output_path=out_tight,
                    bar_width_scale=0.5,
                    gap_scale=0.5,
                )
            )
            try:
                default_patches = ax_default.patches
                tight_patches = ax_tight.patches

                self.assertAlmostEqual(default_patches[0].get_width(), tight_patches[0].get_width(), places=6)

                default_gap = default_patches[1].get_x() - (
                    default_patches[0].get_x() + default_patches[0].get_width()
                )
                tight_gap = tight_patches[1].get_x() - (
                    tight_patches[0].get_x() + tight_patches[0].get_width()
                )

                self.assertAlmostEqual(tight_gap, default_gap * 0.5, places=6)
            finally:
                close_figure(fig_default)
                close_figure(fig_tight)

    def test_plot_bar_delta_offset_scale_raises_delta_labels(self):
        wb = Workbook()
        ws = wb.active
        ws.title = "S"

        ws["A1"].value = 100
        ws["B1"].value = 120
        ws["C1"].value = 150
        ws["A2"].value = "T1"
        ws["B2"].value = "T2"
        ws["C2"].value = "T3"

        with tempfile.TemporaryDirectory() as td:
            td_path = Path(td)
            xlsx_path = td_path / "t.xlsx"
            out_default = td_path / "out-default.png"
            out_raised = td_path / "out-raised.png"
            wb.save(xlsx_path)

            fig_default, ax_default = plot_bar_from_excel(
                ExcelBarChartSpec(
                    file_path=xlsx_path,
                    sheet_name="S",
                    values_range="A1:C1",
                    xlabels_range="A2:C2",
                    output_path=out_default,
                    show_delta_pct=True,
                    show_delta_bracket=True,
                    delta_pairs=((0, 2), (1, 2)),
                )
            )
            fig_raised, ax_raised = plot_bar_from_excel(
                ExcelBarChartSpec(
                    file_path=xlsx_path,
                    sheet_name="S",
                    values_range="A1:C1",
                    xlabels_range="A2:C2",
                    output_path=out_raised,
                    show_delta_pct=True,
                    show_delta_bracket=True,
                    delta_pairs=((0, 2), (1, 2)),
                    delta_offset_scale=1.25,
                )
            )
            try:
                default_texts = [text for text in ax_default.texts if "%" in text.get_text()]
                raised_texts = [text for text in ax_raised.texts if "%" in text.get_text()]

                self.assertEqual([text.get_text() for text in default_texts], ["+50,0%", "+25,0%"])
                self.assertEqual([text.get_text() for text in raised_texts], ["+50,0%", "+25,0%"])
                self.assertGreater(raised_texts[0].get_position()[1], default_texts[0].get_position()[1])
                self.assertGreater(raised_texts[1].get_position()[1], default_texts[1].get_position()[1])
            finally:
                close_figure(fig_default)
                close_figure(fig_raised)

    def test_plot_bar_can_customize_bracket_stroke_color_per_pair(self):
        wb = Workbook()
        ws = wb.active
        ws.title = "S"

        ws["A1"].value = 100
        ws["B1"].value = 120
        ws["C1"].value = 150
        ws["A2"].value = "T1"
        ws["B2"].value = "T2"
        ws["C2"].value = "T3"

        with tempfile.TemporaryDirectory() as td:
            td_path = Path(td)
            xlsx_path = td_path / "t.xlsx"
            out_path = td_path / "out.png"
            wb.save(xlsx_path)

            fig, ax = plot_bar_from_excel(
                ExcelBarChartSpec(
                    file_path=xlsx_path,
                    sheet_name="S",
                    values_range="A1:C1",
                    xlabels_range="A2:C2",
                    output_path=out_path,
                    show_delta_pct=True,
                    show_delta_bracket=True,
                    delta_pairs=((0, 2), (1, 2)),
                    delta_bracket_colors=("#123a7a", "#2f2f2f"),
                )
            )
            try:
                self.assertGreaterEqual(len(ax.lines), 2)
                self.assertEqual(to_hex(ax.lines[0].get_color()), "#123a7a")
                self.assertEqual(to_hex(ax.lines[1].get_color()), "#2f2f2f")

                delta_texts = [text for text in ax.texts if "%" in text.get_text()]
                self.assertEqual([to_hex(text.get_color()) for text in delta_texts], ["#2f2f2f", "#2f2f2f"])
            finally:
                close_figure(fig)

    def test_plot_donut_chart_keeps_group_ring_inside_detail_ring(self):
        with tempfile.TemporaryDirectory() as td:
            out_path = Path(td) / "donut.png"

            fig, ax = plot_donut_chart(
                categories=["A", "B", "B"],
                labels=["A1", "B1", "B2"],
                values=[10, 20, 30],
                center_text="Centro",
                output_path=out_path,
            )
            try:
                wedges = [patch for patch in ax.patches if isinstance(patch, Wedge)]
                radii = [round(float(wedge.r), 2) for wedge in wedges]
                self.assertEqual(radii.count(1.00), 2)
                self.assertEqual(radii.count(0.70), 3)
            finally:
                close_figure(fig)

    def test_plot_donut_chart_can_mirror_horizontally(self):
        with tempfile.TemporaryDirectory() as td:
            out_path = Path(td) / "donut-mirror.png"

            fig, ax = plot_donut_chart(
                categories=["Veiculos Leves", "Growth"],
                labels=["Veiculos Leves Usados", "EGV"],
                values=[50, 50],
                center_text="Centro",
                output_path=out_path,
                mirror_horizontal=True,
            )
            try:
                category_ann = next(
                    text for text in ax.texts if text.get_text().startswith("Veiculos Leves\n")
                )
                self.assertLess(category_ann.get_position()[0], 0.0)
            finally:
                close_figure(fig)

    def test_plot_donut_chart_places_growth_labels_closer_to_donut(self):
        with tempfile.TemporaryDirectory() as td:
            out_path = Path(td) / "donut-growth.png"

            fig, ax = plot_donut_chart(
                categories=["Veiculos Leves", "Growth", "Growth", "Atacado"],
                labels=["Veiculos Leves Usados", "EGV", "Cartões", "Corporate"],
                values=[40, 20, 15, 25],
                center_text="",
                output_path=out_path,
                mirror_horizontal=True,
            )
            try:
                growth_category_ann = next(
                    text for text in ax.texts if text.get_text().startswith("Growth\n")
                )
                egv_ann = next(
                    text for text in ax.texts if text.get_text().startswith("EGV\n")
                )
                cartoes_ann = next(
                    text for text in ax.texts if text.get_text().startswith("Cartões\n")
                )
                self.assertLess(abs(growth_category_ann.get_position()[0]), 1.0)
                self.assertLess(abs(egv_ann.get_position()[0]), 1.5)
                self.assertLess(abs(cartoes_ann.get_position()[0]), 1.5)
            finally:
                close_figure(fig)

    def test_plot_donut_chart_skips_center_text_when_blank(self):
        with tempfile.TemporaryDirectory() as td:
            out_path = Path(td) / "donut-no-center.png"

            fig, ax = plot_donut_chart(
                categories=["Veiculos Leves", "Growth"],
                labels=["Veiculos Leves Usados", "EGV"],
                values=[50, 50],
                center_text="",
                output_path=out_path,
            )
            try:
                center_texts = [
                    text for text in ax.texts
                    if text.get_position() == (0, 0) and str(text.get_text()).strip()
                ]
                self.assertEqual(center_texts, [])
            finally:
                close_figure(fig)


if __name__ == "__main__":
    unittest.main()
