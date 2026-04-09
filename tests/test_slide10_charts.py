import tempfile
import unittest
from pathlib import Path

from openpyxl import Workbook

from utils.slides.slide10_charts import _find_title_index, _pick_remaining, _sanitize_filename, generate_slide10_charts


class TestSlide10Charts(unittest.TestCase):
    def test_helper_functions_cover_title_and_filename_logic(self):
        self.assertEqual(_sanitize_filename("Serie A/B"), "Serie_AB")
        self.assertEqual(_find_title_index(["Varejo", "Atacado"], "atacad"), 1)
        self.assertEqual(_pick_remaining([1, 3], 5), [0, 2, 4])

    def test_generate_slide10_charts_creates_expected_pngs(self):
        wb = Workbook()
        ws = wb.active
        ws.title = "Qualidade Cart 2682"

        titles = ["Varejo", "Veiculos", "Total", "Atacado"]
        for row_offset, title in enumerate(titles, start=7):
            ws.cell(row=row_offset, column=2).value = title

        for col_offset in range(3, 17):
            ws.cell(row=6, column=col_offset).value = f"P{col_offset - 2}"

        base_rows = {
            7: [1.10, 1.12, 1.14, 1.15, 1.17, 1.18, 1.19, 1.20, 1.22, 1.24, 1.26, 1.28, 1.30, 1.32],
            8: [1.30, 1.28, 1.27, 1.25, 1.24, 1.22, 1.20, 1.19, 1.18, 1.16, 1.15, 1.13, 1.12, 1.10],
            9: [2.10, 2.09, 2.08, 2.07, 2.06, 2.05, 2.04, 2.03, 2.02, 2.01, 2.00, 1.99, 1.98, 1.97],
            10: [1.80, 1.78, 1.76, 1.75, 1.73, 1.71, 1.70, 1.69, 1.68, 1.66, 1.65, 1.64, 1.63, 1.61],
        }
        for row, values in base_rows.items():
            for col_offset, value in enumerate(values, start=3):
                ws.cell(row=row, column=col_offset).value = value

        with tempfile.TemporaryDirectory() as td:
            xlsx_path = Path(td) / "test.xlsx"
            output_dir = Path(td) / "out"
            wb.save(xlsx_path)

            files = generate_slide10_charts(xlsx_path=xlsx_path, output_dir=output_dir)
            self.assertEqual(
                [path.name for path in files],
                [
                    "05_qualidade_varejo_veiculos.png",
                    "06_qualidade_total.png",
                    "07_qualidade_atacado.png",
                ],
            )
            for file_path in files:
                self.assertTrue(file_path.exists())
                self.assertGreater(file_path.stat().st_size, 0)


if __name__ == "__main__":
    unittest.main()
