import tempfile
import unittest
from pathlib import Path
from unittest.mock import patch

import matplotlib.pyplot as plt
import numpy as np
from openpyxl import Workbook

from src.utils.slides.slide11_charts import (
    SLIDE11_TRI_DELTA_BRACKET_COLORS,
    SLIDE11_TRI_DELTA_LABEL_X_FRACTIONS,
    SLIDE11_FONT_SCALE,
    SLIDE11_TRI_DELTA_PAIRS,
    _normalize_expense_values,
    _plot_stacked_expenses,
    _read_named_series_rows,
    _wrap_words,
    generate_slide11_charts,
)


class TestSlide11Charts(unittest.TestCase):
    def test_wrap_words_breaks_depreciacao_label(self):
        self.assertEqual(
            _wrap_words("Depreciação e Amortização", max_line_len=16),
            "Depreciação e\nAmortização",
        )

    def test_read_named_series_rows_rejects_blank_ranges(self):
        wb = Workbook()
        ws = wb.active
        ws.title = "Tabelas"

        with self.assertRaisesRegex(ValueError, "Range sem labels"):
            _read_named_series_rows(
                ws,
                xlabels_range="D33:F33",
                series_specs=[
                    ("Pessoal", "D35:F35"),
                    ("Administrativas", "D39:F39"),
                ],
            )

    def test_normalize_expense_values_uses_absolute_magnitude(self):
        values = [[-1837.0, -1379.0, -407.0], [-1940.0, -1397.0, -475.0]]
        self.assertEqual(
            _normalize_expense_values(values).tolist(),
            [[1837.0, 1379.0, 407.0], [1940.0, 1397.0, 475.0]],
        )

    def test_generate_slide11_charts_uses_updated_expense_sources(self):
        wb = Workbook()
        ws_expenses = wb.active
        ws_expenses.title = "Tabelas"

        for col_offset, label in enumerate(["1T25", "2T25", "3T25"], start=4):
            ws_expenses.cell(row=33, column=col_offset).value = label
        for col_offset, value in enumerate([100, 110, 120], start=4):
            ws_expenses.cell(row=35, column=col_offset).value = value
        for col_offset, value in enumerate([40, 42, 44], start=4):
            ws_expenses.cell(row=39, column=col_offset).value = value
        for col_offset, value in enumerate([10, 11, 12], start=4):
            ws_expenses.cell(row=45, column=col_offset).value = value

        for col_offset, label in enumerate(["9M24", "9M25"], start=7):
            ws_expenses.cell(row=33, column=col_offset).value = label
        for col_offset, value in enumerate([330, 360], start=7):
            ws_expenses.cell(row=35, column=col_offset).value = value
        for col_offset, value in enumerate([126, 132], start=7):
            ws_expenses.cell(row=39, column=col_offset).value = value
        for col_offset, value in enumerate([31, 33], start=7):
            ws_expenses.cell(row=45, column=col_offset).value = value

        captured: dict[str, dict[str, object]] = {}

        def _capture_stacked(*, xlabels, series_names, values, output_path):
            captured[Path(output_path).name] = {
                "xlabels": list(xlabels),
                "series_names": list(series_names),
                "values": values.tolist(),
                "kwargs": {},
            }

        def _capture_stacked_with_kwargs(*, xlabels, series_names, values, output_path, **kwargs):
            captured[Path(output_path).name] = {
                "xlabels": list(xlabels),
                "series_names": list(series_names),
                "values": values.tolist(),
                "kwargs": dict(kwargs),
            }

        with tempfile.TemporaryDirectory() as td:
            xlsx_path = Path(td) / "test.xlsx"
            output_dir = Path(td) / "out"
            wb.save(xlsx_path)

            with patch("src.utils.slides.slide11_charts._plot_stacked_expenses", side_effect=_capture_stacked_with_kwargs):
                files = generate_slide11_charts(xlsx_path=xlsx_path, output_dir=output_dir)

        self.assertEqual(len(files), 2)
        self.assertEqual(
            captured["11_despesas_pessoal_adm_trimestres.png"]["xlabels"],
            ["1T25", "2T25", "3T25"],
        )
        self.assertEqual(
            captured["11_despesas_pessoal_adm_trimestres.png"]["series_names"],
            ["Pessoal", "Administrativas", "Depreciação e Amortização"],
        )
        self.assertEqual(
            captured["11_despesas_pessoal_adm_trimestres.png"]["values"],
            [[100.0, 40.0, 10.0], [110.0, 42.0, 11.0], [120.0, 44.0, 12.0]],
        )
        self.assertEqual(
            captured["11_despesas_pessoal_adm_trimestres.png"]["kwargs"],
            {
                "font_scale": SLIDE11_FONT_SCALE,
                "delta_pairs": SLIDE11_TRI_DELTA_PAIRS,
                "delta_bracket_colors": SLIDE11_TRI_DELTA_BRACKET_COLORS,
                "delta_label_x_fractions": SLIDE11_TRI_DELTA_LABEL_X_FRACTIONS,
            },
        )

        self.assertEqual(
            captured["11_despesas_pessoal_adm_9m.png"]["xlabels"],
            ["9M24", "9M25"],
        )
        self.assertEqual(
            captured["11_despesas_pessoal_adm_9m.png"]["values"],
            [[330.0, 126.0, 31.0], [360.0, 132.0, 33.0]],
        )
        self.assertEqual(
            captured["11_despesas_pessoal_adm_9m.png"]["kwargs"],
            {
                "font_scale": SLIDE11_FONT_SCALE,
            },
        )

    def test_generate_slide11_charts_converts_negative_expenses_to_positive(self):
        wb = Workbook()
        ws_expenses = wb.active
        ws_expenses.title = "Tabelas"

        for col_offset, label in enumerate(["2024", "2025"], start=7):
            ws_expenses.cell(row=33, column=col_offset).value = label
        for col_offset, value in enumerate([-1837, -1940], start=7):
            ws_expenses.cell(row=35, column=col_offset).value = value
        for col_offset, value in enumerate([-1379, -1397], start=7):
            ws_expenses.cell(row=39, column=col_offset).value = value
        for col_offset, value in enumerate([-407, -475], start=7):
            ws_expenses.cell(row=45, column=col_offset).value = value

        for col_offset, label in enumerate(["1T25", "2T25", "3T25"], start=4):
            ws_expenses.cell(row=33, column=col_offset).value = label
        for col_offset, value in enumerate([100, 110, 120], start=4):
            ws_expenses.cell(row=35, column=col_offset).value = value
        for col_offset, value in enumerate([40, 42, 44], start=4):
            ws_expenses.cell(row=39, column=col_offset).value = value
        for col_offset, value in enumerate([10, 11, 12], start=4):
            ws_expenses.cell(row=45, column=col_offset).value = value

        captured: dict[str, dict[str, object]] = {}

        def _capture_stacked(*, xlabels, series_names, values, output_path, **kwargs):
            captured[Path(output_path).name] = {
                "xlabels": list(xlabels),
                "series_names": list(series_names),
                "values": values.tolist(),
                "kwargs": dict(kwargs),
            }

        with tempfile.TemporaryDirectory() as td:
            xlsx_path = Path(td) / "test.xlsx"
            output_dir = Path(td) / "out"
            wb.save(xlsx_path)

            with patch("src.utils.slides.slide11_charts._plot_stacked_expenses", side_effect=_capture_stacked):
                generate_slide11_charts(xlsx_path=xlsx_path, output_dir=output_dir)

        self.assertEqual(
            captured["11_despesas_pessoal_adm_9m.png"]["values"],
            [[1837.0, 1379.0, 407.0], [1940.0, 1397.0, 475.0]],
        )

    def test_plot_stacked_expenses_uses_requested_bracket_pairs_and_font_scale(self):
        fig, ax = plt.subplots()

        with tempfile.TemporaryDirectory() as td:
            output_path = Path(td) / "chart.png"
            with patch("matplotlib.pyplot.subplots", return_value=(fig, ax)):
                with patch("src.utils.slides.slide11_charts.close_figure"):
                    _plot_stacked_expenses(
                        xlabels=["1T25", "2T25", "3T25"],
                        series_names=["Pessoal", "Administrativas", "Depreciação e Amortização"],
                        values=np.asarray(
                            [
                                [100.0, 40.0, 10.0],
                                [110.0, 42.0, 11.0],
                                [120.0, 44.0, 12.0],
                            ],
                            dtype=float,
                        ),
                        output_path=output_path,
                        font_scale=SLIDE11_FONT_SCALE,
                        delta_pairs=SLIDE11_TRI_DELTA_PAIRS,
                        delta_bracket_colors=SLIDE11_TRI_DELTA_BRACKET_COLORS,
                        delta_label_x_fractions=SLIDE11_TRI_DELTA_LABEL_X_FRACTIONS,
                    )

        bracket_lines = [line for line in ax.lines if len(line.get_xdata()) == 4]
        self.assertEqual(len(bracket_lines), 2)
        self.assertEqual(bracket_lines[0].get_color(), SLIDE11_TRI_DELTA_BRACKET_COLORS[0])
        self.assertEqual(bracket_lines[1].get_color(), SLIDE11_TRI_DELTA_BRACKET_COLORS[1])
        self.assertAlmostEqual(bracket_lines[0].get_xdata()[0], 0.0, places=2)
        self.assertAlmostEqual(bracket_lines[0].get_xdata()[2], 0.44, places=2)
        self.assertAlmostEqual(bracket_lines[1].get_xdata()[0], 0.22, places=2)
        self.assertAlmostEqual(bracket_lines[1].get_xdata()[2], 0.44, places=2)

        percent_texts = [text for text in ax.texts if "%" in text.get_text()]
        self.assertEqual(len(percent_texts), 2)
        self.assertTrue(all(abs(text.get_fontsize() - 13.5) < 0.01 for text in percent_texts))
        self.assertAlmostEqual(percent_texts[0].get_position()[0], 0.13, places=2)
        self.assertAlmostEqual(percent_texts[1].get_position()[0], 0.33, places=2)

        xtick_labels = ax.get_xticklabels()
        self.assertTrue(all(abs(label.get_fontsize() - 15.0) < 0.01 for label in xtick_labels))

        plt.close(fig)


if __name__ == "__main__":
    unittest.main()
