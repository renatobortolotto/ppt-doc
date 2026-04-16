import tempfile
import unittest
from pathlib import Path
from unittest.mock import patch

import matplotlib.pyplot as plt
import numpy as np
from openpyxl import Workbook

from src.utils.slides.slide12_charts import (
    SLIDE12_DELTA_BRACKET_COLORS,
    SLIDE12_DELTA_LABEL_X_FRACTIONS,
    SLIDE12_DELTA_PAIRS,
    SLIDE12_FONT_SCALE,
    _plot_slide12_stacked,
    generate_slide12_charts,
)


class TestSlide12Charts(unittest.TestCase):
    def test_generate_slide12_charts_uses_carteira_ranges(self):
        wb = Workbook()
        ws = wb.active
        ws.title = "Carteira"

        for col_offset, label in enumerate(["3T25", "4T25", "1T26"], start=4):
            ws.cell(row=12, column=col_offset).value = label

        ws["C15"] = "Varejo - Produtos de Entrada"
        ws["C22"] = "Varejo Relacional"
        ws["C34"] = "Atacado"

        for col_offset, value in enumerate([10.0, 11.0, 12.0], start=4):
            ws.cell(row=15, column=col_offset).value = value
        for col_offset, value in enumerate([20.0, 21.0, 22.0], start=4):
            ws.cell(row=22, column=col_offset).value = value
        for col_offset, value in enumerate([30.0, 31.0, 32.0], start=4):
            ws.cell(row=34, column=col_offset).value = value

        captured: dict[str, object] = {}

        def _capture_plot(
            *,
            xlabels,
            series_names,
            values,
            output_path,
            font_scale=1.0,
            bracket_top_gap_scale=0.20,
            bracket_top_gap_min=1.4,
            delta_pairs=(),
            delta_bracket_colors=(),
            delta_label_x_fractions=(),
        ):
            captured["xlabels"] = list(xlabels)
            captured["series_names"] = list(series_names)
            captured["values"] = values.tolist()
            captured["output_name"] = Path(output_path).name
            captured["font_scale"] = font_scale
            captured["bracket_top_gap_scale"] = bracket_top_gap_scale
            captured["bracket_top_gap_min"] = bracket_top_gap_min
            captured["delta_pairs"] = delta_pairs
            captured["delta_bracket_colors"] = delta_bracket_colors
            captured["delta_label_x_fractions"] = delta_label_x_fractions

        with tempfile.TemporaryDirectory() as td:
            xlsx_path = Path(td) / "test.xlsx"
            output_dir = Path(td) / "out"
            wb.save(xlsx_path)

            with patch("src.utils.slides.slide12_charts._plot_slide12_stacked", side_effect=_capture_plot):
                files = generate_slide12_charts(xlsx_path=xlsx_path, output_dir=output_dir)

        self.assertEqual([path.name for path in files], ["12_slide12_composicao.png"])
        self.assertEqual(captured["output_name"], "12_slide12_composicao.png")
        self.assertEqual(captured["xlabels"], ["3T25", "4T25", "1T26"])
        self.assertEqual(
            captured["series_names"],
            ["Varejo - Produtos de Entrada", "Varejo Relacional", "Atacado"],
        )
        self.assertEqual(
            captured["values"],
            [
                [0.01, 0.02, 0.03],
                [0.011, 0.021, 0.031],
                [0.012, 0.022, 0.032],
            ],
        )
        self.assertEqual(captured["font_scale"], SLIDE12_FONT_SCALE)
        self.assertEqual(captured["bracket_top_gap_scale"], 0.12)
        self.assertEqual(captured["bracket_top_gap_min"], 0.9)
        self.assertEqual(captured["delta_pairs"], SLIDE12_DELTA_PAIRS)
        self.assertEqual(captured["delta_bracket_colors"], SLIDE12_DELTA_BRACKET_COLORS)
        self.assertEqual(captured["delta_label_x_fractions"], SLIDE12_DELTA_LABEL_X_FRACTIONS)

    def test_plot_slide12_stacked_uses_requested_bracket_pairs_and_font_scale(self):
        fig, ax = plt.subplots()

        with tempfile.TemporaryDirectory() as td:
            output_path = Path(td) / "chart.png"
            with patch("matplotlib.pyplot.subplots", return_value=(fig, ax)):
                with patch("src.utils.slides.slide12_charts.close_figure"):
                    _plot_slide12_stacked(
                        xlabels=["3T25", "4T25", "1T26"],
                        series_names=["Entrada", "Relacional", "Atacado"],
                        values=np.asarray(
                            [
                                [10.0, 20.0, 30.0],
                                [11.0, 21.0, 31.0],
                                [12.0, 22.0, 32.0],
                            ],
                            dtype=float,
                        ),
                        output_path=output_path,
                        font_scale=SLIDE12_FONT_SCALE,
                        delta_pairs=SLIDE12_DELTA_PAIRS,
                        delta_bracket_colors=SLIDE12_DELTA_BRACKET_COLORS,
                        delta_label_x_fractions=SLIDE12_DELTA_LABEL_X_FRACTIONS,
                    )

        bracket_lines = [line for line in ax.lines if len(line.get_xdata()) == 4]
        self.assertEqual(len(bracket_lines), 2)
        self.assertEqual(bracket_lines[0].get_color(), SLIDE12_DELTA_BRACKET_COLORS[0])
        self.assertEqual(bracket_lines[1].get_color(), SLIDE12_DELTA_BRACKET_COLORS[1])
        self.assertAlmostEqual(bracket_lines[0].get_xdata()[0], 0.0, places=2)
        self.assertAlmostEqual(bracket_lines[0].get_xdata()[2], 2.0, places=2)
        self.assertAlmostEqual(bracket_lines[1].get_xdata()[0], 1.0, places=2)
        self.assertAlmostEqual(bracket_lines[1].get_xdata()[2], 2.0, places=2)

        percent_texts = [text for text in ax.texts if "%" in text.get_text()]
        self.assertEqual(len(percent_texts), 11)
        bracket_texts = [text for text in percent_texts if text.get_text().startswith(("+", "-"))]
        self.assertEqual(len(bracket_texts), 2)
        self.assertTrue(all(abs(text.get_fontsize() - 12.6) < 0.01 for text in bracket_texts))
        self.assertAlmostEqual(bracket_texts[0].get_position()[0], 0.60, places=2)
        self.assertAlmostEqual(bracket_texts[1].get_position()[0], 1.50, places=2)

        xtick_labels = ax.get_xticklabels()
        self.assertTrue(all(abs(label.get_fontsize() - 14.0) < 0.01 for label in xtick_labels))


if __name__ == "__main__":
    unittest.main()
