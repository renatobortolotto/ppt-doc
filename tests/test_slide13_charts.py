import tempfile
import unittest
from pathlib import Path
from unittest.mock import patch

import numpy as np
from openpyxl import Workbook

from utils.slide13_charts import (
    _fmt_num,
    _share_label_indices,
    _stack_order_for_bar,
    _top_badge_label_indices,
    SLIDE13_ATACADO_PALETTE,
    SLIDE13_PALETTE,
    generate_slide13_charts,
)


class TestSlide13Charts(unittest.TestCase):
    def test_fmt_num_keeps_two_decimals_for_small_non_zero_values(self):
        self.assertEqual(_fmt_num(0.04), "0,04")
        self.assertEqual(_fmt_num(0.4), "0,4")

    def test_helpers_sort_largest_values_to_the_bottom_and_limit_pct_labels(self):
        self.assertEqual(_stack_order_for_bar(np.asarray([1.0, 3.0, 2.0])), (1, 2, 0))
        self.assertEqual(_share_label_indices(np.asarray([9.0, 8.0, 7.0, 6.0, 5.0])), {0, 1, 2})
        self.assertEqual(_share_label_indices(np.asarray([9.0, 8.0, 7.0])), {0, 1})
        self.assertEqual(_top_badge_label_indices(np.asarray([9.0, 5.0, 1.0, 0.6])), {2, 3})
        self.assertEqual(_top_badge_label_indices(np.asarray([9.0, 5.0, 3.0, 2.0])), {3})

    def test_generate_slide13_charts_uses_updated_carteira_ranges(self):
        wb = Workbook()
        ws = wb.active
        ws.title = "Carteira"

        for col_offset, label in enumerate(["3T25", "4T25", "1T26"], start=4):
            ws.cell(row=12, column=col_offset).value = label
            ws.cell(row=115, column=col_offset).value = label

        for col_offset, value in enumerate([1000.0, 1100.0, 1200.0], start=4):
            ws.cell(row=16, column=col_offset).value = value
        for col_offset, value in enumerate([800.0, 900.0, 950.0], start=4):
            ws.cell(row=18, column=col_offset).value = value
        for col_offset, value in enumerate([700.0, 750.0, 800.0], start=4):
            ws.cell(row=17, column=col_offset).value = value
        for col_offset, value in enumerate([100.0, 120.0, 140.0], start=4):
            ws.cell(row=19, column=col_offset).value = value
        for col_offset, value in enumerate([40.0, 50.0, 60.0], start=4):
            ws.cell(row=20, column=col_offset).value = value
        for col_offset, value in enumerate([4.0, 5.0, 6.0], start=4):
            ws.cell(row=21, column=col_offset).value = value

        for col_offset, value in enumerate([500.0, 550.0, 600.0], start=4):
            ws.cell(row=25, column=col_offset).value = value
        for col_offset, value in enumerate([400.0, 450.0, 470.0], start=4):
            ws.cell(row=24, column=col_offset).value = value
        for col_offset, value in enumerate([300.0, 320.0, 340.0], start=4):
            ws.cell(row=23, column=col_offset).value = value

        for col_offset, value in enumerate([2000.0, 2100.0, 2200.0], start=4):
            ws.cell(row=119, column=col_offset).value = value
        for col_offset, value in enumerate([1500.0, 1600.0, 1700.0], start=4):
            ws.cell(row=117, column=col_offset).value = value
        for col_offset, value in enumerate([1000.0, 1050.0, 1100.0], start=4):
            ws.cell(row=118, column=col_offset).value = value

        captured: dict[str, dict[str, object]] = {}

        def _capture_plot(
            *,
            xlabels,
            series_names,
            values,
            output_path,
            palette=SLIDE13_PALETTE,
            figure_size=(7.4, 6.0),
            bar_slot=0.24,
            bar_width=0.20,
        ):
            captured[Path(output_path).name] = {
                "xlabels": list(xlabels),
                "series_names": list(series_names),
                "values": values.tolist(),
                "palette": tuple(palette),
                "figure_size": tuple(figure_size),
                "bar_slot": bar_slot,
                "bar_width": bar_width,
            }

        with tempfile.TemporaryDirectory() as td:
            xlsx_path = Path(td) / "test.xlsx"
            output_dir = Path(td) / "out"
            wb.save(xlsx_path)

            with patch("utils.slide13_charts._plot_slide13_breakdown", side_effect=_capture_plot):
                files = generate_slide13_charts(xlsx_path=xlsx_path, output_dir=output_dir)

        self.assertEqual(
            [path.name for path in files],
            [
                "13_varejo_produtos_entrada.png",
                "13_varejo_relacional.png",
                "13_atacado.png",
            ],
        )
        self.assertEqual(
            captured["13_varejo_produtos_entrada.png"]["xlabels"],
            ["3T25", "4T25", "1T26"],
        )
        self.assertEqual(
            captured["13_varejo_produtos_entrada.png"]["series_names"],
            ["Leves e Usados", "Motos e Novos", "Pesados", "Solar", "Outros"],
        )
        self.assertEqual(captured["13_varejo_produtos_entrada.png"]["palette"], SLIDE13_PALETTE)
        self.assertEqual(captured["13_varejo_produtos_entrada.png"]["figure_size"], (7.4, 6.0))
        self.assertEqual(captured["13_varejo_produtos_entrada.png"]["bar_slot"], 0.24)
        self.assertEqual(captured["13_varejo_produtos_entrada.png"]["bar_width"], 0.20)
        self.assertEqual(
            captured["13_varejo_produtos_entrada.png"]["values"],
            [
                [1.0, 0.8, 0.7, 0.1, 0.044],
                [1.1, 0.9, 0.75, 0.12, 0.055],
                [1.2, 0.95, 0.8, 0.14, 0.066],
            ],
        )
        self.assertEqual(
            captured["13_varejo_relacional.png"]["series_names"],
            ["Crédito Pessoal", "Cartões", "EGV"],
        )
        self.assertEqual(
            captured["13_varejo_relacional.png"]["values"],
            [
                [0.5, 0.4, 0.3],
                [0.55, 0.45, 0.32],
                [0.6, 0.47, 0.34],
            ],
        )
        self.assertEqual(
            captured["13_atacado.png"]["xlabels"],
            ["3T25", "4T25", "1T26"],
        )
        self.assertEqual(
            captured["13_atacado.png"]["series_names"],
            ["PMEs", "Corporate", "Large e IF"],
        )
        self.assertEqual(captured["13_atacado.png"]["palette"], SLIDE13_ATACADO_PALETTE)
        self.assertEqual(
            captured["13_atacado.png"]["values"],
            [
                [2.0, 1.5, 1.0],
                [2.1, 1.6, 1.05],
                [2.2, 1.7, 1.1],
            ],
        )


if __name__ == "__main__":
    unittest.main()
