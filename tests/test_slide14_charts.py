import tempfile
import unittest
from pathlib import Path
from unittest.mock import patch

import matplotlib.pyplot as plt
import numpy as np
from openpyxl import Workbook

from src.utils.slides.slide14_charts import (
    SLIDE14_TRI_DELTA_BRACKET_COLORS,
    SLIDE14_TRI_DELTA_LABEL_X_FRACTIONS,
    SLIDE14_TRI_DELTA_PAIRS,
    _fmt_pct_trunc,
    _largest_segment_indices,
    _normalize_percent_values,
    _plot_stacked_veiculos,
    generate_slide14_charts,
)


class TestSlide14Charts(unittest.TestCase):
    def test_fmt_pct_trunc_uses_half_down_threshold(self):
        self.assertEqual(_fmt_pct_trunc(39.4), "39%")
        self.assertEqual(_fmt_pct_trunc(39.5), "39%")
        self.assertEqual(_fmt_pct_trunc(39.6), "40%")

    def test_largest_segment_indices_marks_only_bigger_values(self):
        self.assertEqual(_largest_segment_indices([100.0, 50.0]), {0})
        self.assertEqual(_largest_segment_indices([50.0, 100.0]), {1})
        self.assertEqual(_largest_segment_indices([50.0, 50.0]), {0, 1})

    def test_normalize_percent_values_accepts_fraction_and_whole_percent(self):
        self.assertEqual(_normalize_percent_values([31.0, 42.0]), [0.31, 0.42])
        self.assertEqual(_normalize_percent_values([0.31, 0.42]), [0.31, 0.42])

    def test_generate_slide14_charts_uses_veiculos_ranges(self):
        wb = Workbook()
        ws = wb.active
        ws.title = "Veículos"

        for col_offset, label in enumerate(["3T25", "4T25", "1T26"], start=4):
            ws.cell(row=3, column=col_offset).value = label
        for col_offset, label in enumerate(["2025", "2026"], start=7):
            ws.cell(row=3, column=col_offset).value = label

        ws["C10"] = "Leves"
        ws["C11"] = "Pesados"

        for col_offset, value in enumerate([100.0, 110.0, 120.0], start=4):
            ws.cell(row=10, column=col_offset).value = value
        for col_offset, value in enumerate([50.0, 55.0, 60.0], start=4):
            ws.cell(row=11, column=col_offset).value = value

        for col_offset, value in enumerate([230.0, 250.0], start=7):
            ws.cell(row=10, column=col_offset).value = value
        for col_offset, value in enumerate([90.0, 100.0], start=7):
            ws.cell(row=11, column=col_offset).value = value

        for col_offset, value in enumerate([31.0, 32.0, 33.0], start=4):
            ws.cell(row=7, column=col_offset).value = value
        for col_offset, value in enumerate([41.0, 43.0], start=7):
            ws.cell(row=7, column=col_offset).value = value

        captured_stacked: dict[str, dict[str, object]] = {}
        captured_percent: dict[str, dict[str, object]] = {}

        def _capture_stacked(*, xlabels, series_names, values, output_path, **kwargs):
            captured_stacked[Path(output_path).name] = {
                "xlabels": list(xlabels),
                "series_names": list(series_names),
                "values": values.tolist(),
                "kwargs": dict(kwargs),
            }

        def _capture_percent(*, xlabels, values, output_path):
            captured_percent[Path(output_path).name] = {
                "xlabels": list(xlabels),
                "values": list(values),
            }

        with tempfile.TemporaryDirectory() as td:
            xlsx_path = Path(td) / "test.xlsx"
            output_dir = Path(td) / "out"
            wb.save(xlsx_path)

            with patch("src.utils.slides.slide14_charts._plot_stacked_veiculos", side_effect=_capture_stacked):
                with patch("src.utils.slides.slide14_charts._plot_percent_bars", side_effect=_capture_percent):
                    files = generate_slide14_charts(xlsx_path=xlsx_path, output_dir=output_dir)

        self.assertEqual(
            [path.name for path in files],
            [
                "14_veiculos_empilhado_trimestres.png",
                "14_veiculos_empilhado_anos.png",
                "14_veiculos_percentual_trimestres.png",
                "14_veiculos_percentual_anos.png",
            ],
        )
        self.assertEqual(
            captured_stacked["14_veiculos_empilhado_trimestres.png"]["xlabels"],
            ["3T25", "4T25", "1T26"],
        )
        self.assertEqual(
            captured_stacked["14_veiculos_empilhado_trimestres.png"]["series_names"],
            ["Leves", "Pesados"],
        )
        self.assertEqual(
            captured_stacked["14_veiculos_empilhado_trimestres.png"]["values"],
            [
                [100.0, 50.0],
                [110.0, 55.0],
                [120.0, 60.0],
            ],
        )
        self.assertEqual(
            captured_stacked["14_veiculos_empilhado_trimestres.png"]["kwargs"],
            {
                "delta_pairs": SLIDE14_TRI_DELTA_PAIRS,
                "delta_bracket_colors": SLIDE14_TRI_DELTA_BRACKET_COLORS,
                "delta_label_x_fractions": SLIDE14_TRI_DELTA_LABEL_X_FRACTIONS,
            },
        )
        self.assertEqual(
            captured_stacked["14_veiculos_empilhado_anos.png"]["xlabels"],
            ["2025", "2026"],
        )
        self.assertEqual(
            captured_stacked["14_veiculos_empilhado_anos.png"]["values"],
            [
                [230.0, 90.0],
                [250.0, 100.0],
            ],
        )
        self.assertEqual(
            captured_stacked["14_veiculos_empilhado_anos.png"]["kwargs"],
            {
                "delta_pairs": (),
                "delta_bracket_colors": (),
                "delta_label_x_fractions": (),
            },
        )
        self.assertEqual(
            captured_percent["14_veiculos_percentual_trimestres.png"]["values"],
            [0.31, 0.32, 0.33],
        )
        self.assertEqual(
            captured_percent["14_veiculos_percentual_anos.png"]["xlabels"],
            ["2025", "2026"],
        )
        self.assertEqual(
            captured_percent["14_veiculos_percentual_anos.png"]["values"],
            [0.41, 0.43],
        )

    def test_plot_stacked_veiculos_supports_first_to_third_bracket_for_trimestres(self):
        fig, ax = plt.subplots()

        with tempfile.TemporaryDirectory() as td:
            output_path = Path(td) / "chart.png"
            with patch("matplotlib.pyplot.subplots", return_value=(fig, ax)):
                with patch("src.utils.slides.slide14_charts.close_figure"):
                    _plot_stacked_veiculos(
                        xlabels=["3T25", "4T25", "1T26"],
                        series_names=["Leves", "Pesados"],
                        values=np.asarray(
                            [
                                [100.0, 50.0],
                                [110.0, 55.0],
                                [120.0, 60.0],
                            ],
                            dtype=float,
                        ),
                        output_path=output_path,
                        delta_pairs=SLIDE14_TRI_DELTA_PAIRS,
                        delta_bracket_colors=SLIDE14_TRI_DELTA_BRACKET_COLORS,
                        delta_label_x_fractions=SLIDE14_TRI_DELTA_LABEL_X_FRACTIONS,
                    )

        bracket_lines = [line for line in ax.lines if len(line.get_xdata()) == 4]
        self.assertEqual(len(bracket_lines), 2)
        self.assertEqual(bracket_lines[0].get_color(), SLIDE14_TRI_DELTA_BRACKET_COLORS[0])
        self.assertEqual(bracket_lines[1].get_color(), SLIDE14_TRI_DELTA_BRACKET_COLORS[1])
        self.assertAlmostEqual(bracket_lines[0].get_xdata()[0], 0.0, places=2)
        self.assertAlmostEqual(bracket_lines[0].get_xdata()[2], 0.48, places=2)
        self.assertAlmostEqual(bracket_lines[1].get_xdata()[0], 0.24, places=2)
        self.assertAlmostEqual(bracket_lines[1].get_xdata()[2], 0.48, places=2)

        percent_texts = [text for text in ax.texts if text.get_text().startswith(("+", "-"))]
        self.assertEqual(len(percent_texts), 2)
        self.assertAlmostEqual(percent_texts[0].get_position()[0], 0.14, places=2)
        self.assertAlmostEqual(percent_texts[1].get_position()[0], 0.36, places=2)

        plt.close(fig)


if __name__ == "__main__":
    unittest.main()
