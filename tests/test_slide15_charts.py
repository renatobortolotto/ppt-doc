import tempfile
import unittest
from pathlib import Path
from unittest.mock import patch

import matplotlib.pyplot as plt
from openpyxl import Workbook

import numpy as np

from src.utils.slides.slide15_charts import (
    SLIDE15_DELTA_BRACKET_COLORS,
    SLIDE15_DELTA_LABEL_X_FRACTIONS,
    SLIDE15_DELTA_PAIRS,
    SLIDE15_OUTPUT,
    _plot_stacked_captacoes,
    _should_render_zero_label,
    _top_share_indices,
    generate_slide15_charts,
)


class TestSlide15Charts(unittest.TestCase):
    def test_helpers_mark_top_two_values_and_fidc_zero_labels(self):
        self.assertEqual(_top_share_indices(np.asarray([100.0, 80.0, 40.0, 30.0])), {0, 1})
        self.assertEqual(_should_render_zero_label("FIDC"), True)
        self.assertEqual(_should_render_zero_label("Outros"), False)

    def test_generate_slide15_charts_uses_captacoes_ranges(self):
        wb = Workbook()
        ws = wb.active
        ws.title = "Captações"

        for col_offset, label in enumerate(["3T25", "4T25", "1T26"], start=3):
            ws.cell(row=3, column=col_offset).value = label

        rows = {
            5: ("Vista", [100.0, 110.0, 120.0]),
            6: ("Prazo", [80.0, 85.0, 90.0]),
            10: ("LCA", [40.0, 42.0, 44.0]),
            11: ("LCI", [30.0, 32.0, 34.0]),
            12: ("LF", [20.0, 21.0, 22.0]),
            13: ("Debêntures", [15.0, 16.0, 17.0]),
            14: ("FIDC", [10.0, 11.0, 12.0]),
            15: ("Outros", [5.0, 6.0, 7.0]),
        }
        for row, (name, values) in rows.items():
            ws.cell(row=row, column=2).value = name
            for col_offset, value in enumerate(values, start=3):
                ws.cell(row=row, column=col_offset).value = value

        captured: dict[str, object] = {}

        def _capture_plot(*, xlabels, series_names, values, output_path, **kwargs):
            captured["xlabels"] = list(xlabels)
            captured["series_names"] = list(series_names)
            captured["values"] = values.tolist()
            captured["output_name"] = Path(output_path).name
            captured["kwargs"] = dict(kwargs)

        with tempfile.TemporaryDirectory() as td:
            xlsx_path = Path(td) / "test.xlsx"
            output_dir = Path(td) / "out"
            wb.save(xlsx_path)

            with patch("src.utils.slides.slide15_charts._plot_stacked_captacoes", side_effect=_capture_plot):
                files = generate_slide15_charts(xlsx_path=xlsx_path, output_dir=output_dir)

        self.assertEqual([path.name for path in files], [SLIDE15_OUTPUT])
        self.assertEqual(captured["output_name"], SLIDE15_OUTPUT)
        self.assertEqual(captured["xlabels"], ["3T25", "4T25", "1T26"])
        self.assertEqual(
            captured["series_names"],
            ["Vista", "Prazo", "LCA", "LCI", "LF", "Debêntures", "FIDC", "Outros"],
        )
        self.assertEqual(
            captured["values"],
            [
                [100.0, 80.0, 40.0, 30.0, 20.0, 15.0, 10.0, 5.0],
                [110.0, 85.0, 42.0, 32.0, 21.0, 16.0, 11.0, 6.0],
                [120.0, 90.0, 44.0, 34.0, 22.0, 17.0, 12.0, 7.0],
            ],
        )
        self.assertEqual(
            captured["kwargs"],
            {
                "delta_pairs": SLIDE15_DELTA_PAIRS,
                "delta_bracket_colors": SLIDE15_DELTA_BRACKET_COLORS,
                "delta_label_x_fractions": SLIDE15_DELTA_LABEL_X_FRACTIONS,
            },
        )

    def test_plot_stacked_captacoes_uses_slide14_bracket_style(self):
        fig, ax = plt.subplots()

        with tempfile.TemporaryDirectory() as td:
            output_path = Path(td) / "chart.png"
            with patch("matplotlib.pyplot.subplots", return_value=(fig, ax)):
                with patch("src.utils.slides.slide15_charts.close_figure"):
                    _plot_stacked_captacoes(
                        xlabels=["3T25", "4T25", "1T26"],
                        series_names=["Vista", "Prazo", "LCA"],
                        values=np.asarray(
                            [
                                [100.0, 80.0, 40.0],
                                [110.0, 85.0, 42.0],
                                [120.0, 90.0, 44.0],
                            ],
                            dtype=float,
                        ),
                        output_path=output_path,
                        delta_pairs=SLIDE15_DELTA_PAIRS,
                        delta_bracket_colors=SLIDE15_DELTA_BRACKET_COLORS,
                        delta_label_x_fractions=SLIDE15_DELTA_LABEL_X_FRACTIONS,
                    )

        bracket_lines = [line for line in ax.lines if len(line.get_xdata()) == 4]
        self.assertEqual(len(bracket_lines), 2)
        self.assertEqual(bracket_lines[0].get_color(), SLIDE15_DELTA_BRACKET_COLORS[0])
        self.assertEqual(bracket_lines[1].get_color(), SLIDE15_DELTA_BRACKET_COLORS[1])
        self.assertAlmostEqual(bracket_lines[0].get_xdata()[0], 0.0, places=2)
        self.assertAlmostEqual(bracket_lines[0].get_xdata()[2], 2.68, places=2)
        self.assertAlmostEqual(bracket_lines[1].get_xdata()[0], 1.34, places=2)
        self.assertAlmostEqual(bracket_lines[1].get_xdata()[2], 2.68, places=2)

        percent_texts = [text for text in ax.texts if text.get_text().startswith(("+", "-"))]
        self.assertEqual(len(percent_texts), 2)
        self.assertAlmostEqual(percent_texts[0].get_position()[0], 0.80, places=2)
        self.assertAlmostEqual(percent_texts[1].get_position()[0], 2.01, places=2)

        plt.close(fig)


if __name__ == "__main__":
    unittest.main()
