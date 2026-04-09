import tempfile
import unittest
from pathlib import Path
from unittest.mock import patch

from openpyxl import Workbook

from utils.slides.slide16_charts import (
    SLIDE16_OUTPUT,
    generate_slide16_charts,
)


class TestSlide16Charts(unittest.TestCase):
    def test_generate_slide16_charts_uses_basileia_ranges(self):
        wb = Workbook()
        ws = wb.active
        ws.title = "Basiléia"

        for col_offset, label in enumerate(["4T24", "3T25", "4T25"], start=3):
            ws.cell(row=3, column=col_offset).value = label

        ws["C18"] = 0.128
        ws["D18"] = 0.131
        ws["E18"] = 0.128
        ws["C19"] = 0.017
        ws["D19"] = 0.024
        ws["E19"] = 0.025
        ws["C20"] = 0.015
        ws["D20"] = 0.014
        ws["E20"] = 0.015

        captured: dict[str, object] = {}

        def _capture_plot(*, xlabels, series_names, values_pct, output_path):
            captured["xlabels"] = list(xlabels)
            captured["series_names"] = list(series_names)
            captured["values_pct"] = values_pct.tolist()
            captured["output_name"] = Path(output_path).name

        with tempfile.TemporaryDirectory() as td:
            xlsx_path = Path(td) / "test.xlsx"
            output_dir = Path(td) / "out"
            wb.save(xlsx_path)

            with patch("utils.slides.slide16_charts._plot_stacked_basileia", side_effect=_capture_plot):
                files = generate_slide16_charts(xlsx_path=xlsx_path, output_dir=output_dir)

        self.assertEqual([path.name for path in files], [SLIDE16_OUTPUT])
        self.assertEqual(captured["output_name"], SLIDE16_OUTPUT)
        self.assertEqual(captured["xlabels"], ["4T24", "3T25", "4T25"])
        self.assertEqual(
            captured["series_names"],
            ["Nível I Principal", "Nível I Complementar", "Nível II"],
        )
        self.assertEqual(
            captured["values_pct"],
            [
                [12.8, 1.7000000000000002, 1.5],
                [13.100000000000001, 2.4, 1.4000000000000001],
                [12.8, 2.5, 1.5],
            ],
        )


if __name__ == "__main__":
    unittest.main()
