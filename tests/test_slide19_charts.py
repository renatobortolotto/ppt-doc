import tempfile
import unittest
from pathlib import Path
from unittest.mock import patch

import numpy as np
from openpyxl import Workbook

from src.utils.slides.slide19_charts import (
    SLIDE19_SEGUROS_ANOS_OUTPUT,
    SLIDE19_SEGUROS_TRIMESTRES_OUTPUT,
    SLIDE19_VEICULOS_OUTPUT,
    generate_slide19_charts,
)


class TestSlide19Charts(unittest.TestCase):
    def test_generate_slide19_charts_uses_veiculos_ranges(self):
        wb = Workbook()
        ws = wb.active
        ws.title = "Veículos"
        ws_seguros = wb.create_sheet("Seguros e Cartões")

        for col_offset, label in enumerate(["4T24", "3T25", "4T25"], start=4):
            ws.cell(row=14, column=col_offset).value = label

        for col_offset, value in enumerate([42.236, 44.852, 46.888], start=4):
            ws.cell(row=21, column=col_offset).value = value

        row22 = [2.219, 2.913, 3.258]
        row23 = [3.548, 4.262, 4.558]
        for col_offset, value in enumerate(row22, start=4):
            ws.cell(row=22, column=col_offset).value = value
        for col_offset, value in enumerate(row23, start=4):
            ws.cell(row=23, column=col_offset).value = value

        for col_offset, label in enumerate(["4T24", "3T25", "4T25"], start=4):
            ws_seguros.cell(row=3, column=col_offset).value = label
        for col_offset, value in enumerate([420.0, 354.0, 418.0], start=4):
            ws_seguros.cell(row=8, column=col_offset).value = value
        for col_offset, label in enumerate(["2024", "2025"], start=7):
            ws_seguros.cell(row=3, column=col_offset).value = label
        for col_offset, value in enumerate([1672.0, 1452.0], start=7):
            ws_seguros.cell(row=8, column=col_offset).value = value

        captured: dict[str, object] = {"simple_calls": []}

        def _capture_plot(*, xlabels, series_names, values, output_path):
            captured["xlabels"] = list(xlabels)
            captured["series_names"] = list(series_names)
            captured["values"] = values.tolist()
            captured["veiculos_output_name"] = Path(output_path).name

        def _capture_simple(
            *,
            xlabels,
            values,
            output_path,
            bar_color="#123A7A",
            bar_width=0.62,
            bar_slot=1.0,
        ):
            captured["simple_calls"].append(
                {
                    "xlabels": list(xlabels),
                    "values": list(values),
                    "output_name": Path(output_path).name,
                    "bar_color": bar_color,
                    "bar_width": bar_width,
                    "bar_slot": bar_slot,
                }
            )

        with tempfile.TemporaryDirectory() as td:
            xlsx_path = Path(td) / "test.xlsx"
            output_dir = Path(td) / "out"
            wb.save(xlsx_path)

            with patch("src.utils.slides.slide19_charts._plot_stacked_veiculos", side_effect=_capture_plot):
                with patch("src.utils.slides.slide19_charts._plot_simple_bars", side_effect=_capture_simple):
                    files = generate_slide19_charts(xlsx_path=xlsx_path, output_dir=output_dir)

        self.assertEqual(
            [path.name for path in files],
            [SLIDE19_VEICULOS_OUTPUT, SLIDE19_SEGUROS_TRIMESTRES_OUTPUT, SLIDE19_SEGUROS_ANOS_OUTPUT],
        )
        self.assertEqual(captured["veiculos_output_name"], SLIDE19_VEICULOS_OUTPUT)
        self.assertEqual(captured["xlabels"], ["4T24", "3T25", "4T25"])
        self.assertEqual(captured["series_names"], ["Leves Usados", "Outros Veículos"])
        np.testing.assert_allclose(
            captured["values"],
            [
                [42.236, 5.767],
                [44.852, 7.175],
                [46.888, 7.816],
            ],
            rtol=0.0,
            atol=1e-9,
        )
        self.assertEqual(
            captured["simple_calls"],
            [
                {
                    "xlabels": ["4T24", "3T25", "4T25"],
                    "values": [420.0, 354.0, 418.0],
                    "output_name": SLIDE19_SEGUROS_TRIMESTRES_OUTPUT,
                    "bar_color": "#123A7A",
                    "bar_width": 0.62,
                    "bar_slot": 1.0,
                },
                {
                    "xlabels": ["2024", "2025"],
                    "values": [1672.0, 1452.0],
                    "output_name": SLIDE19_SEGUROS_ANOS_OUTPUT,
                    "bar_color": "#123A7A",
                    "bar_width": 0.22,
                    "bar_slot": 0.38,
                },
            ],
        )


if __name__ == "__main__":
    unittest.main()
