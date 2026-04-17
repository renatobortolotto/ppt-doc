import tempfile
import unittest
from pathlib import Path
from unittest.mock import patch

import matplotlib.pyplot as plt
import numpy as np
from openpyxl import Workbook

from src.utils.slides.slide20_charts import (
    SLIDE20_DELTA_BRACKET_COLORS,
    SLIDE20_DELTA_LABEL_X_FRACTIONS,
    SLIDE20_DELTA_PAIRS,
    SLIDE20_EMPRESTIMOS_OUTPUT,
    SLIDE20_SEGUROS_OUTPUT,
    _plot_simple_bars,
    _plot_stacked_bars,
    generate_slide20_charts,
)


class TestSlide20Charts(unittest.TestCase):
    def test_generate_slide20_charts_uses_new_ranges(self):
        wb = Workbook()
        ws_emp = wb.active
        ws_emp.title = "Empréstimos"
        ws_seg = wb.create_sheet("Seguros e Cartões")

        for col_offset, label in enumerate(["3T24", "2T25", "3T25"], start=4):
            ws_emp.cell(row=3, column=col_offset).value = label
        for col_offset, value in enumerate([169.0, 93.0, 74.0], start=4):
            ws_emp.cell(row=5, column=col_offset).value = value
        for col_offset, value in enumerate([4167.0, 3795.0, 3731.0], start=4):
            ws_emp.cell(row=6, column=col_offset).value = value
        for col_offset, value in enumerate([513.0, 395.0, 333.0], start=4):
            ws_emp.cell(row=7, column=col_offset).value = value
        for col_offset, value in enumerate([4032.0, 4797.0, 5262.0], start=4):
            ws_emp.cell(row=8, column=col_offset).value = value

        for col_offset, label in enumerate(["4T24", "3T25", "4T25"], start=4):
            ws_seg.cell(row=11, column=col_offset).value = label
        for col_offset, value in enumerate([4761.0, 4829.0, 5230.0], start=4):
            ws_seg.cell(row=15, column=col_offset).value = value

        captured: dict[str, object] = {}

        def _capture_stacked(*, xlabels, series_names, values, output_path, colors, **kwargs):
            captured["stacked_xlabels"] = list(xlabels)
            captured["stacked_series_names"] = list(series_names)
            captured["stacked_values"] = values.tolist()
            captured["stacked_output_name"] = Path(output_path).name
            captured["stacked_colors"] = list(colors)
            captured["stacked_kwargs"] = dict(kwargs)

        def _capture_simple(*, xlabels, values, output_path, bar_color="#123A7A", **kwargs):
            captured["simple_xlabels"] = list(xlabels)
            captured["simple_values"] = list(values)
            captured["simple_output_name"] = Path(output_path).name
            captured["simple_bar_color"] = bar_color
            captured["simple_kwargs"] = dict(kwargs)

        with tempfile.TemporaryDirectory() as td:
            xlsx_path = Path(td) / "test.xlsx"
            output_dir = Path(td) / "out"
            wb.save(xlsx_path)

            with patch("src.utils.slides.slide20_charts._plot_stacked_bars", side_effect=_capture_stacked):
                with patch("src.utils.slides.slide20_charts._plot_simple_bars", side_effect=_capture_simple):
                    files = generate_slide20_charts(xlsx_path=xlsx_path, output_dir=output_dir)

        self.assertEqual([path.name for path in files], [SLIDE20_EMPRESTIMOS_OUTPUT, SLIDE20_SEGUROS_OUTPUT])
        self.assertEqual(captured["stacked_output_name"], SLIDE20_EMPRESTIMOS_OUTPUT)
        self.assertEqual(
            captured["stacked_kwargs"],
            {
                "delta_pairs": SLIDE20_DELTA_PAIRS,
                "delta_bracket_colors": SLIDE20_DELTA_BRACKET_COLORS,
                "delta_label_x_fractions": SLIDE20_DELTA_LABEL_X_FRACTIONS,
            },
        )
        self.assertEqual(captured["stacked_xlabels"], ["3T24", "2T25", "3T25"])
        self.assertEqual(captured["stacked_series_names"], ["EGV", "Placas Solares", "Outros"])
        np.testing.assert_allclose(
            captured["stacked_values"],
            [
                [4.032, 4.167, 0.682],
                [4.797, 3.795, 0.488],
                [5.262, 3.731, 0.407],
            ],
            rtol=0.0,
            atol=1e-9,
        )
        self.assertEqual(captured["simple_output_name"], SLIDE20_SEGUROS_OUTPUT)
        self.assertEqual(captured["simple_xlabels"], ["4T24", "3T25", "4T25"])
        np.testing.assert_allclose(captured["simple_values"], [4.761, 4.829, 5.230], rtol=0.0, atol=1e-9)
        self.assertEqual(captured["simple_bar_color"], "#123A7A")
        self.assertEqual(
            captured["simple_kwargs"],
            {
                "delta_pairs": SLIDE20_DELTA_PAIRS,
                "delta_bracket_colors": SLIDE20_DELTA_BRACKET_COLORS,
                "delta_label_x_fractions": SLIDE20_DELTA_LABEL_X_FRACTIONS,
            },
        )

    def test_plot_stacked_bars_uses_slide15_bracket_style(self):
        fig, ax = plt.subplots()

        with tempfile.TemporaryDirectory() as td:
            output_path = Path(td) / "chart.png"
            with patch("matplotlib.pyplot.subplots", return_value=(fig, ax)):
                with patch("src.utils.slides.slide20_charts.close_figure"):
                    _plot_stacked_bars(
                        xlabels=["3T24", "2T25", "3T25"],
                        series_names=["EGV", "Placas Solares", "Outros"],
                        values=np.asarray(
                            [
                                [4.032, 4.167, 0.682],
                                [4.797, 3.795, 0.488],
                                [5.262, 3.731, 0.407],
                            ],
                            dtype=float,
                        ),
                        output_path=output_path,
                        colors=["#123A7A", "#5B8FF9", "#AFC8F5"],
                        delta_pairs=SLIDE20_DELTA_PAIRS,
                        delta_bracket_colors=SLIDE20_DELTA_BRACKET_COLORS,
                        delta_label_x_fractions=SLIDE20_DELTA_LABEL_X_FRACTIONS,
                    )

        bracket_lines = [line for line in ax.lines if len(line.get_xdata()) == 4]
        self.assertEqual(len(bracket_lines), 2)
        self.assertEqual(bracket_lines[0].get_color(), SLIDE20_DELTA_BRACKET_COLORS[0])
        self.assertEqual(bracket_lines[1].get_color(), SLIDE20_DELTA_BRACKET_COLORS[1])
        self.assertAlmostEqual(bracket_lines[0].get_xdata()[0], 0.0, places=2)
        self.assertAlmostEqual(bracket_lines[0].get_xdata()[2], 2.0, places=2)
        self.assertAlmostEqual(bracket_lines[1].get_xdata()[0], 1.0, places=2)
        self.assertAlmostEqual(bracket_lines[1].get_xdata()[2], 2.0, places=2)

        bracket_texts = [text for text in ax.texts if text.get_text().startswith(("+", "-"))]
        self.assertEqual(len(bracket_texts), 2)
        self.assertAlmostEqual(bracket_texts[0].get_position()[0], 0.60, places=2)
        self.assertAlmostEqual(bracket_texts[1].get_position()[0], 1.50, places=2)

        plt.close(fig)

    def test_plot_simple_bars_uses_slide15_bracket_style(self):
        fig, ax = plt.subplots()

        with tempfile.TemporaryDirectory() as td:
            output_path = Path(td) / "chart.png"
            with patch("matplotlib.pyplot.subplots", return_value=(fig, ax)):
                with patch("src.utils.slides.slide20_charts.close_figure"):
                    _plot_simple_bars(
                        xlabels=["4T24", "3T25", "4T25"],
                        values=[4.761, 4.829, 5.230],
                        output_path=output_path,
                        delta_pairs=SLIDE20_DELTA_PAIRS,
                        delta_bracket_colors=SLIDE20_DELTA_BRACKET_COLORS,
                        delta_label_x_fractions=SLIDE20_DELTA_LABEL_X_FRACTIONS,
                    )

        bracket_lines = [line for line in ax.lines if len(line.get_xdata()) == 4]
        self.assertEqual(len(bracket_lines), 2)
        self.assertEqual(bracket_lines[0].get_color(), SLIDE20_DELTA_BRACKET_COLORS[0])
        self.assertEqual(bracket_lines[1].get_color(), SLIDE20_DELTA_BRACKET_COLORS[1])
        self.assertAlmostEqual(bracket_lines[0].get_xdata()[0], 0.0, places=2)
        self.assertAlmostEqual(bracket_lines[0].get_xdata()[2], 2.0, places=2)
        self.assertAlmostEqual(bracket_lines[1].get_xdata()[0], 1.0, places=2)
        self.assertAlmostEqual(bracket_lines[1].get_xdata()[2], 2.0, places=2)

        bracket_texts = [text for text in ax.texts if text.get_text().startswith(("+", "-"))]
        self.assertEqual(len(bracket_texts), 2)
        self.assertAlmostEqual(bracket_texts[0].get_position()[0], 0.60, places=2)
        self.assertAlmostEqual(bracket_texts[1].get_position()[0], 1.50, places=2)

        plt.close(fig)


if __name__ == "__main__":
    unittest.main()
