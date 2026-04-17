import tempfile
import unittest
from pathlib import Path
from unittest.mock import patch

import matplotlib.pyplot as plt
import numpy as np
from openpyxl import Workbook

from src.utils.slides.slide21_charts import (
    SLIDE21_COMPARATIVE_OUTPUT,
    SLIDE21_DELTA_BRACKET_COLORS,
    SLIDE21_DELTA_LABEL_X_FRACTIONS,
    SLIDE21_DELTA_PAIRS,
    SLIDE21_STACKED_OUTPUT,
    _fmt_comparative_value,
    _plot_stacked_atacado,
    generate_slide21_charts,
)


class TestSlide21Charts(unittest.TestCase):
    def test_fmt_comparative_value_forces_integer_rounding(self):
        self.assertEqual(_fmt_comparative_value(12.4), "12%")
        self.assertEqual(_fmt_comparative_value(12.5), "13%")
        self.assertEqual(_fmt_comparative_value(12.6), "13%")

    def test_generate_slide21_charts_uses_expected_ranges(self):
        wb = Workbook()
        ws_carteira = wb.active
        ws_carteira.title = "Carteira"
        ws_setor = wb.create_sheet("Carteira Atac Setor")

        for col_offset, label in enumerate(["4T24", "3T25", "4T25"], start=4):
            ws_carteira.cell(row=115, column=col_offset).value = label

        ws_carteira["C117"] = "corporate"
        ws_carteira["C118"] = "large + if"
        ws_carteira["C119"] = "pme"

        for col_offset, value in enumerate([14043.0, 14016.0, 15250.0], start=4):
            ws_carteira.cell(row=117, column=col_offset).value = value
        for col_offset, value in enumerate([11864.0, 9844.0, 9327.0], start=4):
            ws_carteira.cell(row=118, column=col_offset).value = value
        for col_offset, value in enumerate([2948.0, 2805.0, 3746.0], start=4):
            ws_carteira.cell(row=119, column=col_offset).value = value

        ws_setor["C2"] = "4T24"
        ws_setor["E2"] = "4T25"
        for row, (sector, left, right) in enumerate(
            [
                ("Agroindustria", 0.12, 0.12),
                ("PMEs", 0.10, 0.13),
                ("Outros", 0.13, 0.18),
            ],
            start=4,
        ):
            ws_setor[f"B{row}"] = sector
            ws_setor[f"D{row}"] = left
            ws_setor[f"F{row}"] = right

        captured: dict[str, object] = {}

        def _capture_stacked(*, xlabels, series_names, values, output_path, **kwargs):
            captured["stacked_xlabels"] = list(xlabels)
            captured["stacked_series_names"] = list(series_names)
            captured["stacked_values"] = values.tolist()
            captured["stacked_output_name"] = Path(output_path).name
            captured["stacked_kwargs"] = dict(kwargs)

        def _capture_comparative(*, categories, label_left, label_right, values_left, values_right, output_path):
            captured["comparative_categories"] = list(categories)
            captured["comparative_label_left"] = label_left
            captured["comparative_label_right"] = label_right
            captured["comparative_values_left"] = list(values_left)
            captured["comparative_values_right"] = list(values_right)
            captured["comparative_output_name"] = Path(output_path).name

        with tempfile.TemporaryDirectory() as td:
            xlsx_path = Path(td) / "test.xlsx"
            output_dir = Path(td) / "out"
            wb.save(xlsx_path)

            with patch("src.utils.slides.slide21_charts._plot_stacked_atacado", side_effect=_capture_stacked):
                with patch("src.utils.slides.slide21_charts._plot_comparative_bars", side_effect=_capture_comparative):
                    files = generate_slide21_charts(xlsx_path=xlsx_path, output_dir=output_dir)

        self.assertEqual([path.name for path in files], [SLIDE21_STACKED_OUTPUT, SLIDE21_COMPARATIVE_OUTPUT])
        self.assertEqual(captured["stacked_output_name"], SLIDE21_STACKED_OUTPUT)
        self.assertEqual(
            captured["stacked_kwargs"],
            {
                "delta_pairs": SLIDE21_DELTA_PAIRS,
                "delta_bracket_colors": SLIDE21_DELTA_BRACKET_COLORS,
                "delta_label_x_fractions": SLIDE21_DELTA_LABEL_X_FRACTIONS,
            },
        )
        self.assertEqual(captured["stacked_xlabels"], ["4T24", "3T25", "4T25"])
        self.assertEqual(captured["stacked_series_names"], ["Corporate", "Large + IF", "PMEs"])
        np.testing.assert_allclose(
            captured["stacked_values"],
            [
                [14.043, 11.864, 2.948],
                [14.016, 9.844, 2.805],
                [15.250, 9.327, 3.746],
            ],
            rtol=0.0,
            atol=1e-9,
        )

        self.assertEqual(captured["comparative_output_name"], SLIDE21_COMPARATIVE_OUTPUT)
        self.assertEqual(captured["comparative_categories"], ["Agroindustria", "PMEs", "Outros"])
        self.assertEqual(captured["comparative_label_left"], "4T24")
        self.assertEqual(captured["comparative_label_right"], "4T25")
        np.testing.assert_allclose(captured["comparative_values_left"], [12.0, 10.0, 13.0], rtol=0.0, atol=1e-9)
        np.testing.assert_allclose(captured["comparative_values_right"], [12.0, 13.0, 18.0], rtol=0.0, atol=1e-9)

    def test_plot_stacked_atacado_uses_slide15_bracket_style(self):
        fig, ax = plt.subplots()

        with tempfile.TemporaryDirectory() as td:
            output_path = Path(td) / "chart.png"
            with patch("matplotlib.pyplot.subplots", return_value=(fig, ax)):
                with patch("src.utils.slides.slide21_charts.close_figure"):
                    _plot_stacked_atacado(
                        xlabels=["4T24", "3T25", "4T25"],
                        series_names=["Corporate", "Large + IF", "PMEs"],
                        values=np.asarray(
                            [
                                [14.043, 11.864, 2.948],
                                [14.016, 9.844, 2.805],
                                [15.250, 9.327, 3.746],
                            ],
                            dtype=float,
                        ),
                        output_path=output_path,
                        delta_pairs=SLIDE21_DELTA_PAIRS,
                        delta_bracket_colors=SLIDE21_DELTA_BRACKET_COLORS,
                        delta_label_x_fractions=SLIDE21_DELTA_LABEL_X_FRACTIONS,
                    )

        bracket_lines = [line for line in ax.lines if len(line.get_xdata()) == 4]
        self.assertEqual(len(bracket_lines), 2)
        self.assertEqual(bracket_lines[0].get_color(), SLIDE21_DELTA_BRACKET_COLORS[0])
        self.assertEqual(bracket_lines[1].get_color(), SLIDE21_DELTA_BRACKET_COLORS[1])
        self.assertAlmostEqual(bracket_lines[0].get_xdata()[0], 0.0, places=2)
        self.assertAlmostEqual(bracket_lines[0].get_xdata()[2], 0.48, places=2)
        self.assertAlmostEqual(bracket_lines[1].get_xdata()[0], 0.24, places=2)
        self.assertAlmostEqual(bracket_lines[1].get_xdata()[2], 0.48, places=2)

        bracket_texts = [text for text in ax.texts if text.get_text().startswith(("+", "-"))]
        self.assertEqual(len(bracket_texts), 2)
        self.assertAlmostEqual(bracket_texts[0].get_position()[0], 0.14, places=2)
        self.assertAlmostEqual(bracket_texts[1].get_position()[0], 0.36, places=2)

        plt.close(fig)


if __name__ == "__main__":
    unittest.main()
