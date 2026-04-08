import tempfile
import unittest
from pathlib import Path
from unittest.mock import patch

import numpy as np
from openpyxl import Workbook

from utils.slide22_charts import (
    SLIDE22_CHARTS,
    SLIDE22_COBERTURA_REESTRUTURADA_LINE_CHART,
    SLIDE22_LINE_CHARTS,
    SLIDE22_NPL_BAR_CHART,
    SLIDE22_NPL_LINE_CHART,
    SLIDE22_REESTRUTURADA_BAR_CHART,
    SLIDE22_REESTRUTURADA_LINE_CHART,
    _callout_indices_for_bar,
    _fmt_pct_int,
    generate_slide22_charts,
)


class TestSlide22Charts(unittest.TestCase):
    def test_fmt_pct_int_rounds_to_integer_percent(self):
        self.assertEqual(_fmt_pct_int(32.4), "32%")
        self.assertEqual(_fmt_pct_int(32.5), "33%")
        self.assertEqual(_fmt_pct_int(74.1), "74%")

    def test_callout_indices_for_small_overlapping_segments(self):
        self.assertEqual(_callout_indices_for_bar(np.asarray([3.1, 93.9, 3.0], dtype=float)), {2})
        self.assertEqual(_callout_indices_for_bar(np.asarray([6.5, 85.0, 8.8], dtype=float)), {0})

    def test_generate_slide22_charts_uses_expected_ranges(self):
        wb = Workbook()
        ws = wb.active
        ws.title = "Qualidade Cart 4966"

        for col_offset, label in enumerate(["3T25", "4T25"], start=5):
            ws.cell(row=30, column=col_offset).value = label
            ws.cell(row=61, column=col_offset).value = label
            ws.cell(row=2, column=col_offset).value = label

        block_values = {
            32: [0.065, 0.058],
            33: [0.850, 0.855],
            34: [0.088, 0.087],
            42: [0.070, 0.065],
            43: [0.828, 0.831],
            44: [0.103, 0.104],
            52: [0.031, 0.029],
            53: [0.939, 0.948],
            54: [0.030, 0.022],
            37: [0.320, 0.326],
            38: [0.740, 0.760],
            47: [0.340, 0.352],
            48: [0.740, 0.756],
            57: [0.140, 0.102],
            58: [0.860, 0.741],
            64: [515299.0, 533967.0],
            65: [0.006, 0.006],
            67: [0.821, 0.872],
            19: [826.0, 457.0],
            20: [0.010, 0.005],
        }
        for row, values in block_values.items():
            for col_offset, value in enumerate(values, start=5):
                ws.cell(row=row, column=col_offset).value = value

        captured_calls: list[dict[str, object]] = []
        captured_line_calls: list[dict[str, object]] = []
        captured_bar_calls: list[dict[str, object]] = []

        def _capture_plot(*, xlabels, values, output_path):
            captured_calls.append(
                {
                    "xlabels": list(xlabels),
                    "values": values.tolist(),
                    "output_name": Path(output_path).name,
                }
            )

        def _capture_line_plot(
            *,
            xlabels,
            values,
            output_path,
            label_decimals=0,
            label_font_scale=1.0,
            x_tick_font_scale=1.0,
            label_offset_factor=0.06,
            label_offset_min=1.6,
            y_pad_factor=0.35,
            y_pad_min=3.5,
            show_xlabels=True,
        ):
            captured_line_calls.append(
                {
                    "xlabels": list(xlabels),
                    "values": values.tolist(),
                    "output_name": Path(output_path).name,
                    "label_decimals": label_decimals,
                    "label_font_scale": label_font_scale,
                    "x_tick_font_scale": x_tick_font_scale,
                    "label_offset_factor": label_offset_factor,
                    "label_offset_min": label_offset_min,
                    "y_pad_factor": y_pad_factor,
                    "y_pad_min": y_pad_min,
                    "show_xlabels": show_xlabels,
                }
            )

        def _capture_bar_plot(
            *,
            xlabels,
            values,
            output_path,
            bar_width=0.58,
            bar_slot=1.0,
            font_scale=1.0,
            **kwargs,
        ):
            captured_bar_calls.append(
                {
                    "xlabels": list(xlabels),
                    "values": list(values),
                    "output_name": Path(output_path).name,
                    "bar_width": bar_width,
                    "bar_slot": bar_slot,
                    "font_scale": font_scale,
                    **kwargs,
                }
            )

        with tempfile.TemporaryDirectory() as td:
            xlsx_path = Path(td) / "test.xlsx"
            output_dir = Path(td) / "out"
            wb.save(xlsx_path)

            with patch("utils.slide22_charts._plot_stacked_pct_block", side_effect=_capture_plot):
                with patch("utils.slide22_charts._plot_two_point_lines", side_effect=_capture_line_plot):
                    with patch("utils.slide22_charts._plot_two_bar_values", side_effect=_capture_bar_plot):
                        files = generate_slide22_charts(xlsx_path=xlsx_path, output_dir=output_dir)

        self.assertEqual(
            [path.name for path in files],
            [spec.output_name for spec in SLIDE22_CHARTS]
            + [spec.output_name for spec in SLIDE22_LINE_CHARTS]
            + [
                SLIDE22_REESTRUTURADA_BAR_CHART.output_name,
                SLIDE22_REESTRUTURADA_LINE_CHART.output_name,
                SLIDE22_COBERTURA_REESTRUTURADA_LINE_CHART.output_name,
                SLIDE22_NPL_BAR_CHART.output_name,
                SLIDE22_NPL_LINE_CHART.output_name,
            ],
        )
        self.assertEqual([call["xlabels"] for call in captured_calls], [["3T25", "4T25"]] * 3)
        self.assertEqual(
            [call["output_name"] for call in captured_calls],
            [
                "22_qualidade_4966_bloco1.png",
                "22_qualidade_4966_bloco2.png",
                "22_qualidade_4966_bloco3.png",
            ],
        )
        np.testing.assert_allclose(
            captured_calls[0]["values"],
            [[6.5, 85.0, 8.8], [5.8, 85.5, 8.7]],
            rtol=0.0,
            atol=1e-9,
        )
        self.assertEqual([call["xlabels"] for call in captured_line_calls], [["3T25", "4T25"]] * 6)
        self.assertEqual(
            [call["output_name"] for call in captured_line_calls[:3]],
            [
                "22_qualidade_4966_linha1.png",
                "22_qualidade_4966_linha2.png",
                "22_qualidade_4966_linha3.png",
            ],
        )
        self.assertEqual([call["label_decimals"] for call in captured_line_calls[:3]], [0, 0, 0])
        np.testing.assert_allclose(
            captured_line_calls[0]["values"],
            [[32.0, 32.6], [74.0, 76.0]],
            rtol=0.0,
            atol=1e-9,
        )
        np.testing.assert_allclose(
            captured_line_calls[1]["values"],
            [[34.0, 35.2], [74.0, 75.6]],
            rtol=0.0,
            atol=1e-9,
        )
        np.testing.assert_allclose(
            captured_line_calls[2]["values"],
            [[14.0, 10.2], [86.0, 74.1]],
            rtol=0.0,
            atol=1e-9,
        )
        self.assertEqual(
            captured_bar_calls,
            [
                {
                    "xlabels": ["3T25", "4T25"],
                    "values": [515.299, 533.967],
                    "output_name": "22_carteira_reestruturada_barras.png",
                    "bar_width": 0.066,
                    "bar_slot": 0.096,
                    "font_scale": 1.5,
                    "bracket_anchor": "center",
                    "bracket_top_gap_scale": 0.18,
                    "bracket_top_gap_min": 55.0,
                    "bracket_label_clearance": 24.0,
                    "x_margin": 0.12,
                },
                {
                    "xlabels": ["3T25", "4T25"],
                    "values": [826.0, 457.0],
                    "output_name": "22_npl_barras.png",
                    "bar_width": 0.58,
                    "bar_slot": 1.0,
                    "font_scale": 1.0,
                    "bracket_top_gap_scale": 0.12,
                    "bracket_top_gap_min": 36.0,
                    "bracket_label_clearance": 18.0,
                }
            ],
        )
        self.assertEqual(
            [call["output_name"] for call in captured_line_calls[-3:]],
            [
                "22_carteira_reestruturada_linha.png",
                "22_cobertura_reestruturada_linha.png",
                "22_npl_linha.png",
            ],
        )
        self.assertEqual([call["label_decimals"] for call in captured_line_calls[-3:]], [1, 1, 1])
        self.assertEqual(captured_line_calls[-3]["label_font_scale"], 2.24)
        self.assertEqual(captured_line_calls[-3]["label_offset_factor"], 0.008)
        self.assertEqual(captured_line_calls[-3]["label_offset_min"], 0.05)
        self.assertEqual(captured_line_calls[-3]["y_pad_factor"], 0.18)
        self.assertEqual(captured_line_calls[-3]["y_pad_min"], 0.28)
        self.assertEqual(captured_line_calls[-3]["show_xlabels"], False)
        np.testing.assert_allclose(
            captured_line_calls[-3]["values"],
            [[0.6, 0.6]],
            rtol=0.0,
            atol=1e-9,
        )
        np.testing.assert_allclose(
            captured_line_calls[-2]["values"],
            [[82.1, 87.2]],
            rtol=0.0,
            atol=1e-9,
        )
        np.testing.assert_allclose(
            captured_line_calls[-1]["values"],
            [[1.0, 0.5]],
            rtol=0.0,
            atol=1e-9,
        )
        self.assertEqual(captured_line_calls[-1]["label_font_scale"], 2.24)
        self.assertEqual(captured_line_calls[-1]["label_offset_factor"], 0.008)
        self.assertEqual(captured_line_calls[-1]["label_offset_min"], 0.05)
        self.assertEqual(captured_line_calls[-1]["show_xlabels"], False)
        np.testing.assert_allclose(
            captured_calls[1]["values"],
            [[7.0, 82.8, 10.3], [6.5, 83.1, 10.4]],
            rtol=0.0,
            atol=1e-9,
        )
        np.testing.assert_allclose(
            captured_calls[2]["values"],
            [[3.1, 93.9, 3.0], [2.9, 94.8, 2.2]],
            rtol=0.0,
            atol=1e-9,
        )


if __name__ == "__main__":
    unittest.main()
