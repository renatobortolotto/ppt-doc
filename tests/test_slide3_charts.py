import tempfile
import unittest
from pathlib import Path

from openpyxl import Workbook

from utils.slide3_charts import _combine_consignado_demais, generate_slide3_charts


class TestSlide3Charts(unittest.TestCase):
    def test_combine_consignado_demais_merges_matching_rows(self):
        rows = [
            ("Demais", [10.0, 20.0]),
            ("Consignado Privado", [1.0, 2.0]),
            ("Veiculos", [3.0, 4.0]),
        ]

        combined = _combine_consignado_demais(rows)

        self.assertEqual(
            combined,
            [("Veiculos", [3.0, 4.0]), ("Demais", [11.0, 22.0])],
        )

    def test_generate_slide3_charts_creates_expected_pngs(self):
        wb = Workbook()
        ws_emprestimos = wb.active
        ws_emprestimos.title = "Emprestimos"
        ws_seguros = wb.create_sheet("Seguros e Cartoes")

        for col_offset, label in enumerate(["3T25", "4T25", "1T26"], start=4):
            ws_emprestimos.cell(row=4, column=col_offset).value = label
            ws_seguros.cell(row=14, column=col_offset).value = label

        rows = {
            5: ("Demais", [1000, 1100, 1200]),
            6: ("Consignado Privado", [200, 210, 220]),
            7: ("Veiculos", [300, 320, 340]),
            8: ("Motos", [400, 420, 440]),
            9: ("Total", [1900, 2050, 2200]),
        }
        for row, (label, values) in rows.items():
            ws_emprestimos.cell(row=row, column=3).value = label
            for col_offset, value in enumerate(values, start=4):
                ws_emprestimos.cell(row=row, column=col_offset).value = value

        for col_offset, value in enumerate([5200, 5600, 6100], start=4):
            ws_seguros.cell(row=15, column=col_offset).value = value

        with tempfile.TemporaryDirectory() as td:
            xlsx_path = Path(td) / "test.xlsx"
            output_dir = Path(td) / "out"
            wb.save(xlsx_path)

            files = generate_slide3_charts(xlsx_path=xlsx_path, output_dir=output_dir)
            self.assertEqual(
                [path.name for path in files],
                ["08_emprestimos_empilhado.png", "09_seguros_cartoes_total.png"],
            )
            for file_path in files:
                self.assertTrue(file_path.exists())
                self.assertGreater(file_path.stat().st_size, 0)


if __name__ == "__main__":
    unittest.main()
