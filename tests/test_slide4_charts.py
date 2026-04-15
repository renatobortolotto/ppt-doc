import tempfile
import unittest
from pathlib import Path
from unittest.mock import patch

from openpyxl import Workbook

from src.utils.slides.slide4_charts import (
    SLIDE4_BAR_FONT_SCALE,
    SLIDE4_BAR_GAP_SCALE,
    SLIDE4_BAR_WIDTH_SCALE,
    SLIDE4_9M_DELTA_OFFSET_SCALE,
    SLIDE4_TRIMESTRES_DELTA_BRACKET_COLORS,
    SLIDE4_TRIMESTRES_DELTA_OFFSET_SCALE,
    _extract_slide4_donut_series,
    generate_slide4_charts,
)


class TestSlide4Charts(unittest.TestCase):
    def test_extract_slide4_donut_series_groups_by_fixed_rows_even_with_unexpected_labels(self):
        wb = Workbook()
        ws = wb.active
        ws.title = "Carteira"

        rows = {
            16: ("qualquer texto 1", 100.0),
            17: ("qualquer texto 2", 50.0),
            18: ("qualquer texto 3", 70.0),
            19: ("qualquer texto 4", 80.0),
            23: ("qualquer texto 5", 110.0),
            24: ("qualquer texto 6", 90.0),
            25: ("qualquer texto 7", 10.0),
            28: ("qualquer texto 8", 200.0),
            29: ("qualquer texto 9", 300.0),
            30: ("qualquer texto 10", 400.0),
            35: ("qualquer texto 11", 20.0),
            36: ("qualquer texto 12", 30.0),
        }

        for row, (label, value) in rows.items():
            ws.cell(row=row, column=3).value = label
            ws.cell(row=row, column=6).value = value

        categories, labels, values = _extract_slide4_donut_series(ws, source_range="C12:F36")

        self.assertEqual(
            categories,
            [
                "Veiculos Leves",
                "Growth",
                "Growth",
                "Growth",
                "Growth",
                "Atacado",
                "Atacado",
                "Atacado",
            ],
        )
        self.assertEqual(
            labels,
            [
                "Veiculos Leves Usados",
                "Outros Veiculos",
                "Paineis Solares",
                "EGV",
                "Cartões",
                "Corporate",
                "Large Corporate + instituicoes financeiras",
                "Pequenas e Medias Empresas (PME)",
            ],
        )
        self.assertEqual(values, [100.0, 120.0, 80.0, 110.0, 100.0, 220.0, 330.0, 400.0])

    def test_extract_slide4_donut_series_reads_carteira_leaf_rows_and_skips_subtotals(self):
        wb = Workbook()
        ws = wb.active
        ws.title = "Carteira"

        rows = {
            14: ("segmento varejo", 69333),
            15: ("Varejo - Produtos de Entrada", 58802),
            16: ("Veiculos Leves Usados", 46888),
            17: ("Veiculos Pesados", 3258),
            18: ("Motos e Veiculos Novos", 4558),
            19: ("Paineis Solares", 3731),
            20: ("Consignado Privado", 333),
            21: ("Outros (FGTS + Saúde)", 35),
            22: ("Varejo Relacional", 10531),
            23: ("Empréstimo com Garantia Vericular (EGV)", 5262),
            24: ("Cartão de Crédito", 5230),
            25: ("CP", 39),
            27: ("Segmento Atacado", 10318),
            28: ("Corporate", 6547),
            29: ("Large Corporate + Instituições Financeiras", 2507),
            30: ("Pequenas e Médias Empresas", 1265),
            34: ("Segmento Atacado Ampliado", 28323),
            35: ("Avais e Fianças Prestados", 6120),
            36: ("TVM privado", 11885),
        }

        for row, (label, value) in rows.items():
            ws.cell(row=row, column=3).value = label
            ws.cell(row=row, column=6).value = value

        categories, labels, values = _extract_slide4_donut_series(ws, source_range="C12:F36")

        self.assertEqual(
            categories,
            [
                "Veiculos Leves",
                "Growth",
                "Growth",
                "Growth",
                "Growth",
                "Atacado",
                "Atacado",
                "Atacado",
            ],
        )
        self.assertEqual(
            labels,
            [
                "Veiculos Leves Usados",
                "Outros Veiculos",
                "Paineis Solares",
                "EGV",
                "Cartões",
                "Corporate",
                "Large Corporate + instituicoes financeiras",
                "Pequenas e Medias Empresas (PME)",
            ],
        )
        self.assertEqual(
            values,
            [46888.0, 7816.0, 3731.0, 5262.0, 5269.0, 12667.0, 14392.0, 1265.0],
        )

    def test_generate_slide4_charts_creates_expected_pngs(self):
        wb = Workbook()
        ws = wb.active
        ws.title = "Carteira"
        ws_pizza = wb.create_sheet("Pizza Teste")

        rows = {
            16: ("Veiculos Leves Usados", 46888),
            17: ("Veiculos Pesados", 3258),
            18: ("Motos e Veiculos Novos", 4558),
            19: ("Paineis Solares", 3731),
            23: ("Empréstimo com Garantia Vericular (EGV)", 5262),
            24: ("Cartão de Crédito", 5230),
            25: ("CP", 39),
            28: ("Corporate", 6547),
            29: ("Large Corporate + Instituições Financeiras", 2507),
            30: ("Pequenas e Médias Empresas", 1265),
            35: ("Avais e Fianças Prestados", 6120),
            36: ("TVM privado", 11885),
        }
        for row, (label, value) in rows.items():
            ws.cell(row=row, column=3).value = label
            ws.cell(row=row, column=6).value = value

        for col_offset, label in enumerate(["3T25", "4T25", "1T26"], start=8):
            ws_pizza.cell(row=3, column=col_offset).value = label
        for col_offset, value in enumerate([120, 135, 150], start=8):
            ws_pizza.cell(row=4, column=col_offset).value = value

        for col_offset, label in enumerate(["9M25", "9M26"], start=11):
            ws_pizza.cell(row=3, column=col_offset).value = label
        for col_offset, value in enumerate([390, 420], start=11):
            ws_pizza.cell(row=4, column=col_offset).value = value

        with tempfile.TemporaryDirectory() as td:
            xlsx_path = Path(td) / "test.xlsx"
            output_dir = Path(td) / "out"
            wb.save(xlsx_path)

            files = generate_slide4_charts(xlsx_path=xlsx_path, output_dir=output_dir)
            self.assertEqual(
                [path.name for path in files],
                ["10_pizza_carteira.png", "11_pizza_trimestres.png", "12_pizza_9m.png"],
            )
            for file_path in files:
                self.assertTrue(file_path.exists())
                self.assertGreater(file_path.stat().st_size, 0)

    @patch("src.utils.slides.slide4_charts.close_figure")
    @patch("src.utils.slides.slide4_charts.plot_bar_from_excel")
    @patch("src.utils.slides.slide4_charts.plot_donut_chart")
    def test_generate_slide4_charts_uses_updated_bar_specs(
        self,
        plot_donut_chart_mock,
        plot_bar_from_excel_mock,
        close_figure_mock,
    ):
        wb = Workbook()
        ws = wb.active
        ws.title = "Carteira"
        ws_pizza = wb.create_sheet("Pizza Teste")

        for row, value in {
            16: 46888,
            17: 3258,
            18: 4558,
            19: 3731,
            23: 5262,
            24: 5230,
            25: 39,
            28: 6547,
            29: 2507,
            30: 1265,
            35: 6120,
            36: 11885,
        }.items():
            ws.cell(row=row, column=6).value = value

        for col_offset, label in enumerate(["3T25", "4T25", "1T26"], start=8):
            ws_pizza.cell(row=3, column=col_offset).value = label
        for col_offset, value in enumerate([120, 135, 150], start=8):
            ws_pizza.cell(row=4, column=col_offset).value = value

        for col_offset, label in enumerate(["9M25", "9M26"], start=11):
            ws_pizza.cell(row=3, column=col_offset).value = label
        for col_offset, value in enumerate([390, 420], start=11):
            ws_pizza.cell(row=4, column=col_offset).value = value

        dummy_fig = object()
        dummy_ax = object()
        plot_donut_chart_mock.return_value = (dummy_fig, dummy_ax)
        plot_bar_from_excel_mock.side_effect = [(dummy_fig, dummy_ax), (dummy_fig, dummy_ax)]

        with tempfile.TemporaryDirectory() as td:
            xlsx_path = Path(td) / "test.xlsx"
            output_dir = Path(td) / "out"
            wb.save(xlsx_path)

            generate_slide4_charts(xlsx_path=xlsx_path, output_dir=output_dir)

        self.assertEqual(plot_bar_from_excel_mock.call_count, 2)

        trimestres_spec = plot_bar_from_excel_mock.call_args_list[0].args[0]
        self.assertEqual(trimestres_spec.font_scale, SLIDE4_BAR_FONT_SCALE)
        self.assertEqual(trimestres_spec.delta_pairs, ((0, 2), (1, 2)))
        self.assertEqual(trimestres_spec.delta_bracket_colors, SLIDE4_TRIMESTRES_DELTA_BRACKET_COLORS)
        self.assertEqual(trimestres_spec.delta_label_x_fractions, (0.30, 0.50))
        self.assertEqual(trimestres_spec.bar_width_scale, SLIDE4_BAR_WIDTH_SCALE)
        self.assertEqual(trimestres_spec.gap_scale, SLIDE4_BAR_GAP_SCALE)
        self.assertEqual(trimestres_spec.delta_offset_scale, SLIDE4_TRIMESTRES_DELTA_OFFSET_SCALE)

        nove_meses_spec = plot_bar_from_excel_mock.call_args_list[1].args[0]
        self.assertEqual(nove_meses_spec.font_scale, SLIDE4_BAR_FONT_SCALE)
        self.assertEqual(nove_meses_spec.delta_pairs, ((0, 1),))
        self.assertEqual(nove_meses_spec.bar_width_scale, SLIDE4_BAR_WIDTH_SCALE)
        self.assertEqual(nove_meses_spec.gap_scale, SLIDE4_BAR_GAP_SCALE)
        self.assertEqual(nove_meses_spec.delta_offset_scale, SLIDE4_9M_DELTA_OFFSET_SCALE)

        self.assertEqual(close_figure_mock.call_count, 3)

if __name__ == "__main__":
    unittest.main()
