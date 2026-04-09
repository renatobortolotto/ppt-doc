import tempfile
import unittest
from pathlib import Path

from openpyxl import Workbook

from utils.slide1_charts import generate_slide1_charts
from utils.slide7_charts import generate_slide7_charts


class TestSlide7Charts(unittest.TestCase):
    def test_legacy_slide1_alias_points_to_slide7_generator(self):
        self.assertIs(generate_slide1_charts, generate_slide7_charts)

    def test_generate_slide7_charts_creates_expected_pngs(self):
        wb = Workbook()
        ws = wb.active
        ws.title = "DRE Saida"

        quarter_labels = ["1T24", "2T24", "3T24", "4T24", "1T25", "2T25", "3T25", "4T25", "1T26"]
        for col_offset, label in enumerate(quarter_labels, start=3):
            ws.cell(row=3, column=col_offset).value = label
        for col_offset, value in enumerate([100, 110, 120, 130, 140, 150, 160, 170, 180], start=3):
            ws.cell(row=18, column=col_offset).value = value
        for col_offset, value in enumerate([0.10, 0.11, 0.115, 0.12, 0.125, 0.13, 0.14, 0.145, 0.15], start=3):
            ws.cell(row=20, column=col_offset).value = value

        for col_offset, label in enumerate(["9M25", "9M26"], start=12):
            ws.cell(row=3, column=col_offset).value = label
        for col_offset, value in enumerate([420, 510], start=12):
            ws.cell(row=18, column=col_offset).value = value
        for col_offset, value in enumerate([0.135, 0.148], start=12):
            ws.cell(row=20, column=col_offset).value = value

        with tempfile.TemporaryDirectory() as td:
            xlsx_path = Path(td) / "test.xlsx"
            output_dir = Path(td) / "out"
            wb.save(xlsx_path)

            files = generate_slide7_charts(xlsx_path=xlsx_path, output_dir=output_dir)
            self.assertEqual(
                [path.name for path in files],
                [
                    "01_lucro_trimestres.png",
                    "02_lucro_9m.png",
                    "03_roe_trimestres.png",
                    "04_roe_9m.png",
                ],
            )
            for file_path in files:
                self.assertTrue(file_path.exists())
                self.assertGreater(file_path.stat().st_size, 0)


if __name__ == "__main__":
    unittest.main()
