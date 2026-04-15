import tempfile
import unittest
from pathlib import Path
from unittest.mock import patch

import numpy as np
from matplotlib.colors import to_rgba
from openpyxl import Workbook

from src.utils.slides.slide8_charts import (
    SLIDE8_9M_BAR_WIDTH_SCALE,
    SLIDE8_9M_GAP_SCALE,
    SLIDE8_FONT_SCALE,
    SLIDE8_LEGEND_MAX_LINE_LEN,
    SLIDE8_MFB_TRIMESTRES_LABEL_BBOX_SERIES,
    SLIDE8_MFB_TRIMESTRES_X_TICK_PAD,
    SLIDE8_TRIMESTRES_BAR_WIDTH_SCALE,
    SLIDE8_TRIMESTRES_DELTA_BRACKET_COLORS,
    SLIDE8_TRIMESTRES_DELTA_LABEL_X_FRACTIONS,
    SLIDE8_TRIMESTRES_DELTA_PAIRS,
    SLIDE8_TRIMESTRES_GAP_SCALE,
    _plot_stacked_vertical,
    _read_stacked_rows,
    _wrap_words,
    generate_slide8_charts,
)


class TestSlide8Charts(unittest.TestCase):
    def test_read_stacked_rows_reads_named_series_from_linear_ranges(self):
        wb = Workbook()
        ws = wb.active
        ws.title = "DRE Saida 2"

        labels = ["3T24", "2T25", "3T25"]
        margem_values = [2372, 2311, 2295]
        servicos_values = [685, 556, 624]

        for col_offset, label in enumerate(labels, start=4):
            ws.cell(row=3, column=col_offset).value = label

        for col_offset, value in enumerate(margem_values, start=4):
            ws.cell(row=5, column=col_offset).value = value

        for col_offset, value in enumerate(servicos_values, start=4):
            ws.cell(row=9, column=col_offset).value = value

        xlabels, series, values = _read_stacked_rows(
            ws,
            xlabels_range="D3:F3",
            series_specs=[
                ("Margem Financeira Bruta", "D5:F5"),
                ("Serviços e Seguros", "D9:F9"),
            ],
        )

        self.assertEqual(xlabels, ["3T24", "2T25", "3T25"])
        self.assertEqual(series, ["Margem Financeira Bruta", "Serviços e Seguros"])
        self.assertEqual(values.tolist(), [[2372.0, 685.0], [2311.0, 556.0], [2295.0, 624.0]])

    def test_wrap_words_breaks_long_label(self):
        self.assertEqual(
            _wrap_words("Serviços e Seguros", max_line_len=11),
            "Serviços e\nSeguros",
        )

    def test_generate_slide8_charts_creates_expected_pngs(self):
        wb = Workbook()
        ws_dre = wb.active
        ws_dre.title = "DRE Saida 2"
        ws_tabelas = wb.create_sheet("Tabelas")

        for col_offset, label in enumerate(["3T25", "4T25", "1T26"], start=4):
            ws_dre.cell(row=3, column=col_offset).value = label
        for col_offset, label in enumerate(["9M25", "9M26"], start=7):
            ws_dre.cell(row=3, column=col_offset).value = label

        for col_offset, value in enumerate([2372, 2311, 2295], start=4):
            ws_dre.cell(row=5, column=col_offset).value = value
        for col_offset, value in enumerate([685, 556, 624], start=4):
            ws_dre.cell(row=9, column=col_offset).value = value
        for col_offset, value in enumerate([1800, 1750, 1720], start=4):
            ws_dre.cell(row=6, column=col_offset).value = value
        for col_offset, value in enumerate([572, 561, 575], start=4):
            ws_dre.cell(row=7, column=col_offset).value = value

        for col_offset, value in enumerate([4683, 4606], start=7):
            ws_dre.cell(row=5, column=col_offset).value = value
        for col_offset, value in enumerate([1241, 1180], start=7):
            ws_dre.cell(row=9, column=col_offset).value = value
        for col_offset, value in enumerate([3520, 3440], start=7):
            ws_dre.cell(row=6, column=col_offset).value = value
        for col_offset, value in enumerate([1163, 1166], start=7):
            ws_dre.cell(row=7, column=col_offset).value = value

        for col_offset, label in enumerate(["3T25", "4T25", "1T26"], start=4):
            ws_tabelas.cell(row=16, column=col_offset).value = label
        for col_offset, label in enumerate(["9M25", "9M26"], start=7):
            ws_tabelas.cell(row=16, column=col_offset).value = label

        for col_offset, value in enumerate([110, 120, 130], start=4):
            ws_tabelas.cell(row=19, column=col_offset).value = value
        for col_offset, value in enumerate([75, 80, 90], start=4):
            ws_tabelas.cell(row=28, column=col_offset).value = value
        for col_offset, value in enumerate([330, 360], start=7):
            ws_tabelas.cell(row=19, column=col_offset).value = value
        for col_offset, value in enumerate([220, 240], start=7):
            ws_tabelas.cell(row=28, column=col_offset).value = value

        with tempfile.TemporaryDirectory() as td:
            xlsx_path = Path(td) / "test.xlsx"
            output_dir = Path(td) / "out"
            wb.save(xlsx_path)

            files = generate_slide8_charts(xlsx_path=xlsx_path, output_dir=output_dir)
            self.assertEqual(
                [path.name for path in files],
                [
                    "13_slide8_trimestres.png",
                    "14_slide8_9m.png",
                    "15_margem_financeira_bruta_total_trimestres.png",
                    "16_margem_financeira_bruta_total_9m.png",
                    "17_servicos_corretagem_trimestres.png",
                    "18_servicos_corretagem_9m.png",
                ],
            )
            for file_path in files:
                self.assertTrue(file_path.exists())
                self.assertGreater(file_path.stat().st_size, 0)

    @patch("src.utils.slides.slide8_charts._plot_stacked_vertical")
    def test_generate_slide8_charts_uses_updated_specs_for_trimestres(self, plot_mock):
        wb = Workbook()
        ws_dre = wb.active
        ws_dre.title = "DRE Saida 2"
        ws_tabelas = wb.create_sheet("Tabelas")

        for col_offset, label in enumerate(["3T25", "4T25", "1T26"], start=4):
            ws_dre.cell(row=3, column=col_offset).value = label
            ws_tabelas.cell(row=16, column=col_offset).value = label
        for col_offset, label in enumerate(["9M25", "9M26"], start=7):
            ws_dre.cell(row=3, column=col_offset).value = label
            ws_tabelas.cell(row=16, column=col_offset).value = label

        for col_offset, value in enumerate([2372, 2311, 2295], start=4):
            ws_dre.cell(row=5, column=col_offset).value = value
        for col_offset, value in enumerate([685, 556, 624], start=4):
            ws_dre.cell(row=9, column=col_offset).value = value
        for col_offset, value in enumerate([1800, 1750, 1720], start=4):
            ws_dre.cell(row=6, column=col_offset).value = value
        for col_offset, value in enumerate([572, 561, 575], start=4):
            ws_dre.cell(row=7, column=col_offset).value = value

        for col_offset, value in enumerate([4683, 4606], start=7):
            ws_dre.cell(row=5, column=col_offset).value = value
        for col_offset, value in enumerate([1241, 1180], start=7):
            ws_dre.cell(row=9, column=col_offset).value = value
        for col_offset, value in enumerate([3520, 3440], start=7):
            ws_dre.cell(row=6, column=col_offset).value = value
        for col_offset, value in enumerate([1163, 1166], start=7):
            ws_dre.cell(row=7, column=col_offset).value = value

        for col_offset, value in enumerate([110, 120, 130], start=4):
            ws_tabelas.cell(row=19, column=col_offset).value = value
        for col_offset, value in enumerate([75, 80, 90], start=4):
            ws_tabelas.cell(row=28, column=col_offset).value = value
        for col_offset, value in enumerate([330, 360], start=7):
            ws_tabelas.cell(row=19, column=col_offset).value = value
        for col_offset, value in enumerate([220, 240], start=7):
            ws_tabelas.cell(row=28, column=col_offset).value = value

        with tempfile.TemporaryDirectory() as td:
            xlsx_path = Path(td) / "test.xlsx"
            output_dir = Path(td) / "out"
            wb.save(xlsx_path)

            generate_slide8_charts(xlsx_path=xlsx_path, output_dir=output_dir)

        self.assertEqual(plot_mock.call_count, 6)

        for index in (0, 2, 4):
            kwargs = plot_mock.call_args_list[index].kwargs
            self.assertEqual(kwargs["font_scale"], SLIDE8_FONT_SCALE)
            self.assertEqual(kwargs["delta_pairs"], SLIDE8_TRIMESTRES_DELTA_PAIRS)
            self.assertEqual(kwargs["delta_bracket_colors"], SLIDE8_TRIMESTRES_DELTA_BRACKET_COLORS)
            self.assertEqual(kwargs["delta_label_x_fractions"], SLIDE8_TRIMESTRES_DELTA_LABEL_X_FRACTIONS)
            self.assertEqual(kwargs["bar_width_scale"], SLIDE8_TRIMESTRES_BAR_WIDTH_SCALE)
            self.assertEqual(kwargs["gap_scale"], SLIDE8_TRIMESTRES_GAP_SCALE)
            if index == 2:
                self.assertEqual(kwargs["x_tick_pad"], SLIDE8_MFB_TRIMESTRES_X_TICK_PAD)
                self.assertEqual(kwargs["segment_label_bbox_series_names"], SLIDE8_MFB_TRIMESTRES_LABEL_BBOX_SERIES)
            else:
                self.assertNotIn("x_tick_pad", kwargs)
                self.assertNotIn("segment_label_bbox_series_names", kwargs)

        for index in (1, 3, 5):
            kwargs = plot_mock.call_args_list[index].kwargs
            self.assertEqual(kwargs["font_scale"], SLIDE8_FONT_SCALE)
            self.assertEqual(kwargs["bar_width_scale"], SLIDE8_9M_BAR_WIDTH_SCALE)
            self.assertEqual(kwargs["gap_scale"], SLIDE8_9M_GAP_SCALE)
            self.assertNotIn("delta_pairs", kwargs)
            self.assertNotIn("delta_bracket_colors", kwargs)
            self.assertNotIn("delta_label_x_fractions", kwargs)

    def test_plot_stacked_vertical_respects_explicit_delta_pairs_and_colors(self):
        with tempfile.TemporaryDirectory() as td:
            output_path = Path(td) / "out.png"
            captured: dict[str, object] = {}

            import matplotlib.pyplot as plt

            original_subplots = plt.subplots

            def _capture_subplots(*args, **kwargs):
                fig, ax = original_subplots(*args, **kwargs)
                captured["fig"] = fig
                captured["ax"] = ax
                return fig, ax

            with (
                patch("matplotlib.pyplot.subplots", side_effect=_capture_subplots),
                patch("src.utils.slides.slide8_charts.close_figure"),
            ):
                _plot_stacked_vertical(
                    xlabels=["3T25", "4T25", "1T26"],
                    series_names=["Margem Financeira Bruta", "Serviços e Seguros"],
                    values=np.array([[70.0, 30.0], [80.0, 40.0], [90.0, 60.0]]),
                    output_path=output_path,
                    colors=["#123a7a", "#5B8FF9"],
                    font_scale=SLIDE8_FONT_SCALE,
                    show_delta_bracket=True,
                    show_delta_pct=True,
                    show_segment_pct=False,
                    bold_last_bar=True,
                    bold_text=True,
                    inline_left_legend=True,
                    delta_pairs=SLIDE8_TRIMESTRES_DELTA_PAIRS,
                    delta_bracket_colors=SLIDE8_TRIMESTRES_DELTA_BRACKET_COLORS,
                    delta_label_x_fractions=SLIDE8_TRIMESTRES_DELTA_LABEL_X_FRACTIONS,
                )

            ax = captured["ax"]
            delta_texts = [text for text in ax.texts if "%" in text.get_text()]
            self.assertEqual([text.get_text() for text in delta_texts], ["+50,0%", "+25,0%"])
            bracket_lines = [line for line in ax.lines if len(line.get_xdata()) == 4]
            self.assertEqual([line.get_color() for line in bracket_lines[:2]], list(SLIDE8_TRIMESTRES_DELTA_BRACKET_COLORS))

            expected_legend_texts = {
                _wrap_words("Margem Financeira Bruta", max_line_len=SLIDE8_LEGEND_MAX_LINE_LEN),
                _wrap_words("Serviços e Seguros", max_line_len=SLIDE8_LEGEND_MAX_LINE_LEN),
            }
            legend_texts = [text for text in ax.texts if text.get_text() in expected_legend_texts]
            self.assertEqual(len(legend_texts), 2)
            for legend_text in legend_texts:
                self.assertEqual(legend_text.get_ha(), "right")
                self.assertIn("\n", legend_text.get_text())

            bar_centers = [
                patch_obj.get_x() + patch_obj.get_width() / 2.0
                for patch_obj in ax.patches[:3]
            ]
            for legend_text in legend_texts:
                self.assertLess(legend_text.get_position()[0], bar_centers[0])
            first_midpoint = (bar_centers[0] + bar_centers[2]) / 2.0
            second_midpoint = (bar_centers[1] + bar_centers[2]) / 2.0
            self.assertLess(delta_texts[0].get_position()[0], first_midpoint)
            self.assertAlmostEqual(delta_texts[1].get_position()[0], second_midpoint, places=6)

            captured["fig"].clf()

    def test_plot_stacked_vertical_can_reduce_9m_bar_width_and_gap(self):
        with tempfile.TemporaryDirectory() as td:
            td_path = Path(td)
            default_output = td_path / "default.png"
            compact_output = td_path / "compact.png"
            captured_default: dict[str, object] = {}
            captured_compact: dict[str, object] = {}

            import matplotlib.pyplot as plt

            original_subplots = plt.subplots

            def _capture_default(*args, **kwargs):
                fig, ax = original_subplots(*args, **kwargs)
                captured_default["fig"] = fig
                captured_default["ax"] = ax
                return fig, ax

            def _capture_compact(*args, **kwargs):
                fig, ax = original_subplots(*args, **kwargs)
                captured_compact["fig"] = fig
                captured_compact["ax"] = ax
                return fig, ax

            with (
                patch("matplotlib.pyplot.subplots", side_effect=_capture_default),
                patch("src.utils.slides.slide8_charts.close_figure"),
            ):
                _plot_stacked_vertical(
                    xlabels=["9M25", "9M26"],
                    series_names=["Clientes", "Mercado"],
                    values=np.array([[70.0, 30.0], [90.0, 60.0]]),
                    output_path=default_output,
                    colors=["#123a7a", "#5B8FF9"],
                    show_delta_bracket=False,
                    show_delta_pct=False,
                    inline_left_legend=False,
                )

            with (
                patch("matplotlib.pyplot.subplots", side_effect=_capture_compact),
                patch("src.utils.slides.slide8_charts.close_figure"),
            ):
                _plot_stacked_vertical(
                    xlabels=["9M25", "9M26"],
                    series_names=["Clientes", "Mercado"],
                    values=np.array([[70.0, 30.0], [90.0, 60.0]]),
                    output_path=compact_output,
                    colors=["#123a7a", "#5B8FF9"],
                    show_delta_bracket=False,
                    show_delta_pct=False,
                    inline_left_legend=False,
                    bar_width_scale=SLIDE8_9M_BAR_WIDTH_SCALE,
                    gap_scale=SLIDE8_9M_GAP_SCALE,
                )

            default_patches = captured_default["ax"].patches
            compact_patches = captured_compact["ax"].patches

            default_width = default_patches[0].get_width()
            compact_width = compact_patches[0].get_width()
            self.assertAlmostEqual(compact_width, default_width * 0.5, places=6)

            default_gap = default_patches[1].get_x() - (default_patches[0].get_x() + default_patches[0].get_width())
            compact_gap = compact_patches[1].get_x() - (compact_patches[0].get_x() + compact_patches[0].get_width())
            self.assertAlmostEqual(compact_gap, default_gap * 0.5, places=6)

            captured_default["fig"].clf()
            captured_compact["fig"].clf()

    def test_plot_stacked_vertical_can_reduce_trimestres_bar_width_and_gap(self):
        with tempfile.TemporaryDirectory() as td:
            td_path = Path(td)
            default_output = td_path / "default.png"
            compact_output = td_path / "compact.png"
            captured_default: dict[str, object] = {}
            captured_compact: dict[str, object] = {}

            import matplotlib.pyplot as plt

            original_subplots = plt.subplots

            def _capture_default(*args, **kwargs):
                fig, ax = original_subplots(*args, **kwargs)
                captured_default["fig"] = fig
                captured_default["ax"] = ax
                return fig, ax

            def _capture_compact(*args, **kwargs):
                fig, ax = original_subplots(*args, **kwargs)
                captured_compact["fig"] = fig
                captured_compact["ax"] = ax
                return fig, ax

            with (
                patch("matplotlib.pyplot.subplots", side_effect=_capture_default),
                patch("src.utils.slides.slide8_charts.close_figure"),
            ):
                _plot_stacked_vertical(
                    xlabels=["3T25", "4T25", "1T26"],
                    series_names=["Clientes", "Mercado"],
                    values=np.array([[70.0, 30.0], [90.0, 60.0], [110.0, 80.0]]),
                    output_path=default_output,
                    colors=["#123a7a", "#5B8FF9"],
                    show_delta_bracket=False,
                    show_delta_pct=False,
                    inline_left_legend=False,
                )

            with (
                patch("matplotlib.pyplot.subplots", side_effect=_capture_compact),
                patch("src.utils.slides.slide8_charts.close_figure"),
            ):
                _plot_stacked_vertical(
                    xlabels=["3T25", "4T25", "1T26"],
                    series_names=["Clientes", "Mercado"],
                    values=np.array([[70.0, 30.0], [90.0, 60.0], [110.0, 80.0]]),
                    output_path=compact_output,
                    colors=["#123a7a", "#5B8FF9"],
                    show_delta_bracket=False,
                    show_delta_pct=False,
                    inline_left_legend=False,
                    bar_width_scale=SLIDE8_TRIMESTRES_BAR_WIDTH_SCALE,
                    gap_scale=SLIDE8_TRIMESTRES_GAP_SCALE,
                )

            default_patches = captured_default["ax"].patches
            compact_patches = captured_compact["ax"].patches

            default_width = default_patches[0].get_width()
            compact_width = compact_patches[0].get_width()
            self.assertAlmostEqual(compact_width, default_width * SLIDE8_TRIMESTRES_BAR_WIDTH_SCALE, places=6)

            default_gap = default_patches[1].get_x() - (default_patches[0].get_x() + default_patches[0].get_width())
            compact_gap = compact_patches[1].get_x() - (compact_patches[0].get_x() + compact_patches[0].get_width())
            self.assertAlmostEqual(compact_gap, default_gap * SLIDE8_TRIMESTRES_GAP_SCALE, places=6)

            captured_default["fig"].clf()
            captured_compact["fig"].clf()

    def test_plot_stacked_vertical_can_add_bbox_for_mercado_and_increase_x_tick_pad(self):
        with tempfile.TemporaryDirectory() as td:
            output_path = Path(td) / "out.png"
            captured: dict[str, object] = {}

            import matplotlib.pyplot as plt

            original_subplots = plt.subplots

            def _capture_subplots(*args, **kwargs):
                fig, ax = original_subplots(*args, **kwargs)
                captured["fig"] = fig
                captured["ax"] = ax
                return fig, ax

            with (
                patch("matplotlib.pyplot.subplots", side_effect=_capture_subplots),
                patch("src.utils.slides.slide8_charts.close_figure"),
            ):
                _plot_stacked_vertical(
                    xlabels=["4T24", "3T25", "4T25"],
                    series_names=["Mercado", "Clientes"],
                    values=np.array([[152.0, 2367.0], [244.0, 2051.0], [172.0, 2131.0]]),
                    output_path=output_path,
                    colors=["#123a7a", "#5B8FF9"],
                    font_scale=SLIDE8_FONT_SCALE,
                    show_delta_bracket=True,
                    show_delta_pct=True,
                    show_segment_pct=True,
                    bold_last_bar=True,
                    bold_text=True,
                    inline_left_legend=True,
                    x_tick_pad=SLIDE8_MFB_TRIMESTRES_X_TICK_PAD,
                    segment_label_bbox_series_names=SLIDE8_MFB_TRIMESTRES_LABEL_BBOX_SERIES,
                )

            ax = captured["ax"]
            mercado_text = next(
                text for text in ax.texts
                if text.get_text() == "152\n(6%)"
            )
            bbox_patch = mercado_text.get_bbox_patch()
            self.assertIsNotNone(bbox_patch)
            self.assertEqual(tuple(round(v, 3) for v in bbox_patch.get_facecolor()), tuple(round(v, 3) for v in to_rgba("#123a7a")))

            self.assertEqual(ax.xaxis.majorTicks[0].get_pad(), SLIDE8_MFB_TRIMESTRES_X_TICK_PAD)

            captured["fig"].clf()


if __name__ == "__main__":
    unittest.main()
