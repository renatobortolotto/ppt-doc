import tempfile
import unittest
from pathlib import Path
from unittest.mock import patch

import matplotlib.pyplot as plt
import numpy as np
from matplotlib.patches import FancyBboxPatch
from matplotlib.patches import Rectangle
from openpyxl import Workbook

from src.utils.slides.slide9_charts import (
    SLIDE9_9M_DELTA_BRACKET_COLORS,
    SLIDE9_9M_DELTA_LABEL_X_FRACTIONS,
    SLIDE9_9M_DELTA_PAIRS,
    SLIDE9_COVERAGE_FONT_SCALE,
    SLIDE9_COVERAGE_XTICK_PAD,
    SLIDE9_LINE_9M_LABEL_FONTSIZE,
    SLIDE9_LINE_9M_LABEL_HORIZONTAL_ALIGNMENTS,
    SLIDE9_LINE_9M_LABEL_OFFSETS_PTS,
    SLIDE9_LINE_9M_LABEL_X_OFFSETS_PTS,
    SLIDE9_LINE_LABEL_OFFSET_PTS,
    SLIDE9_LINE_TRI_LABEL_FONTSIZE,
    SLIDE9_STACKED_BAR_WIDTH,
    SLIDE9_STACKED_FONT_SCALE,
    SLIDE9_STACKED_GROUP_SPACING,
    SLIDE9_STACKED_LEGEND_X_OFFSET,
    SLIDE9_STACKED_XTICK_PAD,
    _plot_indice_cobertura_percent,
    _plot_stacked_bars_with_total,
    _wrap_words,
    generate_slide9_charts,
)


class TestSlide9Charts(unittest.TestCase):
    def test_wrap_words_breaks_long_legend_labels(self):
        self.assertEqual(
            _wrap_words("Recuperação de Crédito", max_line_len=15),
            "Recuperação de\nCrédito",
        )

    def test_generate_slide9_charts_uses_updated_sources(self):
        wb = Workbook()
        ws_custo = wb.active
        ws_custo.title = "Tabelas"
        ws_cobertura = wb.create_sheet("Qualidade Cart 4966")

        for col_offset, label in enumerate(["9M23", "9M24", "9M25"], start=4):
            ws_custo.cell(row=2, column=col_offset).value = label
        for col_offset, value in enumerate([100, -110, 120], start=4):
            ws_custo.cell(row=13, column=col_offset).value = value
        for col_offset, value in enumerate([10, 12, 14], start=4):
            ws_custo.cell(row=5, column=col_offset).value = value

        for col_offset, label in enumerate(["4T24", "1T25"], start=7):
            ws_custo.cell(row=2, column=col_offset).value = label
        for col_offset, value in enumerate([-40, 45], start=7):
            ws_custo.cell(row=13, column=col_offset).value = value
        for col_offset, value in enumerate([4, 5], start=7):
            ws_custo.cell(row=5, column=col_offset).value = value

        for col_offset, label in enumerate(["4T24", "1T25", "2T25"], start=4):
            ws_cobertura.cell(row=2, column=col_offset).value = label
        for col_offset, value in enumerate([1.6, 1.7, 1.8], start=4):
            ws_cobertura.cell(row=17, column=col_offset).value = value

        captured: dict[str, dict[str, object]] = {}
        captured_lines: dict[str, dict[str, object]] = {}

        def _capture_stacked(*, xlabels, series_names, values, output_path, colors, **kwargs):
            captured[Path(output_path).name] = {
                "xlabels": list(xlabels),
                "series_names": list(series_names),
                "values": values.tolist(),
                "colors": list(colors),
                "kwargs": dict(kwargs),
            }

        def _capture_cobertura(*, xlabels, values, output_path, highlight_last_count=3, **kwargs):
            captured[Path(output_path).name] = {
                "xlabels": list(xlabels),
                "values": list(values),
                "highlight_last_count": highlight_last_count,
                "kwargs": dict(kwargs),
            }

        def _capture_line(*, file_path, sheet_name, values_range, xlabels_range, output_path, **kwargs):
            captured_lines[Path(output_path).name] = {
                "file_path": str(file_path),
                "sheet_name": sheet_name,
                "values_range": values_range,
                "xlabels_range": xlabels_range,
                "kwargs": dict(kwargs),
            }
            return object(), object()

        with tempfile.TemporaryDirectory() as td:
            xlsx_path = Path(td) / "test.xlsx"
            output_dir = Path(td) / "out"
            wb.save(xlsx_path)

            with patch("src.utils.slides.slide9_charts._plot_stacked_bars_with_total", side_effect=_capture_stacked):
                with patch("src.utils.slides.slide9_charts._plot_indice_cobertura_percent", side_effect=_capture_cobertura):
                    with patch("src.utils.slides.slide9_charts.plot_line_from_excel", side_effect=_capture_line):
                        with patch("src.utils.slides.slide9_charts.close_figure"):
                            files = generate_slide9_charts(xlsx_path=xlsx_path, output_dir=output_dir)

        self.assertEqual(len(files), 5)
        self.assertEqual(
            captured["09_custo_credito_trimestres.png"]["xlabels"],
            ["4T24", "1T25"],
        )
        self.assertEqual(
            captured["09_custo_credito_trimestres.png"]["series_names"],
            ["PDD Expandida", "Recuperação de Crédito"],
        )
        self.assertEqual(
            captured["09_custo_credito_trimestres.png"]["values"],
            [[40.0, -4.0], [45.0, -5.0]],
        )
        self.assertEqual(
            captured["09_custo_credito_trimestres.png"]["kwargs"],
            {
                "font_scale": SLIDE9_STACKED_FONT_SCALE,
                "legend_x_offset": SLIDE9_STACKED_LEGEND_X_OFFSET,
                "x_tick_pad": SLIDE9_STACKED_XTICK_PAD,
                "bar_width": SLIDE9_STACKED_BAR_WIDTH,
                "group_spacing": SLIDE9_STACKED_GROUP_SPACING,
            },
        )

        self.assertEqual(
            captured["09_custo_credito_9m.png"]["xlabels"],
            ["9M23", "9M24", "9M25"],
        )
        self.assertEqual(
            captured["09_custo_credito_9m.png"]["values"],
            [[100.0, -10.0], [110.0, -12.0], [120.0, -14.0]],
        )
        self.assertEqual(
            captured["09_custo_credito_9m.png"]["kwargs"],
            {
                "font_scale": SLIDE9_STACKED_FONT_SCALE,
                "legend_x_offset": SLIDE9_STACKED_LEGEND_X_OFFSET,
                "x_tick_pad": SLIDE9_STACKED_XTICK_PAD,
                "bar_width": SLIDE9_STACKED_BAR_WIDTH,
                "group_spacing": SLIDE9_STACKED_GROUP_SPACING,
                "delta_pairs": SLIDE9_9M_DELTA_PAIRS,
                "delta_bracket_colors": SLIDE9_9M_DELTA_BRACKET_COLORS,
                "delta_label_x_fractions": SLIDE9_9M_DELTA_LABEL_X_FRACTIONS,
            },
        )

        self.assertEqual(
            captured["09_indice_cobertura.png"]["xlabels"],
            ["4T24", "1T25", "2T25"],
        )
        self.assertEqual(
            captured["09_indice_cobertura.png"]["values"],
            [1.6, 1.7, 1.8],
        )
        self.assertEqual(captured["09_indice_cobertura.png"]["highlight_last_count"], 2)
        self.assertEqual(
            captured["09_indice_cobertura.png"]["kwargs"],
            {
                "font_scale": SLIDE9_COVERAGE_FONT_SCALE,
                "x_tick_pad": SLIDE9_COVERAGE_XTICK_PAD,
            },
        )

        self.assertEqual(
            captured_lines["09_custo_variacao_custo_credito.png"]["sheet_name"],
            "Tabelas",
        )
        self.assertEqual(
            captured_lines["09_custo_variacao_custo_credito.png"]["xlabels_range"],
            "D2:F2",
        )
        self.assertEqual(
            captured_lines["09_custo_variacao_custo_credito.png"]["values_range"],
            "D10:F10",
        )
        self.assertTrue(captured_lines["09_custo_variacao_custo_credito.png"]["kwargs"]["fmt_as_percent"])
        self.assertEqual(
            captured_lines["09_custo_variacao_custo_credito.png"]["kwargs"]["label_fontsize"],
            SLIDE9_LINE_TRI_LABEL_FONTSIZE,
        )
        self.assertEqual(captured_lines["09_custo_variacao_custo_credito.png"]["kwargs"]["marker_size"], 160.0)
        self.assertEqual(
            captured_lines["09_custo_variacao_custo_credito.png"]["kwargs"]["label_offset_pts"],
            SLIDE9_LINE_LABEL_OFFSET_PTS,
        )

        self.assertEqual(
            captured_lines["09_custo_variacao_custo_credito_9m.png"]["sheet_name"],
            "Tabelas",
        )
        self.assertEqual(
            captured_lines["09_custo_variacao_custo_credito_9m.png"]["xlabels_range"],
            "G2:H2",
        )
        self.assertEqual(
            captured_lines["09_custo_variacao_custo_credito_9m.png"]["values_range"],
            "G10:H10",
        )
        self.assertTrue(captured_lines["09_custo_variacao_custo_credito_9m.png"]["kwargs"]["fmt_as_percent"])
        self.assertEqual(
            captured_lines["09_custo_variacao_custo_credito_9m.png"]["kwargs"]["label_fontsize"],
            SLIDE9_LINE_9M_LABEL_FONTSIZE,
        )
        self.assertEqual(captured_lines["09_custo_variacao_custo_credito_9m.png"]["kwargs"]["marker_size"], 220.0)
        self.assertEqual(captured_lines["09_custo_variacao_custo_credito_9m.png"]["kwargs"]["x_margin"], 0.55)
        self.assertEqual(
            captured_lines["09_custo_variacao_custo_credito_9m.png"]["kwargs"]["label_offset_pts"],
            SLIDE9_LINE_LABEL_OFFSET_PTS,
        )
        self.assertEqual(
            captured_lines["09_custo_variacao_custo_credito_9m.png"]["kwargs"]["label_offsets_pts"],
            SLIDE9_LINE_9M_LABEL_OFFSETS_PTS,
        )
        self.assertEqual(
            captured_lines["09_custo_variacao_custo_credito_9m.png"]["kwargs"]["label_x_offsets_pts"],
            SLIDE9_LINE_9M_LABEL_X_OFFSETS_PTS,
        )
        self.assertEqual(
            captured_lines["09_custo_variacao_custo_credito_9m.png"]["kwargs"]["label_horizontal_alignments"],
            SLIDE9_LINE_9M_LABEL_HORIZONTAL_ALIGNMENTS,
        )

    def test_plot_stacked_bars_with_total_supports_custom_bracket_layout(self):
        fig, ax = plt.subplots()

        with tempfile.TemporaryDirectory() as td:
            output_path = Path(td) / "chart.png"
            with patch("matplotlib.pyplot.subplots", return_value=(fig, ax)):
                with patch("src.utils.slides.slide9_charts.close_figure"):
                    _plot_stacked_bars_with_total(
                        xlabels=["9M23", "9M24", "9M25"],
                        series_names=["PDD Expandida", "Recuperação de Crédito"],
                        values=np.asarray(
                            [
                                [100.0, -10.0],
                                [110.0, -12.0],
                                [120.0, -14.0],
                            ],
                            dtype=float,
                        ),
                        output_path=output_path,
                        colors=["#0B2E6B", "#5B8FF9"],
                        delta_pairs=SLIDE9_9M_DELTA_PAIRS,
                        delta_bracket_colors=SLIDE9_9M_DELTA_BRACKET_COLORS,
                        delta_label_x_fractions=SLIDE9_9M_DELTA_LABEL_X_FRACTIONS,
                    )

        bracket_lines = [line for line in ax.lines if len(line.get_xdata()) == 4]
        self.assertEqual(len(bracket_lines), 2)
        self.assertEqual(bracket_lines[0].get_color(), SLIDE9_9M_DELTA_BRACKET_COLORS[0])
        self.assertEqual(bracket_lines[1].get_color(), SLIDE9_9M_DELTA_BRACKET_COLORS[1])

        percent_texts = [text for text in ax.texts if "%" in text.get_text()]
        self.assertEqual(len(percent_texts), 2)
        self.assertAlmostEqual(percent_texts[0].get_position()[0], 0.60, places=2)
        self.assertAlmostEqual(percent_texts[1].get_position()[0], 1.50, places=2)

        plt.close(fig)

    def test_plot_stacked_bars_with_total_uses_slide14_legend_style(self):
        fig, ax = plt.subplots()

        with tempfile.TemporaryDirectory() as td:
            output_path = Path(td) / "chart.png"
            with patch("matplotlib.pyplot.subplots", return_value=(fig, ax)):
                with patch("src.utils.slides.slide9_charts.close_figure"):
                    _plot_stacked_bars_with_total(
                        xlabels=["9M23", "9M24", "9M25"],
                        series_names=["PDD Expandida", "Recuperação de Crédito"],
                        values=np.asarray(
                            [
                                [100.0, -10.0],
                                [110.0, -12.0],
                                [120.0, -14.0],
                            ],
                            dtype=float,
                        ),
                        output_path=output_path,
                        colors=["#0B2E6B", "#5B8FF9"],
                    )

        legend_texts = [text for text in ax.texts if text.get_text() in {"PDD Expandida", "Recuperação de\nCrédito"}]
        self.assertEqual(len(legend_texts), 2)
        self.assertTrue(all(text.get_ha() == "right" for text in legend_texts))
        connector_lines = [
            line
            for line in ax.lines
            if len(line.get_xdata()) == 2 and line.get_color() in {"#0B2E6B", "#5B8FF9"}
        ]
        self.assertEqual(connector_lines, [])

        plt.close(fig)

    def test_plot_stacked_bars_with_total_supports_custom_bar_width_and_spacing(self):
        fig, ax = plt.subplots()

        with tempfile.TemporaryDirectory() as td:
            output_path = Path(td) / "chart.png"
            with patch("matplotlib.pyplot.subplots", return_value=(fig, ax)):
                with patch("src.utils.slides.slide9_charts.close_figure"):
                    _plot_stacked_bars_with_total(
                        xlabels=["4T24", "1T25"],
                        series_names=["PDD Expandida", "Recuperação de Crédito"],
                        values=np.asarray(
                            [
                                [40.0, -4.0],
                                [45.0, -5.0],
                            ],
                            dtype=float,
                        ),
                        output_path=output_path,
                        colors=["#0B2E6B", "#5B8FF9"],
                        bar_width=SLIDE9_STACKED_BAR_WIDTH,
                        group_spacing=SLIDE9_STACKED_GROUP_SPACING,
                    )

        bar_patches = [
            patch
            for patch in ax.patches
            if isinstance(patch, Rectangle) and patch.get_width() > 0 and patch.get_height() != 0
        ]
        self.assertEqual(len(bar_patches), 4)
        self.assertAlmostEqual(bar_patches[0].get_width(), SLIDE9_STACKED_BAR_WIDTH, places=2)

        bar_centers = sorted(
            {
                round(float(patch.get_x() + patch.get_width() / 2.0), 2)
                for patch in bar_patches
            }
        )
        self.assertEqual(bar_centers, [0.0, 0.5])

        plt.close(fig)

    def test_plot_indice_cobertura_percent_adds_extra_top_padding_to_highlight_box(self):
        fig, ax = plt.subplots()

        with tempfile.TemporaryDirectory() as td:
            output_path = Path(td) / "chart.png"
            with patch("matplotlib.pyplot.subplots", return_value=(fig, ax)):
                with patch("src.utils.slides.slide9_charts.close_figure"):
                    _plot_indice_cobertura_percent(
                        xlabels=["4T24", "3T25", "4T25"],
                        values=[1.68, 1.78, 1.69],
                        output_path=output_path,
                        highlight_last_count=2,
                        font_scale=SLIDE9_COVERAGE_FONT_SCALE,
                        x_tick_pad=SLIDE9_COVERAGE_XTICK_PAD,
                    )

        highlight_boxes = [patch for patch in ax.patches if isinstance(patch, FancyBboxPatch)]
        self.assertEqual(len(highlight_boxes), 1)
        box = highlight_boxes[0]
        box_top = box.get_y() + box.get_height()
        self.assertGreater(box_top, 178.0 + 18.0)

        plt.close(fig)


if __name__ == "__main__":
    unittest.main()
