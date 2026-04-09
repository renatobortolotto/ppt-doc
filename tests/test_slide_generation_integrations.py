import tempfile
import unittest
from pathlib import Path

from openpyxl import Workbook

from src.utils.slides.slide9_charts import generate_slide9_charts
from src.utils.slides.slide11_charts import generate_slide11_charts
from src.utils.slides.slide12_charts import generate_slide12_charts
from src.utils.slides.slide13_charts import generate_slide13_charts
from src.utils.slides.slide14_charts import generate_slide14_charts
from src.utils.slides.slide15_charts import generate_slide15_charts
from src.utils.slides.slide16_charts import generate_slide16_charts
from src.utils.slides.slide19_charts import generate_slide19_charts
from src.utils.slides.slide20_charts import generate_slide20_charts
from src.utils.slides.slide21_charts import generate_slide21_charts
from src.utils.slides.slide22_charts import generate_slide22_charts


class TestSlideGenerationIntegrations(unittest.TestCase):
    def _assert_generated_files(self, files, expected_names):
        self.assertEqual([path.name for path in files], expected_names)
        for file_path in files:
            self.assertTrue(file_path.exists())
            self.assertGreater(file_path.stat().st_size, 0)

    def test_generate_slide9_charts_creates_expected_pngs(self):
        wb = Workbook()
        ws_custo = wb.active
        ws_custo.title = "Tabelas"
        ws_cobertura = wb.create_sheet("Qualidade Cart 4966")

        for col_offset, label in enumerate(["9M23", "9M24", "9M25"], start=4):
            ws_custo.cell(row=2, column=col_offset).value = label
        for col_offset, label in enumerate(["4T25", "1T26"], start=7):
            ws_custo.cell(row=2, column=col_offset).value = label

        for col_offset, value in enumerate([100, 110, 125], start=4):
            ws_custo.cell(row=13, column=col_offset).value = value
        for col_offset, value in enumerate([12, 14, 16], start=4):
            ws_custo.cell(row=5, column=col_offset).value = value
        for col_offset, value in enumerate([0.08, 0.09, 0.11], start=4):
            ws_custo.cell(row=10, column=col_offset).value = value

        for col_offset, value in enumerate([40, 46], start=7):
            ws_custo.cell(row=13, column=col_offset).value = value
        for col_offset, value in enumerate([4, 5], start=7):
            ws_custo.cell(row=5, column=col_offset).value = value
        for col_offset, value in enumerate([0.05, 0.07], start=7):
            ws_custo.cell(row=10, column=col_offset).value = value

        for col_offset, label in enumerate(["4T25", "1T26", "2T26"], start=4):
            ws_cobertura.cell(row=2, column=col_offset).value = label
        for col_offset, value in enumerate([1.6, 1.7, 1.8], start=4):
            ws_cobertura.cell(row=17, column=col_offset).value = value

        with tempfile.TemporaryDirectory() as td:
            xlsx_path = Path(td) / "slide9.xlsx"
            output_dir = Path(td) / "out"
            wb.save(xlsx_path)
            files = generate_slide9_charts(xlsx_path=xlsx_path, output_dir=output_dir)
            self._assert_generated_files(
                files,
                [
                    "09_custo_credito_trimestres.png",
                    "09_custo_credito_9m.png",
                    "09_indice_cobertura.png",
                    "09_custo_variacao_custo_credito.png",
                    "09_custo_variacao_custo_credito_9m.png",
                ],
            )

    def test_generate_slide11_charts_creates_expected_pngs(self):
        wb = Workbook()
        ws_expenses = wb.active
        ws_expenses.title = "Tabelas"
        ws_index = wb.create_sheet("slide_11")

        for col_offset, label in enumerate(["1T25", "2T25", "3T25"], start=4):
            ws_expenses.cell(row=33, column=col_offset).value = label
        for col_offset, value in enumerate([100, 110, 120], start=4):
            ws_expenses.cell(row=35, column=col_offset).value = value
        for col_offset, value in enumerate([40, 42, 44], start=4):
            ws_expenses.cell(row=39, column=col_offset).value = value
        for col_offset, value in enumerate([10, 11, 12], start=4):
            ws_expenses.cell(row=45, column=col_offset).value = value

        for col_offset, label in enumerate(["9M24", "9M25"], start=7):
            ws_expenses.cell(row=33, column=col_offset).value = label
        for col_offset, value in enumerate([330, 360], start=7):
            ws_expenses.cell(row=35, column=col_offset).value = value
        for col_offset, value in enumerate([126, 132], start=7):
            ws_expenses.cell(row=39, column=col_offset).value = value
        for col_offset, value in enumerate([31, 33], start=7):
            ws_expenses.cell(row=45, column=col_offset).value = value

        for col_offset, label in enumerate(["1T25", "2T25", "3T25", "9M24", "9M25"], start=11):
            ws_index.cell(row=3, column=col_offset).value = label
        for col_offset, value in enumerate([0.37, 0.38, 0.39, 0.40, 0.41], start=11):
            ws_index.cell(row=4, column=col_offset).value = value

        with tempfile.TemporaryDirectory() as td:
            xlsx_path = Path(td) / "slide11.xlsx"
            output_dir = Path(td) / "out"
            wb.save(xlsx_path)
            files = generate_slide11_charts(xlsx_path=xlsx_path, output_dir=output_dir)
            self._assert_generated_files(
                files,
                [
                    "11_despesas_pessoal_adm_trimestres.png",
                    "11_despesas_pessoal_adm_9m.png",
                    "11_indice_eficiencia.png",
                ],
            )

    def test_generate_slide12_charts_creates_expected_pngs(self):
        wb = Workbook()
        ws = wb.active
        ws.title = "Carteira"

        for col_offset, label in enumerate(["3T25", "4T25", "1T26"], start=4):
            ws.cell(row=12, column=col_offset).value = label

        ws["C15"] = "Varejo - Produtos de Entrada"
        ws["C22"] = "Varejo Relacional"
        ws["C34"] = "Atacado"

        for col_offset, value in enumerate([10_000, 11_000, 12_000], start=4):
            ws.cell(row=15, column=col_offset).value = value
        for col_offset, value in enumerate([20_000, 21_000, 22_000], start=4):
            ws.cell(row=22, column=col_offset).value = value
        for col_offset, value in enumerate([30_000, 31_000, 32_000], start=4):
            ws.cell(row=34, column=col_offset).value = value

        with tempfile.TemporaryDirectory() as td:
            xlsx_path = Path(td) / "slide12.xlsx"
            output_dir = Path(td) / "out"
            wb.save(xlsx_path)
            files = generate_slide12_charts(xlsx_path=xlsx_path, output_dir=output_dir)
            self._assert_generated_files(files, ["12_slide12_composicao.png"])

    def test_generate_slide13_charts_creates_expected_pngs(self):
        wb = Workbook()
        ws = wb.active
        ws.title = "Carteira"

        for col_offset, label in enumerate(["3T25", "4T25", "1T26"], start=4):
            ws.cell(row=12, column=col_offset).value = label
            ws.cell(row=115, column=col_offset).value = label

        sources = {
            16: [46888, 47000, 48000],
            17: [3258, 3300, 3400],
            18: [4558, 4600, 4700],
            19: [3731, 3800, 3900],
            20: [333, 350, 360],
            21: [35, 40, 45],
            23: [5262, 5400, 5500],
            24: [5230, 5300, 5400],
            25: [39, 42, 45],
            117: [6547, 6600, 6800],
            118: [2507, 2600, 2700],
            119: [1265, 1300, 1400],
        }
        for row, values in sources.items():
            for col_offset, value in enumerate(values, start=4):
                ws.cell(row=row, column=col_offset).value = value

        with tempfile.TemporaryDirectory() as td:
            xlsx_path = Path(td) / "slide13.xlsx"
            output_dir = Path(td) / "out"
            wb.save(xlsx_path)
            files = generate_slide13_charts(xlsx_path=xlsx_path, output_dir=output_dir)
            self._assert_generated_files(
                files,
                [
                    "13_varejo_produtos_entrada.png",
                    "13_varejo_relacional.png",
                    "13_atacado.png",
                ],
            )

    def test_generate_slide14_charts_creates_expected_pngs(self):
        wb = Workbook()
        ws = wb.active
        ws.title = "Veiculos"

        ws["C10"] = "Leves"
        ws["C11"] = "Pesados"

        for col_offset, label in enumerate(["3T25", "4T25", "1T26"], start=4):
            ws.cell(row=3, column=col_offset).value = label
        for col_offset, label in enumerate(["2025", "2026"], start=7):
            ws.cell(row=3, column=col_offset).value = label

        for col_offset, value in enumerate([120, 130, 140], start=4):
            ws.cell(row=10, column=col_offset).value = value
        for col_offset, value in enumerate([30, 35, 40], start=4):
            ws.cell(row=11, column=col_offset).value = value
        for col_offset, value in enumerate([0.68, 0.70, 0.72], start=4):
            ws.cell(row=7, column=col_offset).value = value

        for col_offset, value in enumerate([390, 420], start=7):
            ws.cell(row=10, column=col_offset).value = value
        for col_offset, value in enumerate([90, 95], start=7):
            ws.cell(row=11, column=col_offset).value = value
        for col_offset, value in enumerate([68, 70], start=7):
            ws.cell(row=7, column=col_offset).value = value

        with tempfile.TemporaryDirectory() as td:
            xlsx_path = Path(td) / "slide14.xlsx"
            output_dir = Path(td) / "out"
            wb.save(xlsx_path)
            files = generate_slide14_charts(xlsx_path=xlsx_path, output_dir=output_dir)
            self._assert_generated_files(
                files,
                [
                    "14_veiculos_empilhado_trimestres.png",
                    "14_veiculos_empilhado_anos.png",
                    "14_veiculos_percentual_trimestres.png",
                    "14_veiculos_percentual_anos.png",
                ],
            )

    def test_generate_slide15_charts_creates_expected_pngs(self):
        wb = Workbook()
        ws = wb.active
        ws.title = "Captacoes"

        for col_offset, label in enumerate(["3T25", "4T25", "1T26"], start=3):
            ws.cell(row=3, column=col_offset).value = label

        series_rows = {
            5: ("Depósito a Prazo", [120, 125, 130]),
            6: ("LCA", [90, 92, 94]),
            10: ("LF", [75, 78, 80]),
            11: ("Poupança", [30, 31, 32]),
            12: ("FIDC", [0, 0, 0]),
            13: ("Letras Financeiras", [45, 47, 48]),
            14: ("DPGE", [18, 20, 21]),
            15: ("Outros", [12, 14, 15]),
        }
        for row, (label, values) in series_rows.items():
            ws.cell(row=row, column=2).value = label
            for col_offset, value in enumerate(values, start=3):
                ws.cell(row=row, column=col_offset).value = value

        with tempfile.TemporaryDirectory() as td:
            xlsx_path = Path(td) / "slide15.xlsx"
            output_dir = Path(td) / "out"
            wb.save(xlsx_path)
            files = generate_slide15_charts(xlsx_path=xlsx_path, output_dir=output_dir)
            self._assert_generated_files(files, ["15_captacoes_trimestres.png"])

    def test_generate_slide16_charts_creates_expected_pngs(self):
        wb = Workbook()
        ws = wb.active
        ws.title = "Basileia"

        for col_offset, label in enumerate(["3T25", "4T25", "1T26"], start=3):
            ws.cell(row=3, column=col_offset).value = label

        values_by_row = {
            18: [0.09, 0.10, 0.11],
            19: [0.03, 0.03, 0.03],
            20: [0.02, 0.02, 0.02],
        }
        for row, values in values_by_row.items():
            for col_offset, value in enumerate(values, start=3):
                ws.cell(row=row, column=col_offset).value = value

        with tempfile.TemporaryDirectory() as td:
            xlsx_path = Path(td) / "slide16.xlsx"
            output_dir = Path(td) / "out"
            wb.save(xlsx_path)
            files = generate_slide16_charts(xlsx_path=xlsx_path, output_dir=output_dir)
            self._assert_generated_files(files, ["16_indice_basileia_trimestres.png"])

    def test_generate_slide19_charts_creates_expected_pngs(self):
        wb = Workbook()
        ws_veiculos = wb.active
        ws_veiculos.title = "Veiculos"
        ws_seguros = wb.create_sheet("Seguros e Cartoes")

        for col_offset, label in enumerate(["3T25", "4T25", "1T26"], start=4):
            ws_veiculos.cell(row=14, column=col_offset).value = label
            ws_seguros.cell(row=3, column=col_offset).value = label
        for col_offset, label in enumerate(["2025", "2026"], start=7):
            ws_seguros.cell(row=3, column=col_offset).value = label

        for col_offset, value in enumerate([120, 130, 140], start=4):
            ws_veiculos.cell(row=21, column=col_offset).value = value
        for col_offset, value in enumerate([20, 22, 24], start=4):
            ws_veiculos.cell(row=22, column=col_offset).value = value
        for col_offset, value in enumerate([15, 17, 19], start=4):
            ws_veiculos.cell(row=23, column=col_offset).value = value

        for col_offset, value in enumerate([310, 330, 350], start=4):
            ws_seguros.cell(row=8, column=col_offset).value = value
        for col_offset, value in enumerate([900, 950], start=7):
            ws_seguros.cell(row=8, column=col_offset).value = value

        with tempfile.TemporaryDirectory() as td:
            xlsx_path = Path(td) / "slide19.xlsx"
            output_dir = Path(td) / "out"
            wb.save(xlsx_path)
            files = generate_slide19_charts(xlsx_path=xlsx_path, output_dir=output_dir)
            self._assert_generated_files(
                files,
                [
                    "19_veiculos_empilhado.png",
                    "19_seguros_cartoes_trimestres.png",
                    "19_seguros_cartoes_anos.png",
                ],
            )

    def test_generate_slide20_charts_creates_expected_pngs(self):
        wb = Workbook()
        ws_emprestimos = wb.active
        ws_emprestimos.title = "Emprestimos"
        ws_seguros = wb.create_sheet("Seguros e Cartoes")

        for col_offset, label in enumerate(["3T25", "4T25", "1T26"], start=4):
            ws_emprestimos.cell(row=3, column=col_offset).value = label
            ws_seguros.cell(row=11, column=col_offset).value = label

        for col_offset, value in enumerate([2000, 2200, 2400], start=4):
            ws_emprestimos.cell(row=5, column=col_offset).value = value
        for col_offset, value in enumerate([1500, 1550, 1600], start=4):
            ws_emprestimos.cell(row=6, column=col_offset).value = value
        for col_offset, value in enumerate([500, 550, 600], start=4):
            ws_emprestimos.cell(row=7, column=col_offset).value = value
        for col_offset, value in enumerate([1200, 1250, 1300], start=4):
            ws_emprestimos.cell(row=8, column=col_offset).value = value

        for col_offset, value in enumerate([3100, 3300, 3500], start=4):
            ws_seguros.cell(row=15, column=col_offset).value = value

        with tempfile.TemporaryDirectory() as td:
            xlsx_path = Path(td) / "slide20.xlsx"
            output_dir = Path(td) / "out"
            wb.save(xlsx_path)
            files = generate_slide20_charts(xlsx_path=xlsx_path, output_dir=output_dir)
            self._assert_generated_files(
                files,
                [
                    "20_emprestimos_empilhado.png",
                    "20_seguros_cartoes_trimestres.png",
                ],
            )

    def test_generate_slide21_charts_creates_expected_pngs(self):
        wb = Workbook()
        ws_carteira = wb.active
        ws_carteira.title = "Carteira"
        ws_setor = wb.create_sheet("Carteira Atac Setor")

        for col_offset, label in enumerate(["3T25", "4T25", "1T26"], start=4):
            ws_carteira.cell(row=115, column=col_offset).value = label

        ws_carteira["C117"] = "Corporate"
        ws_carteira["C118"] = "Large + IF"
        ws_carteira["C119"] = "PME"
        values_by_row = {
            117: [6547, 6600, 6800],
            118: [2507, 2600, 2700],
            119: [1265, 1300, 1400],
        }
        for row, values in values_by_row.items():
            for col_offset, value in enumerate(values, start=4):
                ws_carteira.cell(row=row, column=col_offset).value = value

        ws_setor["C2"] = "2025"
        ws_setor["E2"] = "2026"
        categories = ["Agro", "Indústria", "Serviços", "Comércio"]
        left_values = [0.22, 0.18, 0.27, 0.33]
        right_values = [0.24, 0.20, 0.26, 0.30]
        for row_offset, category in enumerate(categories, start=4):
            ws_setor.cell(row=row_offset, column=2).value = category
        for row_offset, value in enumerate(left_values, start=4):
            ws_setor.cell(row=row_offset, column=4).value = value
        for row_offset, value in enumerate(right_values, start=4):
            ws_setor.cell(row=row_offset, column=6).value = value

        with tempfile.TemporaryDirectory() as td:
            xlsx_path = Path(td) / "slide21.xlsx"
            output_dir = Path(td) / "out"
            wb.save(xlsx_path)
            files = generate_slide21_charts(xlsx_path=xlsx_path, output_dir=output_dir)
            self._assert_generated_files(
                files,
                [
                    "21_carteira_atacado_empilhado.png",
                    "21_carteira_atacado_comparativo.png",
                ],
            )

    def test_generate_slide22_charts_creates_expected_pngs(self):
        wb = Workbook()
        ws = wb.active
        ws.title = "Qualidade Cart 4966"

        for col_offset, label in enumerate(["4T25", "1T26"], start=5):
            ws.cell(row=2, column=col_offset).value = label
            ws.cell(row=30, column=col_offset).value = label
            ws.cell(row=61, column=col_offset).value = label

        stacked_blocks = {
            (32, 34): ((0.20, 0.25), (0.30, 0.28), (0.50, 0.47)),
            (42, 44): ((0.15, 0.18), (0.35, 0.32), (0.50, 0.50)),
            (52, 54): ((0.18, 0.19), (0.22, 0.24), (0.60, 0.57)),
        }
        for (start_row, end_row), rows in stacked_blocks.items():
            for row, values in zip(range(start_row, end_row + 1), rows):
                ws.cell(row=row, column=5).value = values[0]
                ws.cell(row=row, column=6).value = values[1]

        line_blocks = {
            (37, 38): ((0.42, 0.39), (0.50, 0.47)),
            (47, 48): ((0.35, 0.33), (0.44, 0.41)),
            (57, 58): ((0.28, 0.30), (0.36, 0.34)),
        }
        for (start_row, end_row), rows in line_blocks.items():
            for row, values in zip(range(start_row, end_row + 1), rows):
                ws.cell(row=row, column=5).value = values[0]
                ws.cell(row=row, column=6).value = values[1]

        ws["E64"] = 1200
        ws["F64"] = 1500
        ws["E65"] = 0.12
        ws["F65"] = 0.15
        ws["E67"] = 0.45
        ws["F67"] = 0.48

        ws["E19"] = 210
        ws["F19"] = 240
        ws["E20"] = 0.13
        ws["F20"] = 0.14

        with tempfile.TemporaryDirectory() as td:
            xlsx_path = Path(td) / "slide22.xlsx"
            output_dir = Path(td) / "out"
            wb.save(xlsx_path)
            files = generate_slide22_charts(xlsx_path=xlsx_path, output_dir=output_dir)
            self._assert_generated_files(
                files,
                [
                    "22_qualidade_4966_bloco1.png",
                    "22_qualidade_4966_bloco2.png",
                    "22_qualidade_4966_bloco3.png",
                    "22_qualidade_4966_linha1.png",
                    "22_qualidade_4966_linha2.png",
                    "22_qualidade_4966_linha3.png",
                    "22_carteira_reestruturada_barras.png",
                    "22_carteira_reestruturada_linha.png",
                    "22_cobertura_reestruturada_linha.png",
                    "22_npl_barras.png",
                    "22_npl_linha.png",
                ],
            )


if __name__ == "__main__":
    unittest.main()
