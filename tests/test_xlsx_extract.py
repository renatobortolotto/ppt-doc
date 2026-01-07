import json
import tempfile
import unittest
from io import BytesIO
from pathlib import Path

from openpyxl import Workbook

from utils.xlsx_extract import (
    ExtractSpec,
    extract_xlsx_bytes_to_dict,
    extract_xlsx_to_dict,
    parse_specs_args,
    parse_specs_json,
)


def _make_workbook_bytes() -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "DRE Saida"

    # Labels row
    ws["C3"].value = "3T25"
    ws["D3"].value = "4T25"

    # Values row
    ws["C18"].value = 461
    ws["D18"].value = 500

    # Single cell cases
    ws["B2"].value = "ONLY_LABEL"
    ws["B5"].value = 123.45

    bio = BytesIO()
    wb.save(bio)
    return bio.getvalue()


class TestXlsxExtract(unittest.TestCase):
    def test_parse_specs_json_accepts_expected_shape(self):
        specs = [
            {
                "id": "lucroTrimestre",
                "sheet": "DRE Saida",
                "labels_range": "C3:D3",
                "values_range": "C18:D18",
            }
        ]

        with tempfile.TemporaryDirectory() as td:
            path = Path(td) / "specs.json"
            path.write_text(json.dumps(specs), encoding="utf-8")

            parsed = parse_specs_json(path)

        self.assertEqual(len(parsed), 1)
        self.assertEqual(parsed[0].id, "lucroTrimestre")
        self.assertEqual(parsed[0].sheet, "DRE Saida")
        self.assertEqual(parsed[0].labels_range, "C3:D3")
        self.assertEqual(parsed[0].values_range, "C18:D18")

    def test_parse_specs_json_accepts_legacy_keys(self):
        specs = [
            {
                "ID": "lucro9M",
                "sheet": "DRE Saida",
                "labels": "L3:M3",
                "values": "L18:M18",
            }
        ]

        with tempfile.TemporaryDirectory() as td:
            path = Path(td) / "specs.json"
            path.write_text(json.dumps(specs), encoding="utf-8")

            parsed = parse_specs_json(path)

        self.assertEqual(parsed[0].id, "lucro9M")
        self.assertEqual(parsed[0].labels_range, "L3:M3")
        self.assertEqual(parsed[0].values_range, "L18:M18")

    def test_parse_specs_args_expanded_format_no_sheet(self):
        parsed = parse_specs_args(["ROE_9M:L3:M3:L20:M20"], default_sheet="DRE Saida")
        self.assertEqual(parsed[0].id, "ROE_9M")
        self.assertEqual(parsed[0].sheet, "DRE Saida")
        self.assertEqual(parsed[0].labels_range, "L3:M3")
        self.assertEqual(parsed[0].values_range, "L20:M20")

    def test_parse_specs_args_expanded_format_with_sheet(self):
        parsed = parse_specs_args(["ROE_9M:DRE Saida:L3:M3:L20:M20"], default_sheet=None)
        self.assertEqual(parsed[0].sheet, "DRE Saida")

    def test_extract_xlsx_bytes_to_dict_basic(self):
        xlsx_bytes = _make_workbook_bytes()
        specs = [
            ExtractSpec(
                id="lucroTrimestre",
                sheet="DRE Saida",
                labels_range="C3:D3",
                values_range="C18:D18",
            )
        ]

        out = extract_xlsx_bytes_to_dict(xlsx_bytes, specs, include_meta=True)
        self.assertIn("lucroTrimestre", out)
        self.assertEqual(out["lucroTrimestre"]["Labels"], ["3T25", "4T25"])
        self.assertEqual(out["lucroTrimestre"]["Values"], [461.0, 500.0])
        self.assertEqual(out["lucroTrimestre"]["Sheet"], "DRE Saida")

    def test_extract_xlsx_bytes_to_dict_lowercase_fields(self):
        xlsx_bytes = _make_workbook_bytes()
        specs = [
            ExtractSpec(
                id="lucroTrimestre",
                sheet="DRE Saida",
                labels_range="C3:D3",
                values_range="C18:D18",
            )
        ]

        out = extract_xlsx_bytes_to_dict(
            xlsx_bytes,
            specs,
            include_meta=True,
            lowercase_fields=True,
        )
        self.assertEqual(out["lucroTrimestre"]["labels"], ["3T25", "4T25"])
        self.assertEqual(out["lucroTrimestre"]["values"], [461.0, 500.0])
        self.assertEqual(out["lucroTrimestre"]["sheet"], "DRE Saida")
        self.assertEqual(out["lucroTrimestre"]["ranges"]["labels"], "C3:D3")

    def test_extract_xlsx_bytes_to_dict_raises_on_empty_bytes(self):
        with self.assertRaises(ValueError):
            extract_xlsx_bytes_to_dict(b"", [ExtractSpec("x", "A1", "A1", sheet="S")])

    def test_extract_xlsx_bytes_to_dict_raises_on_invalid_xlsx(self):
        bad = b"not an xlsx"
        specs = [
            ExtractSpec(
                id="x",
                sheet="DRE Saida",
                labels_range="A1",
                values_range="A1",
            )
        ]
        with self.assertRaises(ValueError):
            extract_xlsx_bytes_to_dict(bad, specs)

    def test_extract_xlsx_to_dict_from_path(self):
        xlsx_bytes = _make_workbook_bytes()
        with tempfile.TemporaryDirectory() as td:
            xlsx_path = Path(td) / "file.xlsx"
            xlsx_path.write_bytes(xlsx_bytes)

            specs = [
                ExtractSpec(
                    id="single",
                    sheet="DRE Saida",
                    labels_range="B2",
                    values_range="B5",
                )
            ]

            out = extract_xlsx_to_dict(xlsx_path, specs, include_meta=False)

        self.assertEqual(out["single"]["Labels"], ["ONLY_LABEL"])
        self.assertEqual(out["single"]["Values"], [123.45])


if __name__ == "__main__":
    unittest.main()
