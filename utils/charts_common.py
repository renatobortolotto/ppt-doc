from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path
from typing import List, Optional, Sequence, Tuple, Union
import unicodedata

import matplotlib

# Headless rendering (safe for CLI jobs)
matplotlib.use("Agg", force=True)

import matplotlib.pyplot as plt
import numpy as np
from openpyxl import load_workbook
from openpyxl.utils.cell import range_boundaries


def _parse_number_like(value: str) -> float:
    """Parse numeric strings, including pt-BR formats and percent strings.

    Examples accepted:
    - "9%", "9 %", "9,5%"
    - "1.234,56" (pt-BR thousands + decimal)
    - "1,234.56" (en-US thousands + decimal)
    """

    s = value.replace("\u00a0", " ").strip()
    if s == "":
        return 0.0

    negative = False
    if s.startswith("(") and s.endswith(")"):
        negative = True
        s = s[1:-1].strip()

    if s.startswith("+"):
        s = s[1:].strip()

    if s.endswith("%"):
        s = s[:-1].strip()

    s = s.replace(" ", "")
    if s == "":
        return 0.0

    # Heuristic for decimal/thousand separators
    if "," in s and "." in s:
        last_comma = s.rfind(",")
        last_dot = s.rfind(".")
        if last_comma > last_dot:
            # pt-BR style: 1.234,56
            s = s.replace(".", "")
            s = s.replace(",", ".")
        else:
            # en-US style: 1,234.56
            s = s.replace(",", "")
    elif "," in s:
        # Likely pt-BR decimal: 12,34
        s = s.replace(".", "")
        s = s.replace(",", ".")
    else:
        # If only dots exist, decide between decimal vs thousands grouping.
        # Treat patterns like 1.234 or 1.234.567 as thousands separators.
        if "." in s:
            parts = s.split(".")
            if len(parts) > 1 and all(p.isdigit() for p in parts) and all(len(p) == 3 for p in parts[1:]):
                s = "".join(parts)
        # Remove thousands commas if present
        s = s.replace(",", "")

    num = float(s)
    return -num if negative else num


@dataclass(frozen=True)
class ExcelBarChartSpec:
    file_path: Union[str, Path]
    sheet_name: str
    values_range: str
    xlabels_range: str
    ylabel_cell: Optional[str] = None
    title: Optional[str] = None
    highlight_last: bool = True
    # Optional: force a single color for all bars (e.g. "#123a7a")
    bar_color: Optional[str] = None
    output_path: Optional[Union[str, Path]] = None
    show_delta_pct: bool = False
    show_delta_bracket: bool = False
    delta_pairs: Tuple[Tuple[int, int], ...] = ()
    fixed_slot_count: Optional[int] = None
    # Optional: format bar-top values with decimals (None keeps existing 0-decimal format)
    value_decimals: Optional[int] = None
    # Optional: multiply computed bar width (e.g. 0.70 for 30% thinner)
    bar_width_scale: float = 1.0
    # Optional: scale all font sizes (e.g. 1.5 to increase by ~50%)
    font_scale: float = 1.0


@dataclass(frozen=True)
class ExcelDonutChartSpec:
    """Spec for a nested donut chart with categories (outer) and segments (inner)."""
    file_path: Union[str, Path]
    sheet_name: str
    # Range with segment labels (column B in the example)
    labels_range: str
    # Range with segment values (column C)
    values_range: str
    # Range with category names (column A) - used to group segments
    categories_range: str
    # Center text (e.g., "Carteira\nAmpliada\nR$ 92.7 bi")
    center_text: str
    # Output path for the PNG
    output_path: Union[str, Path]
    # Optional title
    title: Optional[str] = None
    # Colors for inner segments (one per segment)
    inner_colors: Optional[List[str]] = None
    # Colors for outer categories (one per category)
    outer_colors: Optional[List[str]] = None
    # Figure size
    figsize: Tuple[float, float] = (16, 12)
    # Optional: scale all font sizes (e.g. 1.5 to increase by ~50%)
    font_scale: float = 1.0


def _read_range_row(ws, cell_range: str) -> List[object]:
    min_col, min_row, max_col, max_row = range_boundaries(cell_range)
    out: List[object] = []
    for r in range(min_row, max_row + 1):
        for c in range(min_col, max_col + 1):
            out.append(ws.cell(row=r, column=c).value)
    return out


def _cell_is_percent_formatted(cell) -> bool:
    try:
        fmt = cell.number_format
    except Exception:
        return False
    if not fmt:
        return False
    return "%" in str(fmt)


def read_range_col(ws, cell_range: str) -> List[object]:
    """Read a vertical A1 range and return a list."""

    min_col, min_row, max_col, max_row = range_boundaries(cell_range)
    out: List[object] = []
    for r in range(min_row, max_row + 1):
        for c in range(min_col, max_col + 1):
            out.append(ws.cell(row=r, column=c).value)
    return out


def to_float_list(values: Sequence[object]) -> List[float]:
    out: List[float] = []
    for v in values:
        if v is None or (isinstance(v, str) and v.strip() == ""):
            out.append(0.0)
            continue
        try:
            if isinstance(v, str):
                out.append(_parse_number_like(v))
            else:
                out.append(float(v))
        except (TypeError, ValueError) as exc:
            raise ValueError(f"Valor não numérico no range: {v!r}") from exc
    return out


def pchip_interpolate(x: np.ndarray, y: np.ndarray, x_new: np.ndarray) -> np.ndarray:
    """Monotone cubic interpolation (PCHIP) in NumPy only."""

    x = np.asarray(x, dtype=float)
    y = np.asarray(y, dtype=float)
    x_new = np.asarray(x_new, dtype=float)
    n = x.size
    if n < 2:
        return np.full_like(x_new, y[0] if n else np.nan, dtype=float)
    if np.any(np.diff(x) <= 0):
        raise ValueError("x deve ser estritamente crescente")

    h = np.diff(x)
    delta = np.diff(y) / h
    d = np.zeros(n, dtype=float)

    if n == 2:
        d[0] = delta[0]
        d[1] = delta[0]
    else:
        for i in range(1, n - 1):
            if delta[i - 1] == 0.0 or delta[i] == 0.0 or np.sign(delta[i - 1]) != np.sign(delta[i]):
                d[i] = 0.0
            else:
                w1 = 2 * h[i] + h[i - 1]
                w2 = h[i] + 2 * h[i - 1]
                d[i] = (w1 + w2) / (w1 / delta[i - 1] + w2 / delta[i])

        d0 = ((2 * h[0] + h[1]) * delta[0] - h[0] * delta[1]) / (h[0] + h[1])
        if np.sign(d0) != np.sign(delta[0]):
            d0 = 0.0
        elif (np.sign(delta[0]) != np.sign(delta[1])) and (abs(d0) > abs(3 * delta[0])):
            d0 = 3 * delta[0]
        d[0] = d0

        dn = ((2 * h[-1] + h[-2]) * delta[-1] - h[-1] * delta[-2]) / (h[-1] + h[-2])
        if np.sign(dn) != np.sign(delta[-1]):
            dn = 0.0
        elif (np.sign(delta[-1]) != np.sign(delta[-2])) and (abs(dn) > abs(3 * delta[-1])):
            dn = 3 * delta[-1]
        d[-1] = dn

    idx = np.searchsorted(x, x_new, side="right") - 1
    idx = np.clip(idx, 0, n - 2)
    xk = x[idx]
    xk1 = x[idx + 1]
    hk = xk1 - xk
    t = (x_new - xk) / hk

    yk = y[idx]
    yk1 = y[idx + 1]
    dk = d[idx]
    dk1 = d[idx + 1]

    t2 = t * t
    t3 = t2 * t
    h00 = 2 * t3 - 3 * t2 + 1
    h10 = t3 - 2 * t2 + t
    h01 = -2 * t3 + 3 * t2
    h11 = t3 - t2

    return h00 * yk + h10 * hk * dk + h01 * yk1 + h11 * hk * dk1


def plot_bar_from_excel(spec: ExcelBarChartSpec) -> Tuple[plt.Figure, plt.Axes]:
    file_path = Path(spec.file_path)
    if not file_path.exists():
        raise FileNotFoundError(f"Arquivo não encontrado: {file_path}")

    wb = load_workbook(filename=file_path, data_only=True)
    if spec.sheet_name not in wb.sheetnames:
        raise ValueError(f"Aba não encontrada: {spec.sheet_name!r}. Disponíveis: {wb.sheetnames}")

    ws = wb[spec.sheet_name]

    values = to_float_list(_read_range_row(ws, spec.values_range))
    xlabels = ["" if v is None else str(v) for v in _read_range_row(ws, spec.xlabels_range)]

    if len(values) != len(xlabels):
        raise ValueError(f"Tamanhos diferentes: valores={len(values)} xlabels={len(xlabels)}")

    fig, ax = plt.subplots(figsize=(10, 4.8))
    fig.patch.set_alpha(0)
    ax.set_facecolor("none")

    n = len(values)

    if spec.bar_color:
        colors = [str(spec.bar_color)] * n
    else:
        colors = ["#8d98a6"] * n
        if spec.highlight_last and colors:
            colors[-1] = "#123a7a"

    base_width = 0.8
    slot_count = int(spec.fixed_slot_count) if spec.fixed_slot_count else n
    slot_count = max(slot_count, n)
    ratio = n / slot_count if slot_count else 1.0
    width = base_width if slot_count == n else base_width * float(np.sqrt(ratio))
    try:
        width = float(width) * float(spec.bar_width_scale)
    except Exception:
        pass

    is_9m_two_bars = bool(spec.fixed_slot_count) and n == 2 and slot_count > n

    try:
        font_scale = float(spec.font_scale)
    except Exception:
        font_scale = 1.0
    if not np.isfinite(font_scale) or font_scale <= 0:
        font_scale = 1.0

    font_base = 9
    if is_9m_two_bars:
        font_base = 12
    font_value = float(font_base) * font_scale
    font_xtick = float(font_base) * font_scale
    font_delta = float(font_base) * font_scale

    step = 1.0
    x_pad = 0.5
    if is_9m_two_bars:
        step = 0.48
        x_pad = 0.35
    x_pos = np.arange(n, dtype=float) * step

    bars = ax.bar(x_pos, values, width=width, color=colors, edgecolor="none")
    ax.set_xlim(x_pos.min() - x_pad, x_pos.max() + x_pad)

    last_idx = n - 1
    for i, (rect, val) in enumerate(zip(bars, values)):
        if spec.value_decimals is None:
            value_label = f"{val:,.0f}".replace(",", ".")
        else:
            dec = max(0, int(spec.value_decimals))
            # Requested formatting for some charts: X.X using dot as decimal separator.
            value_label = f"{float(val):.{dec}f}"
        ax.text(
            rect.get_x() + rect.get_width() / 2,
            rect.get_height(),
            value_label,
            ha="center",
            va="bottom",
            fontsize=font_value,
            fontweight="bold" if i == last_idx else "normal",
            color="#2f2f2f",
        )

    delta_label_top: Optional[float] = None
    if spec.show_delta_pct and n >= 2:
        vals = np.asarray(values, dtype=float)
        abs_max = float(np.nanmax(np.abs(vals))) if np.isfinite(np.nanmax(np.abs(vals))) else 0.0
        offset_y = max(abs_max * 0.06, 0.5)
        bracket_h = max(abs_max * 0.03, 0.5)

        pairs = list(spec.delta_pairs) if spec.delta_pairs else [(i - 1, i) for i in range(1, n)]

        def _norm_index(idx: int) -> int:
            return idx + n if idx < 0 else idx

        norm_pairs: List[Tuple[int, int]] = []
        for prev_i, curr_i in pairs:
            pi = _norm_index(int(prev_i))
            ci = _norm_index(int(curr_i))
            if pi < 0 or pi >= n or ci < 0 or ci >= n or pi == ci:
                continue
            norm_pairs.append((pi, ci))

        norm_pairs_sorted = sorted(norm_pairs, key=lambda p: (abs(p[1] - p[0]), p[0], p[1]))

        for level, (pi, ci) in enumerate(norm_pairs_sorted):
            prev = vals[pi]
            curr = vals[ci]
            if not np.isfinite(prev) or not np.isfinite(curr) or prev == 0:
                continue
            pct = (curr / prev - 1.0) * 100.0
            label = f"{pct:+.1f}%".replace(".", ",")

            x1 = bars[pi].get_x() + bars[pi].get_width() / 2
            x2 = bars[ci].get_x() + bars[ci].get_width() / 2

            top = max(prev, curr)
            base = top + offset_y if top >= 0 else top - offset_y
            y_anchor = base + level * (bracket_h + offset_y * 0.9)

            if spec.show_delta_bracket:
                ax.plot(
                    [x1, x1, x2, x2],
                    [y_anchor, y_anchor + bracket_h, y_anchor + bracket_h, y_anchor],
                    color="#2f2f2f",
                    linewidth=1.2,
                    solid_capstyle="round",
                    zorder=4,
                )
                text_y = y_anchor + bracket_h + offset_y * 0.25
            else:
                text_y = y_anchor

            ax.text(
                (x1 + x2) / 2.0,
                text_y,
                label,
                ha="center",
                va="bottom",
                fontsize=font_delta,
                color="#2f2f2f",
                zorder=5,
            )
            delta_label_top = text_y if delta_label_top is None else max(delta_label_top, text_y)

        if delta_label_top is not None:
            cur_ymin, cur_ymax = ax.get_ylim()
            ax.set_ylim(cur_ymin, max(cur_ymax, delta_label_top + offset_y))

    ax.set_xticks(x_pos)
    ax.set_xticklabels(xlabels, rotation=0, fontsize=font_xtick)
    ax.set_ylabel("")
    ax.set_yticks([])
    for s in ("left", "right", "top"):
        ax.spines[s].set_visible(False)
    ax.spines["bottom"].set_visible(True)
    ax.tick_params(axis="y", left=False, labelleft=False)
    ax.yaxis.grid(False)
    ax.margins(x=0.02, y=0.10)

    fig.tight_layout(pad=0.2)

    if spec.output_path:
        out = Path(spec.output_path)
        out.parent.mkdir(parents=True, exist_ok=True)
        fig.savefig(out, dpi=220, transparent=True, bbox_inches="tight", pad_inches=0.05)

    return fig, ax


def plot_line_from_excel(
    *,
    file_path: Union[str, Path],
    sheet_name: str,
    values_range: str,
    xlabels_range: str,
    output_path: Union[str, Path],
    fmt_as_percent: bool = True,
    smooth: bool = True,
    smooth_points: int = 250,
    show_markers: bool = True,
    y_baseline: Optional[float] = None,
    y_expand: float = 0.0,
    marker_color: str = "#123a7a",
    label_offset_pts: float = 10.0,
    line_width: float = 2.2,
    label_fontsize: float = 9.0,
    marker_size: float = 26.0,
) -> Tuple[plt.Figure, plt.Axes]:
    file_path = Path(file_path)
    if not file_path.exists():
        raise FileNotFoundError(f"Arquivo não encontrado: {file_path}")

    wb = load_workbook(filename=file_path, data_only=True)
    if sheet_name not in wb.sheetnames:
        raise ValueError(f"Aba não encontrada: {sheet_name!r}. Disponíveis: {wb.sheetnames}")

    ws = wb[sheet_name]

    # Excel stores percent-formatted numeric cells as fractions (e.g. 0.09 shown as 9%).
    # When fmt_as_percent=True, we want the chart labels in percentage points.
    min_col, min_row, max_col, max_row = range_boundaries(values_range)
    values: List[float] = []
    for r in range(min_row, max_row + 1):
        for c in range(min_col, max_col + 1):
            cell = ws.cell(row=r, column=c)
            raw = cell.value
            if raw is None or (isinstance(raw, str) and raw.strip() == ""):
                v = 0.0
            elif isinstance(raw, str):
                v = _parse_number_like(raw)
            else:
                v = float(raw)

            if fmt_as_percent and _cell_is_percent_formatted(cell) and isinstance(raw, (int, float)):
                v *= 100.0
            values.append(v)

    xlabels = ["" if v is None else str(v) for v in _read_range_row(ws, xlabels_range)]
    if len(values) != len(xlabels):
        raise ValueError(f"Tamanhos diferentes: valores={len(values)} xlabels={len(xlabels)}")

    x = np.arange(len(values), dtype=float)
    y = np.asarray(values, dtype=float)

    fig, ax = plt.subplots(figsize=(10, 4.2))
    fig.patch.set_alpha(0)
    ax.set_facecolor("none")

    if smooth and len(values) >= 3:
        xs = np.linspace(x.min(), x.max(), num=max(int(smooth_points), len(values) * 50))
        ys = pchip_interpolate(x, y, xs)
        ax.plot(
            xs,
            ys,
            linewidth=float(line_width),
            color="#2f2f2f",
            solid_joinstyle="round",
            solid_capstyle="round",
            zorder=2,
        )
    else:
        ax.plot(
            x,
            y,
            linewidth=float(line_width),
            color="#2f2f2f",
            solid_joinstyle="round",
            solid_capstyle="round",
            zorder=2,
        )

    if show_markers:
        ax.scatter(x, y, s=float(marker_size), color=marker_color, zorder=3)

    last_idx = len(values) - 1
    for i, (xi, yi) in enumerate(zip(x, y)):
        label = f"{yi:.1f}%".replace(".", ",") if fmt_as_percent else str(yi).replace(".", ",")
        ax.annotate(
            label,
            (xi, yi),
            textcoords="offset points",
            xytext=(0, label_offset_pts),
            ha="center",
            va="bottom",
            fontsize=float(label_fontsize),
            color="#2f2f2f",
            fontweight="bold" if i == last_idx else "normal",
        )

    ymin = float(np.nanmin(y))
    ymax = float(np.nanmax(y))
    if y_baseline is not None:
        ymin = min(ymin, float(y_baseline))
        ymax = max(ymax, float(y_baseline))
    yr = ymax - ymin
    if not np.isfinite(yr) or yr <= 0:
        yr = 1.0
    if y_expand and y_expand > 0:
        ymin -= yr * float(y_expand)
        ymax += yr * float(y_expand)
    ax.set_ylim(ymin, ymax)

    for s in ("left", "right", "top", "bottom"):
        ax.spines[s].set_visible(False)
    ax.set_xticks([])
    ax.set_yticks([])
    ax.tick_params(left=False, bottom=False, labelleft=False, labelbottom=False)
    ax.grid(False)
    ax.set_title("")
    ax.set_xlabel("")
    ax.set_ylabel("")
    ax.margins(x=0.03, y=0.08)

    fig.tight_layout(pad=0.2)

    out = Path(output_path)
    out.parent.mkdir(parents=True, exist_ok=True)
    fig.savefig(out, dpi=220, transparent=True, bbox_inches="tight", pad_inches=0.08)

    return fig, ax


def plot_donut_from_excel(spec: ExcelDonutChartSpec) -> Tuple[plt.Figure, plt.Axes]:
    """
    Generate a nested donut chart from Excel data.
    
    - Outer ring: aggregated categories
    - Inner ring: individual segments with box labels
    """
    from collections import OrderedDict
    
    file_path = Path(spec.file_path)
    if not file_path.exists():
        raise FileNotFoundError(f"Arquivo não encontrado: {file_path}")

    wb = load_workbook(filename=file_path, data_only=True)
    if spec.sheet_name not in wb.sheetnames:
        raise ValueError(f"Aba não encontrada: {spec.sheet_name!r}. Disponíveis: {wb.sheetnames}")

    ws = wb[spec.sheet_name]

    # Read data from Excel
    categories_raw = _read_range_row(ws, spec.categories_range)
    labels_raw = _read_range_row(ws, spec.labels_range)
    values_raw = _read_range_row(ws, spec.values_range)

    # Filter out empty rows and build lists
    categories = []
    labels = []
    values = []
    for cat, lbl, val in zip(categories_raw, labels_raw, values_raw):
        if lbl is not None and val is not None:
            categories.append(str(cat) if cat else "")
            labels.append(str(lbl))
            values.append(float(val) if val else 0.0)

    # Aggregate by category for outer ring
    category_totals = OrderedDict()
    for cat, val in zip(categories, values):
        category_totals[cat] = category_totals.get(cat, 0) + val
    
    cat_labels = list(category_totals.keys())
    cat_values = list(category_totals.values())

    def _norm_text(s: str) -> str:
        s = "" if s is None else str(s)
        s = unicodedata.normalize("NFKD", s)
        s = "".join(ch for ch in s if not unicodedata.combining(ch))
        s = " ".join(s.strip().lower().split())
        return s

    def _hex_to_rgb(hex_color: str) -> Tuple[float, float, float]:
        s = str(hex_color).strip().lstrip("#")
        if len(s) == 3:
            s = "".join(ch * 2 for ch in s)
        if len(s) != 6:
            raise ValueError(f"Cor inválida: {hex_color!r}")
        r = int(s[0:2], 16) / 255.0
        g = int(s[2:4], 16) / 255.0
        b = int(s[4:6], 16) / 255.0
        return (r, g, b)

    def _rgb_to_hex(rgb: Tuple[float, float, float]) -> str:
        r, g, b = rgb
        r_i = int(max(0, min(255, round(r * 255.0))))
        g_i = int(max(0, min(255, round(g * 255.0))))
        b_i = int(max(0, min(255, round(b * 255.0))))
        return f"#{r_i:02x}{g_i:02x}{b_i:02x}"

    def _lighten(hex_color: str, t: float) -> str:
        """Blend color with white. t=0 keeps original, t=1 becomes white."""
        t = float(t)
        if not np.isfinite(t):
            t = 0.0
        t = max(0.0, min(1.0, t))
        r, g, b = _hex_to_rgb(hex_color)
        r2 = r + (1.0 - r) * t
        g2 = g + (1.0 - g) * t
        b2 = b + (1.0 - b) * t
        return _rgb_to_hex((r2, g2, b2))

    # Palette by top-level category (outer ring). Outer stays darkest; inner uses lighter shades.
    # NOTE: normalize to match variants like "Veículos" vs "Veiculos".
    base_by_category = {
        "veiculos leves": "#1f3a8a",  # navy (dark)
        "atacado": "#b11226",  # red (dark)
        "growth": "#0f766e",  # teal (dark)
    }

    def _base_color_for_category(cat: str) -> str:
        key = _norm_text(cat)
        if key in base_by_category:
            return base_by_category[key]
        # Fallback: stable palette
        fallback = ["#1f3a8a", "#b11226", "#0f766e", "#7c3aed", "#f59e0b"]
        return fallback[abs(hash(key)) % len(fallback)]

    if spec.outer_colors:
        # Keep explicit override (but ensure length matches).
        outer_colors = list(spec.outer_colors)[: len(cat_labels)]
    else:
        outer_colors = [_base_color_for_category(cat) for cat in cat_labels]

    if spec.inner_colors:
        inner_colors = list(spec.inner_colors)[: len(labels)]
    else:
        # Generate inner colors as lighter shades within each category palette.
        inner_colors = ["#cccccc"] * len(labels)
        cat_to_indices: "OrderedDict[str, List[int]]" = OrderedDict()
        for idx, cat in enumerate(categories):
            cat_to_indices.setdefault(cat, []).append(idx)

        for cat, idxs in cat_to_indices.items():
            base = _base_color_for_category(cat)
            m = len(idxs)
            if m <= 1:
                ts = [0.40]
            else:
                # Keep inner segments noticeably lighter than the outer ring.
                ts = list(np.linspace(0.18, 0.68, num=m))
            for j, idx in enumerate(idxs):
                inner_colors[idx] = _lighten(base, ts[j])

    font_scale = float(spec.font_scale) if spec.font_scale else 1.0

    # Create figure
    fig, ax = plt.subplots(figsize=spec.figsize)
    fig.patch.set_alpha(0)
    ax.set_facecolor("none")

    # --- OUTER RING (Categories) ---
    outer_result = ax.pie(
        cat_values,
        startangle=90,
        counterclock=False,
        colors=outer_colors,
        wedgeprops=dict(width=0.25, edgecolor="white", linewidth=2),
        radius=1.0,
    )
    outer_wedges = outer_result[0]

    # --- INNER RING (Segments) ---
    inner_result = ax.pie(
        values,
        startangle=90,
        counterclock=False,
        colors=inner_colors,
        wedgeprops=dict(width=0.30, edgecolor="white", linewidth=1),
        radius=0.70,
    )
    inner_wedges = inner_result[0]

    # --- CATEGORY LABELS (outer ring) with boxes ---
    outer_total = sum(cat_values)
    for i, (wedge, label, value, color) in enumerate(zip(outer_wedges, cat_labels, cat_values, outer_colors)):
        ang = (wedge.theta2 + wedge.theta1) / 2
        pct = value / outer_total * 100

        # Special adjustment for "Veiculos Leves" to avoid overlap
        if "Veiculos" in label or "Veículos" in label:
            r = 1.35
            ang_adjusted = ang + 15
        else:
            r = 1.18
            ang_adjusted = ang

        x = r * np.cos(np.deg2rad(ang_adjusted))
        y = r * np.sin(np.deg2rad(ang_adjusted))

        ax.annotate(
            f"{label}\n{pct:.0f}%",
            xy=(0.9 * np.cos(np.deg2rad(ang)), 0.9 * np.sin(np.deg2rad(ang))),
            xytext=(x, y),
            fontsize=10 * font_scale,
            fontweight="bold",
            ha="center",
            va="center",
            color="white",
            bbox=dict(
                boxstyle="round,pad=0.4",
                facecolor=color,
                edgecolor="white",
                linewidth=2,
            ),
            arrowprops=dict(
                arrowstyle="-",
                color=color,
                lw=2,
            ),
        )

    # --- SEGMENT LABELS (inner ring) with boxes ---
    inner_total = sum(values)
    for i, (wedge, label, value, color) in enumerate(zip(inner_wedges, labels, values, inner_colors)):
        ang = (wedge.theta2 + wedge.theta1) / 2
        pct = value / inner_total * 100

        r_start = 0.55
        x_start = r_start * np.cos(np.deg2rad(ang))
        y_start = r_start * np.sin(np.deg2rad(ang))

        # Special adjustment for "Veiculos Leves Usados" to avoid overlap
        if "Usados" in label:
            r_end = 1.65
            ang_adjusted = ang - 15
        else:
            r_end = 1.50
            ang_adjusted = ang

        x_end = r_end * np.cos(np.deg2rad(ang_adjusted))
        y_end = r_end * np.sin(np.deg2rad(ang_adjusted))

        ax.annotate(
            f"{label}\n{pct:.0f}%",
            xy=(x_start, y_start),
            xytext=(x_end, y_end),
            fontsize=8 * font_scale,
            ha="center",
            va="center",
            bbox=dict(
                boxstyle="round,pad=0.3",
                facecolor="white",
                edgecolor=color,
                linewidth=1.5,
            ),
            arrowprops=dict(
                arrowstyle="-",
                color=color,
                lw=1,
            ),
        )

    # --- CENTER TEXT ---
    ax.text(
        0, 0, spec.center_text,
        ha="center", va="center",
        fontsize=14 * font_scale, fontweight="bold",
        linespacing=1.4,
        bbox=dict(
            boxstyle="circle,pad=0.5",
            facecolor="white",
            edgecolor="lightgray",
            linewidth=1,
        ),
    )

    if spec.title:
        ax.set_title(spec.title, fontsize=15 * font_scale, fontweight="bold", pad=20)

    ax.set_aspect('equal')
    ax.set_xlim(-2.2, 2.2)
    ax.set_ylim(-2.0, 2.0)
    fig.tight_layout()

    # Save to file
    out = Path(spec.output_path)
    out.parent.mkdir(parents=True, exist_ok=True)
    fig.savefig(out, dpi=220, transparent=True, bbox_inches="tight", pad_inches=0.08)

    return fig, ax


def close_figure(fig: plt.Figure) -> None:
    try:
        plt.close(fig)
    except Exception:
        pass
