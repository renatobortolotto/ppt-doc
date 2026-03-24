from __future__ import annotations

from pathlib import Path

import numpy as np
from openpyxl import load_workbook
from openpyxl.utils.cell import range_boundaries

from utils.charts_common import close_figure, to_float_list


def _read_range_row(ws, cell_range: str) -> list[object]:
    min_col, min_row, max_col, max_row = range_boundaries(cell_range)
    out: list[object] = []
    for r in range(min_row, max_row + 1):
        for c in range(min_col, max_col + 1):
            out.append(ws.cell(row=r, column=c).value)
    return out


def _fmt_int(v: float) -> str:
    return f"{float(v):,.0f}".replace(",", ".")


def _wrap_words(text: str, *, max_line_len: int = 16) -> str:
    s = (text or "").strip()
    if not s:
        return ""

    words = s.split()
    lines: list[str] = []
    current: list[str] = []
    current_len = 0

    for word in words:
        word_len = len(word)
        if not current:
            current = [word]
            current_len = word_len
            continue

        if current_len + 1 + word_len <= max_line_len:
            current.append(word)
            current_len += 1 + word_len
            continue

        lines.append(" ".join(current))
        current = [word]
        current_len = word_len

    if current:
        lines.append(" ".join(current))

    return "\n".join(lines)


def _is_blank(value: object) -> bool:
    return value is None or (isinstance(value, str) and value.strip() == "")


def _read_named_series_rows(
    ws,
    *,
    xlabels_range: str,
    series_specs: list[tuple[str, str]],
) -> tuple[list[str], list[str], np.ndarray]:
    raw_xlabels = _read_range_row(ws, xlabels_range)
    if all(_is_blank(value) for value in raw_xlabels):
        raise ValueError(f"Range sem labels para gráfico em {ws.title}!{xlabels_range}")

    xlabels = [("" if v is None else str(v)).strip() for v in raw_xlabels]
    point_count = len(xlabels)
    if point_count == 0:
        raise ValueError(f"Sem labels para gráfico em {ws.title}!{xlabels_range}")

    series_names: list[str] = []
    series_columns: list[list[float]] = []
    for series_name, values_range in series_specs:
        raw_values = _read_range_row(ws, values_range)
        if all(_is_blank(value) for value in raw_values):
            raise ValueError(f"Range sem dados para gráfico em {ws.title}!{values_range}")

        series_values = to_float_list(raw_values)
        if len(series_values) != point_count:
            raise ValueError(
                f"Quantidade de labels em {xlabels_range} difere de {values_range}: "
                f"{point_count} != {len(series_values)}"
            )
        series_names.append(series_name)
        series_columns.append(series_values)

    values = np.column_stack(series_columns)
    return xlabels, series_names, values


def _plot_stacked_expenses(
    *,
    xlabels: list[str],
    series_names: list[str],
    values: np.ndarray,  # [n_bars, n_series]
    output_path: Path,
) -> None:
    import matplotlib.pyplot as plt
    from matplotlib.colors import to_rgba

    n, m = values.shape
    if n == 0 or m == 0:
        raise ValueError("Dados vazios para plotar")

    colors = ["#0B2E6B", "#5B8FF9", "#8CB6FF"]

    fig, ax = plt.subplots(figsize=(10, 4.8), dpi=240)
    fig.patch.set_alpha(0)
    ax.set_facecolor("none")

    x = np.arange(n, dtype=float)
    width = 0.62
    bottom = np.zeros(n, dtype=float)
    segment_centers: list[list[float]] = [[] for _ in range(m)]

    for j in range(m):
        y = np.asarray(values[:, j], dtype=float)
        ax.bar(x, y, width=width, bottom=bottom, color=colors[j % len(colors)], edgecolor="none", zorder=2)

        rgba = to_rgba(colors[j % len(colors)])
        lum = 0.2126 * rgba[0] + 0.7152 * rgba[1] + 0.0722 * rgba[2]
        txt_color = "#ffffff" if lum < 0.50 else "#2f2f2f"

        for i in range(n):
            v = float(y[i])
            if not np.isfinite(v) or abs(v) < 1e-12:
                segment_centers[j].append(float("nan"))
                continue
            yc = float(bottom[i]) + v / 2.0
            segment_centers[j].append(yc)
            ax.text(
                float(x[i]),
                yc,
                _fmt_int(v),
                ha="center",
                va="center",
                fontsize=9.0,
                color=txt_color,
                zorder=4,
            )
        bottom = bottom + np.nan_to_num(y, nan=0.0)

    totals = bottom.copy()
    total_label_tops: list[float] = []
    for i, total in enumerate(totals):
        if not np.isfinite(total):
            continue
        y_label = float(total) + max(abs(float(total)) * 0.02, 0.6)
        total_label_tops.append(y_label)
        ax.text(
            float(x[i]),
            y_label,
            _fmt_int(total),
            ha="center",
            va="bottom",
            fontsize=10.0,
            fontweight="bold" if i == n - 1 else "normal",
            color="#2f2f2f",
            zorder=5,
            clip_on=False,
        )

    # Brackets/deltas no topo para pares consecutivos, alinhados.
    if n >= 2:
        abs_max = float(np.nanmax(np.abs(totals))) if np.isfinite(np.nanmax(np.abs(totals))) else 0.0
        offset_y = max(abs_max * 0.10, 1.0)
        bracket_h = max(abs_max * 0.035, 0.8)
        top_labels_max = max(total_label_tops) if total_label_tops else float(np.nanmax(totals))
        top_base = float(top_labels_max) + max(abs_max * 0.20, 2.0)
        max_text_y: float | None = None

        for i in range(1, n):
            prev = float(totals[i - 1])
            curr = float(totals[i])
            if not np.isfinite(prev) or not np.isfinite(curr) or prev == 0:
                continue
            pct = (curr / prev - 1.0) * 100.0
            label = f"{pct:+.1f}%".replace(".", ",")

            x1 = float(x[i - 1])
            x2 = float(x[i])
            y_anchor = top_base
            ax.plot(
                [x1, x1, x2, x2],
                [y_anchor, y_anchor + bracket_h, y_anchor + bracket_h, y_anchor],
                color="#2f2f2f",
                linewidth=1.2,
                solid_capstyle="round",
                zorder=4,
            )
            text_y = y_anchor + bracket_h + offset_y * 0.25
            ax.text(
                (x1 + x2) / 2.0,
                text_y,
                label,
                ha="center",
                va="bottom",
                fontsize=9.0,
                color="#2f2f2f",
                zorder=5,
            )
            max_text_y = text_y if max_text_y is None else max(max_text_y, text_y)

        if max_text_y is not None:
            ymin, ymax = ax.get_ylim()
            ax.set_ylim(ymin, max(ymax, max_text_y + offset_y * 1.4))

    # Legenda inline à esquerda.
    x_leg = float(x.min()) - 0.95
    for j, name in enumerate(series_names):
        y_ref = segment_centers[j][-1] if segment_centers[j] else float("nan")
        if not np.isfinite(y_ref):
            for yc in segment_centers[j]:
                if np.isfinite(yc):
                    y_ref = yc
                    break
        if not np.isfinite(y_ref):
            continue
        ax.scatter([x_leg], [float(y_ref)], s=90.0, marker="s", color=colors[j % len(colors)], edgecolors="none", zorder=6)
        ax.text(
            x_leg + 0.12,
            float(y_ref),
            _wrap_words(str(name), max_line_len=16),
            ha="left",
            va="center",
            fontsize=9.0,
            color="#2f2f2f",
            zorder=6,
            clip_on=False,
        )

    ax.set_xlim(float(x.min()) - 1.25, float(x.max()) + 0.65)
    ax.set_xticks(x)
    ax.set_xticklabels(xlabels, fontsize=10.0)
    ax.tick_params(axis="x", bottom=False, pad=8)
    ax.set_yticks([])
    for s in ("left", "right", "top", "bottom"):
        ax.spines[s].set_visible(False)
    ax.tick_params(axis="y", left=False, labelleft=False)
    ax.grid(False)
    ax.margins(x=0.05, y=0.12)

    fig.tight_layout(pad=0.2)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    fig.savefig(output_path, dpi=450, transparent=True, bbox_inches="tight", pad_inches=0.05)
    close_figure(fig)


def _plot_efficiency_index(
    *,
    xlabels: list[str],
    values: list[float],  # fraction values, e.g. 0.377
    output_path: Path,
) -> None:
    import matplotlib.pyplot as plt

    vals_pct = np.asarray(values, dtype=float) * 100.0
    x = np.arange(len(vals_pct), dtype=float)

    fig, ax = plt.subplots(figsize=(10, 4.8), dpi=240)
    fig.patch.set_alpha(0)
    ax.set_facecolor("none")

    colors = ["#123a7a"] * len(vals_pct)
    bars = ax.bar(x, vals_pct, width=0.62, color=colors, edgecolor="none", zorder=2)

    for i, (rect, v) in enumerate(zip(bars, vals_pct)):
        ax.text(
            rect.get_x() + rect.get_width() / 2,
            rect.get_height(),
            f"{v:.1f}%".replace(".", ","),
            ha="center",
            va="bottom",
            fontsize=10.0,
            fontweight="bold" if i == len(vals_pct) - 1 else "normal",
            color="#2f2f2f",
            zorder=4,
            clip_on=False,
        )

    # Bracket em p.p. entre primeiro e último ponto.
    if len(vals_pct) >= 2:
        first = float(vals_pct[0])
        last = float(vals_pct[-1])
        delta_pp = last - first
        abs_max = float(np.nanmax(np.abs(vals_pct))) if np.isfinite(np.nanmax(np.abs(vals_pct))) else 0.0
        offset_y = max(abs_max * 0.10, 0.6)
        bracket_h = max(abs_max * 0.03, 0.6)
        y_anchor = float(np.nanmax(vals_pct)) + max(abs_max * 0.20, 1.2)

        x1 = float(x[0])
        x2 = float(x[-1])
        ax.plot(
            [x1, x1, x2, x2],
            [y_anchor, y_anchor + bracket_h, y_anchor + bracket_h, y_anchor],
            color="#2f2f2f",
            linewidth=1.2,
            solid_capstyle="round",
            zorder=4,
        )
        label = f"{delta_pp:+.1f} p.p.".replace(".", ",")
        text_y = y_anchor + bracket_h + offset_y * 0.25
        ax.text(
            (x1 + x2) / 2.0,
            text_y,
            label,
            ha="center",
            va="bottom",
            fontsize=9.0,
            color="#2f2f2f",
            zorder=5,
        )
        ymin, ymax = ax.get_ylim()
        ax.set_ylim(ymin, max(ymax, text_y + offset_y * 1.2))

    ax.set_xticks(x)
    ax.set_xticklabels(xlabels, fontsize=10.0)
    ax.tick_params(axis="x", bottom=False, pad=8)
    ax.set_yticks([])
    for s in ("left", "right", "top", "bottom"):
        ax.spines[s].set_visible(False)
    ax.grid(False)
    ax.margins(x=0.05, y=0.20)

    fig.tight_layout(pad=0.2)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    fig.savefig(output_path, dpi=450, transparent=True, bbox_inches="tight", pad_inches=0.05)
    close_figure(fig)


def generate_slide11_charts(*, xlsx_path: Path, output_dir: Path) -> list[Path]:
    """Slide 11: despesas (tri + 9M) e índice de eficiência."""

    output_dir.mkdir(parents=True, exist_ok=True)

    wb = load_workbook(filename=xlsx_path, data_only=True)
    expenses_sheet_name = "Tabelas"
    index_sheet_name = "slide_11"
    if expenses_sheet_name not in wb.sheetnames:
        raise ValueError(f"Aba não encontrada: {expenses_sheet_name!r}. Disponíveis: {wb.sheetnames}")
    if index_sheet_name not in wb.sheetnames:
        raise ValueError(f"Aba não encontrada: {index_sheet_name!r}. Disponíveis: {wb.sheetnames}")
    ws_expenses = wb[expenses_sheet_name]
    ws_index = wb[index_sheet_name]

    tri_labels, series, tri_values = _read_named_series_rows(
        ws_expenses,
        xlabels_range="D33:F33",
        series_specs=[
            ("Pessoal", "D35:F35"),
            ("Administrativas", "D39:F39"),
            ("Depreciação e Amortização", "D45:F45"),
        ],
    )

    nm_labels, _, nm_values = _read_named_series_rows(
        ws_expenses,
        xlabels_range="G33:H33",
        series_specs=[
            ("Pessoal", "G35:H35"),
            ("Administrativas", "G39:H39"),
            ("Depreciação e Amortização", "G45:H45"),
        ],
    )

    generated: list[Path] = []

    out22 = output_dir / "11_despesas_pessoal_adm_trimestres.png"
    _plot_stacked_expenses(
        xlabels=tri_labels,
        series_names=series,
        values=tri_values,
        output_path=out22,
    )
    generated.append(out22)

    out23 = output_dir / "11_despesas_pessoal_adm_9m.png"
    _plot_stacked_expenses(
        xlabels=nm_labels,
        series_names=series,
        values=nm_values,
        output_path=out23,
    )
    generated.append(out23)

    idx_labels = [("" if v is None else str(v)).strip() for v in _read_range_row(ws_index, "K3:O3")]
    idx_values = to_float_list(_read_range_row(ws_index, "K4:O4"))
    out24 = output_dir / "11_indice_eficiencia.png"
    _plot_efficiency_index(
        xlabels=idx_labels,
        values=idx_values,
        output_path=out24,
    )
    generated.append(out24)

    return generated


if __name__ == "__main__":
    xlsx = Path("testing.xlsx")
    out = Path(".")
    if xlsx.exists():
        files = generate_slide11_charts(xlsx_path=xlsx, output_dir=out)
        print(f"Gerados: {files}")
    else:
        print(f"Arquivo {xlsx} não encontrado")
