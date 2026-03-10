from __future__ import annotations

from pathlib import Path

import numpy as np
from openpyxl import load_workbook
from openpyxl.utils.cell import range_boundaries

from utils.charts_common import close_figure, to_float_list


def _read_range_row(ws, cell_range: str) -> list[object]:
    min_col, min_row, max_col, max_row = range_boundaries(cell_range)
    out: list[object] = []
    for r in range(min_row, max_row + 1):
        for c in range(min_col, max_col + 1):
            out.append(ws.cell(row=r, column=c).value)
    return out


def _read_range_col(ws, cell_range: str) -> list[object]:
    min_col, min_row, max_col, max_row = range_boundaries(cell_range)
    out: list[object] = []
    for r in range(min_row, max_row + 1):
        for c in range(min_col, max_col + 1):
            out.append(ws.cell(row=r, column=c).value)
    return out


def _text_color_for_bg_rgba(rgba) -> str:
    r, g, b = float(rgba[0]), float(rgba[1]), float(rgba[2])
    lum = 0.2126 * r + 0.7152 * g + 0.0722 * b
    return "#ffffff" if lum < 0.50 else "#2f2f2f"


def _fmt_num(v: float, *, decimals: int = 1) -> str:
    return f"{float(v):.{int(decimals)}f}".replace(".", ",")


def _fmt_share(v: float, *, decimals: int = 0) -> str:
    return f"{float(v):.{int(decimals)}f}%".replace(".", ",")


def _plot_slide12_stacked(
    *,
    xlabels: list[str],
    series_names: list[str],
    values: np.ndarray,  # [n_bars, n_series]
    output_path: Path,
) -> None:
    import matplotlib.pyplot as plt
    from matplotlib.colors import to_rgba

    n, m = values.shape
    if n != 3:
        raise ValueError(f"Esperado 3 barras para slide_12; recebido={n}")
    if m == 0:
        raise ValueError("Sem séries para plotar")

    # Paleta solicitada: roxo, cinza, azul, verde, amarelo, rosa claro, rosa escuro.
    colors = [
        "#6F42C1",  # roxo
        "#9AA0A6",  # cinza
        "#1F77B4",  # azul
        "#2CA02C",  # verde
        "#F2C94C",  # amarelo
        "#F8BBD0",  # rosa claro
        "#C2185B",  # rosa escuro
        "#8E44AD",  # variação de roxo para série extra
    ]

    # Force the series containing the highest value to dark blue.
    try:
        max_flat_idx = int(np.nanargmax(values))
        _row_max, col_max = np.unravel_index(max_flat_idx, values.shape)
        colors[col_max % len(colors)] = "#123a7a"
    except Exception:
        pass

    fig, ax = plt.subplots(figsize=(10.6, 5.2), dpi=240)
    fig.patch.set_alpha(0)
    ax.set_facecolor("none")

    x = np.arange(n, dtype=float)
    width = 0.62

    bottom = np.zeros(n, dtype=float)
    segment_bottoms = np.zeros_like(values, dtype=float)
    for j in range(m):
        y = np.asarray(values[:, j], dtype=float)
        segment_bottoms[:, j] = bottom
        ax.bar(
            x,
            y,
            width=width,
            bottom=bottom,
            color=colors[j % len(colors)],
            edgecolor="none",
            zorder=2,
        )
        bottom = bottom + np.nan_to_num(y, nan=0.0)

    totals = np.nansum(values, axis=1).astype(float)

    # Totais no topo.
    total_label_tops: list[float] = []
    for i, total in enumerate(totals):
        if not np.isfinite(total):
            continue
        y_label = float(total) + max(abs(float(total)) * 0.02, 0.45)
        total_label_tops.append(y_label)
        ax.text(
            float(x[i]),
            y_label,
            _fmt_num(total, decimals=1),
            ha="center",
            va="bottom",
            fontsize=10.0,
            fontweight="bold" if i == n - 1 else "normal",
            color="#2f2f2f",
            zorder=5,
            clip_on=False,
        )

    # Mostrar rótulos para todos os segmentos.
    # Top 2: valor + participação no total. Demais: valor com fonte proporcional.
    for i in range(n):
        total = float(totals[i])
        if not np.isfinite(total) or abs(total) < 1e-12:
            continue

        col = values[i, :]
        top2_idx = set(np.argsort(col)[-2:].tolist())  # dois maiores
        for j in range(m):
            v = float(col[j])
            if not np.isfinite(v) or v <= 0:
                continue
            share = (v / total) * 100.0
            yc = float(segment_bottoms[i, j]) + v / 2.0
            rgba = to_rgba(colors[j % len(colors)])
            txt_color = _text_color_for_bg_rgba(rgba)
            is_top2 = j in top2_idx
            if is_top2:
                label = f"{_fmt_num(v, decimals=1)} ({_fmt_share(share, decimals=0)})"
                fontsize = 8.6
                fw = "bold"
                y_adj = 0.0
            else:
                # Segments menores também exibem valor; fonte acompanha a relevância.
                label = _fmt_num(v, decimals=1)
                fontsize = float(np.clip(6.6 + share * 0.05, 6.6, 8.0))
                fw = "normal"
                y_adj = 0.0
            ax.text(
                float(x[i]),
                yc + y_adj,
                label,
                ha="center",
                va="center",
                fontsize=fontsize,
                color=txt_color,
                fontweight=fw,
                zorder=6,
            )

    # Brackets/deltas no topo (pares consecutivos), alinhados.
    abs_max = float(np.nanmax(np.abs(totals))) if np.isfinite(np.nanmax(np.abs(totals))) else 0.0
    offset_y = max(abs_max * 0.10, 0.8)
    bracket_h = max(abs_max * 0.03, 0.6)
    top_labels_max = max(total_label_tops) if total_label_tops else float(np.nanmax(totals))
    top_base = float(top_labels_max) + max(abs_max * 0.20, 1.4)
    max_text_y: float | None = None

    for i in range(1, n):
        prev = float(totals[i - 1])
        curr = float(totals[i])
        if not np.isfinite(prev) or not np.isfinite(curr) or prev == 0:
            continue
        pct = (curr / prev - 1.0) * 100.0
        label = f"{pct:+.1f}%".replace(".", ",")

        x1 = float(x[i - 1])
        x2 = float(x[i])
        y_anchor = top_base
        ax.plot(
            [x1, x1, x2, x2],
            [y_anchor, y_anchor + bracket_h, y_anchor + bracket_h, y_anchor],
            color="#2f2f2f",
            linewidth=1.2,
            solid_capstyle="round",
            zorder=4,
        )
        text_y = y_anchor + bracket_h + offset_y * 0.25
        ax.text(
            (x1 + x2) / 2.0,
            text_y,
            label,
            ha="center",
            va="bottom",
            fontsize=9.0,
            color="#2f2f2f",
            zorder=5,
        )
        max_text_y = text_y if max_text_y is None else max(max_text_y, text_y)

    if max_text_y is not None:
        ymin, ymax = ax.get_ylim()
        ax.set_ylim(ymin, max(ymax, max_text_y + offset_y * 1.3))

    # Rótulos de séries ao lado esquerdo do gráfico (sem caixas), com linhas-guia.
    # Base: posiciona pelos centros dos segmentos da primeira barra.
    i_ref = 0
    centers: list[tuple[int, float]] = []
    for j in range(m):
        v = float(values[i_ref, j])
        if not np.isfinite(v) or abs(v) < 1e-12:
            continue
        yc = float(segment_bottoms[i_ref, j]) + v / 2.0
        centers.append((j, yc))

    centers.sort(key=lambda t: t[1])
    if centers:
        y_min = centers[0][1]
        y_max = centers[-1][1]
        span = max(0.8, (y_max - y_min))
        min_sep = max(0.35, span / 18.0)

        adjusted: list[tuple[int, float, float]] = []  # (series_idx, y_src, y_lbl)
        last_y = -1e18
        for j, y_src in centers:
            y_lbl = max(y_src, last_y + min_sep)
            adjusted.append((j, y_src, y_lbl))
            last_y = y_lbl

        x_anchor = float(x[i_ref]) - width / 2.0
        x_text = x_anchor - 0.58

        for j, y_src, y_lbl in adjusted:
            ax.plot(
                [x_anchor - 0.03, x_text + 0.06],
                [y_src, y_lbl],
                color="#7a7a7a",
                linewidth=0.9,
                zorder=6,
            )
            ax.text(
                x_text,
                y_lbl,
                str(series_names[j]),
                ha="right",
                va="center",
                fontsize=8.1,
                color="#2f2f2f",
                zorder=7,
                clip_on=False,
            )

    ax.set_xlim(float(x.min()) - 1.45, float(x.max()) + 0.65)
    ax.set_xticks(x)
    ax.set_xticklabels(xlabels, fontsize=10.0)
    ax.tick_params(axis="x", bottom=False, pad=8)
    ax.set_yticks([])
    for s in ("left", "right", "top", "bottom"):
        ax.spines[s].set_visible(False)
    ax.grid(False)
    ax.margins(x=0.04, y=0.12)

    fig.tight_layout(pad=0.3)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    fig.savefig(output_path, dpi=450, transparent=True, bbox_inches="tight", pad_inches=0.06)
    close_figure(fig)


def generate_slide12_charts(*, xlsx_path: Path, output_dir: Path) -> list[Path]:
    """Slide 12: gráfico empilhado (B3:E11) com top2 de participação, totais e brackets."""

    output_dir.mkdir(parents=True, exist_ok=True)

    wb = load_workbook(filename=xlsx_path, data_only=True)
    sheet_name = "slide_12"
    if sheet_name not in wb.sheetnames:
        raise ValueError(f"Aba não encontrada: {sheet_name!r}. Disponíveis: {wb.sheetnames}")
    ws = wb[sheet_name]

    xlabels = [("" if v is None else str(v)).strip() for v in _read_range_row(ws, "C3:E3")]
    series_names = [("" if v is None else str(v)).strip() for v in _read_range_col(ws, "B4:B11")]

    raw_rows: list[list[float]] = []
    for r in range(4, 12):
        raw_rows.append(to_float_list(_read_range_row(ws, f"C{r}:E{r}")))
    # raw_rows: [n_series, n_bars] -> transpose to [n_bars, n_series]
    values = np.asarray(raw_rows, dtype=float).T

    out25 = output_dir / "25_slide12_composicao.png"
    _plot_slide12_stacked(
        xlabels=xlabels,
        series_names=series_names,
        values=values,
        output_path=out25,
    )
    return [out25]


if __name__ == "__main__":
    xlsx = Path("testing.xlsx")
    out = Path(".")
    if xlsx.exists():
        files = generate_slide12_charts(xlsx_path=xlsx, output_dir=out)
        print(f"Gerados: {files}")
    else:
        print(f"Arquivo {xlsx} não encontrado")
