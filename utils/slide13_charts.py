from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path

import numpy as np
from openpyxl import load_workbook
from openpyxl.utils.cell import range_boundaries

from utils.charts_common import close_figure, to_float_list


SLIDE13_PALETTE = (
    "#123A7A",
    "#5B8FF9",
    "#9AA0A6",
    "#2CA02C",
    "#F2C94C",
    "#C2185B",
)

SLIDE13_ATACADO_PALETTE = (
    "#C2185B",
    "#9AA0A6",
    "#F8BBD0",
)


@dataclass(frozen=True)
class SeriesRangeSpec:
    name: str
    ranges: tuple[str, ...]


@dataclass(frozen=True)
class ChartSourceSpec:
    output_name: str
    labels_range: str
    series: tuple[SeriesRangeSpec, ...]
    palette: tuple[str, ...] | None = None


SLIDE13_CHARTS: tuple[ChartSourceSpec, ...] = (
    ChartSourceSpec(
        output_name="13_varejo_produtos_entrada.png",
        labels_range="D12:F12",
        series=(
            SeriesRangeSpec("Leves e Usados", ("D16:F16",)),
            SeriesRangeSpec("Motos e Novos", ("D18:F18",)),
            SeriesRangeSpec("Pesados", ("D17:F17",)),
            SeriesRangeSpec("Solar", ("D19:F19",)),
            SeriesRangeSpec("Outros", ("D20:F20", "D21:F21")),
        ),
    ),
    ChartSourceSpec(
        output_name="13_varejo_relacional.png",
        labels_range="D12:F12",
        series=(
            SeriesRangeSpec("Crédito Pessoal", ("D25:F25",)),
            SeriesRangeSpec("Cartões", ("D24:F24",)),
            SeriesRangeSpec("EGV", ("D23:F23",)),
        ),
    ),
    ChartSourceSpec(
        output_name="13_atacado.png",
        labels_range="D115:F115",
        series=(
            SeriesRangeSpec("PMEs", ("D119:F119",)),
            SeriesRangeSpec("Corporate", ("D117:F117",)),
            SeriesRangeSpec("Large e IF", ("D118:F118",)),
        ),
        palette=SLIDE13_ATACADO_PALETTE,
    ),
)


def _read_range_row(ws, cell_range: str) -> list[object]:
    min_col, min_row, max_col, max_row = range_boundaries(cell_range)
    out: list[object] = []
    for r in range(min_row, max_row + 1):
        for c in range(min_col, max_col + 1):
            out.append(ws.cell(row=r, column=c).value)
    return out


def _text_color_for_bg_rgba(rgba) -> str:
    r, g, b = float(rgba[0]), float(rgba[1]), float(rgba[2])
    lum = 0.2126 * r + 0.7152 * g + 0.0722 * b
    return "#ffffff" if lum < 0.50 else "#2f2f2f"


def _fmt_num(v: float) -> str:
    value = float(v)
    decimals = 2 if 0 < abs(value) < 0.1 else 1
    return f"{value:.{decimals}f}".replace(".", ",")


def _fmt_pct(v: float) -> str:
    return f"{float(v):.0f}%".replace(".", ",")


def _stack_order_for_bar(values_row: np.ndarray) -> tuple[int, ...]:
    row = np.asarray(values_row, dtype=float)
    return tuple(
        sorted(
            range(len(row)),
            key=lambda idx: (
                1 if not np.isfinite(float(row[idx])) else 0,
                -float(row[idx]) if np.isfinite(float(row[idx])) else 0.0,
                idx,
            ),
        )
    )


def _share_label_indices(values_row: np.ndarray) -> set[int]:
    ordered_positive = [
        idx
        for idx in _stack_order_for_bar(values_row)
        if np.isfinite(float(values_row[idx])) and float(values_row[idx]) > 0
    ]
    if len(ordered_positive) >= 4:
        top_n = 3
    elif len(ordered_positive) >= 3:
        top_n = 2
    else:
        top_n = len(ordered_positive)
    return set(ordered_positive[:top_n])


def _sum_ranges(ws, ranges: tuple[str, ...]) -> np.ndarray:
    total: np.ndarray | None = None
    for cell_range in ranges:
        row = np.asarray(to_float_list(_read_range_row(ws, cell_range)), dtype=float)
        if total is None:
            total = np.zeros_like(row, dtype=float)
        if row.shape != total.shape:
            raise ValueError(f"Ranges incompatíveis para soma: {ranges}")
        total = total + np.nan_to_num(row, nan=0.0)
    if total is None:
        raise ValueError("Nenhum range informado para soma")
    return total


def _plot_slide13_breakdown(
    *,
    xlabels: list[str],
    series_names: list[str],
    values: np.ndarray,  # [n_bars, n_series]
    output_path: Path,
    palette: tuple[str, ...] = SLIDE13_PALETTE,
    figure_size: tuple[float, float] = (7.4, 6.0),
    bar_slot: float = 0.24,
    bar_width: float = 0.20,
) -> None:
    import matplotlib.pyplot as plt
    from matplotlib.colors import to_rgba

    values_arr = np.asarray(values, dtype=float)
    n, m = values_arr.shape
    if n == 0 or m == 0:
        raise ValueError("Slide 13 sem dados para plotar")

    fig, ax = plt.subplots(figsize=figure_size, dpi=240)
    fig.patch.set_alpha(0)
    ax.set_facecolor("none")

    x = np.arange(n, dtype=float) * float(bar_slot)
    width = float(bar_width)

    segment_centers = np.full_like(values_arr, np.nan, dtype=float)
    totals = np.zeros(n, dtype=float)
    orders = tuple(_stack_order_for_bar(values_arr[i, :]) for i in range(n))

    for i in range(n):
        bottom = 0.0
        for idx in orders[i]:
            value = float(values_arr[i, idx])
            if not np.isfinite(value) or value <= 0:
                continue
            color = palette[idx % len(palette)]
            ax.bar(
                float(x[i]),
                value,
                width=width,
                bottom=bottom,
                color=color,
                edgecolor="none",
                zorder=2,
            )
            segment_centers[i, idx] = bottom + value / 2.0
            bottom += value
        totals[i] = bottom

    finite_totals = totals[np.isfinite(totals)]
    abs_max = float(np.nanmax(np.abs(finite_totals))) if finite_totals.size else 0.0
    total_label_gap = max(abs_max * 0.02, 0.15)

    total_label_tops: list[float] = []
    for i, total in enumerate(totals):
        if not np.isfinite(total):
            continue
        y_label = float(total) + total_label_gap
        total_label_tops.append(y_label)
        ax.text(
            float(x[i]),
            y_label,
            _fmt_num(total),
            ha="center",
            va="bottom",
            fontsize=10.0,
            fontweight="bold" if i == n - 1 else "normal",
            color="#2f2f2f",
            zorder=5,
            clip_on=False,
        )

    for i in range(n):
        total = float(totals[i])
        if not np.isfinite(total) or total <= 0:
            continue

        pct_indices = _share_label_indices(values_arr[i, :])
        for idx in orders[i]:
            value = float(values_arr[i, idx])
            if not np.isfinite(value) or value <= 0:
                continue
            share = (value / total) * 100.0
            label = _fmt_num(value)
            if idx in pct_indices:
                label = f"{label} ({_fmt_pct(share)})"

            rgba = to_rgba(palette[idx % len(palette)])
            txt_color = _text_color_for_bg_rgba(rgba)
            ax.text(
                float(x[i]),
                float(segment_centers[i, idx]),
                label,
                ha="center",
                va="center",
                fontsize=8.6 if idx in pct_indices else 7.8,
                color=txt_color,
                fontweight="bold" if idx in pct_indices else "normal",
                zorder=6,
            )

    if n >= 2:
        offset_y = max(abs_max * 0.08, 0.30)
        bracket_h = max(abs_max * 0.03, 0.18)
        top_labels_max = max(total_label_tops) if total_label_tops else float(np.nanmax(totals))
        top_base = float(top_labels_max) + max(abs_max * 0.12, 0.55)
        max_text_y: float | None = None

        for i in range(1, n):
            prev = float(totals[i - 1])
            curr = float(totals[i])
            if not np.isfinite(prev) or not np.isfinite(curr) or prev == 0:
                continue
            pct = (curr / prev - 1.0) * 100.0
            label = f"{pct:+.1f}%".replace(".", ",")

            x1 = float(x[i - 1])
            x2 = float(x[i])
            y_anchor = top_base
            ax.plot(
                [x1, x1, x2, x2],
                [y_anchor, y_anchor + bracket_h, y_anchor + bracket_h, y_anchor],
                color="#2f2f2f",
                linewidth=1.2,
                solid_capstyle="round",
                zorder=4,
            )
            text_y = y_anchor + bracket_h + offset_y * 0.25
            ax.text(
                (x1 + x2) / 2.0,
                text_y,
                label,
                ha="center",
                va="bottom",
                fontsize=9.0,
                color="#2f2f2f",
                zorder=5,
            )
            max_text_y = text_y if max_text_y is None else max(max_text_y, text_y)

        if max_text_y is not None:
            ymin, ymax = ax.get_ylim()
            ax.set_ylim(ymin, max(ymax, max_text_y + offset_y * 0.9))

    x_text = float(x[0]) - width / 2.0 - 0.06
    for idx in orders[0]:
        y_ref = float(segment_centers[0, idx])
        if not np.isfinite(y_ref):
            for row_idx in range(n):
                candidate = float(segment_centers[row_idx, idx])
                if np.isfinite(candidate):
                    y_ref = candidate
                    break
        if not np.isfinite(y_ref):
            continue
        ax.text(
            x_text,
            y_ref,
            str(series_names[idx]),
            ha="right",
            va="center",
            fontsize=8.4,
            color="#2f2f2f",
            zorder=7,
            clip_on=False,
        )

    ax.set_xlim(float(x.min()) - 0.72, float(x.max()) + 0.28)
    ax.set_xticks(x)
    ax.set_xticklabels(xlabels, fontsize=10.0)
    ax.tick_params(axis="x", bottom=False, pad=8)
    ax.set_yticks([])
    for spine in ("left", "right", "top", "bottom"):
        ax.spines[spine].set_visible(False)
    ax.grid(False)
    ax.margins(x=0.04, y=0.12)

    fig.tight_layout(pad=0.3)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    fig.savefig(output_path, dpi=450, transparent=True, bbox_inches="tight", pad_inches=0.06)
    close_figure(fig)


def generate_slide13_charts(*, xlsx_path: Path, output_dir: Path) -> list[Path]:
    """Slide 13: três gráficos empilhados com base na aba Carteira."""

    output_dir.mkdir(parents=True, exist_ok=True)

    wb = load_workbook(filename=xlsx_path, data_only=True)
    sheet_name = "Carteira"
    if sheet_name not in wb.sheetnames:
        raise ValueError(f"Aba não encontrada: {sheet_name!r}. Disponíveis: {wb.sheetnames}")
    ws = wb[sheet_name]

    generated: list[Path] = []
    for chart_spec in SLIDE13_CHARTS:
        xlabels = [
            ("" if value is None else str(value)).strip()
            for value in _read_range_row(ws, chart_spec.labels_range)
        ]
        series_names = [series_spec.name for series_spec in chart_spec.series]
        raw_rows = [_sum_ranges(ws, series_spec.ranges) for series_spec in chart_spec.series]
        values = np.asarray(raw_rows, dtype=float).T / 1000.0

        output_path = output_dir / chart_spec.output_name
        _plot_slide13_breakdown(
            xlabels=xlabels,
            series_names=series_names,
            values=values,
            output_path=output_path,
            palette=chart_spec.palette or SLIDE13_PALETTE,
        )
        generated.append(output_path)

    return generated


if __name__ == "__main__":
    xlsx = Path("testing.xlsx")
    out = Path(".")
    if xlsx.exists():
        files = generate_slide13_charts(xlsx_path=xlsx, output_dir=out)
        print(f"Gerados: {files}")
    else:
        print(f"Arquivo {xlsx} não encontrado")
