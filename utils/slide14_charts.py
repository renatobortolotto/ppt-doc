from __future__ import annotations

from dataclasses import dataclass
import math
from pathlib import Path

import numpy as np
from openpyxl import load_workbook
from openpyxl.utils.cell import range_boundaries

from utils.charts_common import close_figure, to_float_list


STACKED_COLORS = ("#123A7A", "#5B8FF9")
PERCENT_BAR_COLOR = "#123A7A"
SLIDE14_FONT_SCALE = 1.5
SLIDE14_PERCENT_FIGSIZE = (6.4, 2.6)


@dataclass(frozen=True)
class StackedChartSpec:
    output_name: str
    labels_range: str
    values_ranges: tuple[str, ...]
    name_cells: tuple[str, ...]


@dataclass(frozen=True)
class PercentChartSpec:
    output_name: str
    labels_range: str
    values_range: str


STACKED_SPECS: tuple[StackedChartSpec, ...] = (
    StackedChartSpec(
        output_name="14_veiculos_empilhado_trimestres.png",
        labels_range="D3:F3",
        values_ranges=("D10:F10", "D11:F11"),
        name_cells=("C10", "C11"),
    ),
    StackedChartSpec(
        output_name="14_veiculos_empilhado_anos.png",
        labels_range="G3:H3",
        values_ranges=("G10:H10", "G11:H11"),
        name_cells=("C10", "C11"),
    ),
)

PERCENT_SPECS: tuple[PercentChartSpec, ...] = (
    PercentChartSpec(
        output_name="14_veiculos_percentual_trimestres.png",
        labels_range="D3:F3",
        values_range="D7:F7",
    ),
    PercentChartSpec(
        output_name="14_veiculos_percentual_anos.png",
        labels_range="G3:H3",
        values_range="G7:H7",
    ),
)


def _read_range_row(ws, cell_range: str) -> list[object]:
    min_col, min_row, max_col, max_row = range_boundaries(cell_range)
    out: list[object] = []
    for r in range(min_row, max_row + 1):
        for c in range(min_col, max_col + 1):
            out.append(ws.cell(row=r, column=c).value)
    return out


def _text_color_for_bg_rgba(rgba) -> str:
    r, g, b = float(rgba[0]), float(rgba[1]), float(rgba[2])
    lum = 0.2126 * r + 0.7152 * g + 0.0722 * b
    return "#ffffff" if lum < 0.50 else "#2f2f2f"


def _fmt_value(v: float) -> str:
    value = float(v)
    if abs(value - round(value)) < 1e-9:
        return f"{int(round(value))}"
    return f"{value:.1f}".replace(".", ",")


def _round_pct_half_down(v: float) -> int:
    value = float(v)
    sign = -1 if value < 0 else 1
    abs_value = abs(value)
    base = math.floor(abs_value)
    frac = abs_value - base
    if frac > 0.5:
        base += 1
    return sign * base


def _fmt_pct(v: float) -> str:
    return f"{_round_pct_half_down(v)}%"


def _fmt_pct_trunc(v: float) -> str:
    return _fmt_pct(v)


def _largest_segment_indices(values_row: np.ndarray) -> set[int]:
    row = np.asarray(values_row, dtype=float)
    positive_indices = [
        idx
        for idx, value in enumerate(row)
        if np.isfinite(float(value)) and float(value) > 0
    ]
    if not positive_indices:
        return set()

    max_value = max(float(row[idx]) for idx in positive_indices)
    return {idx for idx in positive_indices if float(row[idx]) == max_value}


def _resolve_veiculos_sheet_name(wb) -> str:
    for candidate in ("Veículos", "Veiculos"):
        if candidate in wb.sheetnames:
            return candidate
    raise ValueError(f"Aba não encontrada: 'Veículos'. Disponíveis: {wb.sheetnames}")


def _normalize_percent_values(values: list[float]) -> list[float]:
    arr = np.asarray(values, dtype=float)
    finite = arr[np.isfinite(arr)]
    if finite.size and float(np.nanmax(np.abs(finite))) > 1.0:
        arr = arr / 100.0
    return arr.tolist()


def _plot_stacked_veiculos(
    *,
    xlabels: list[str],
    series_names: list[str],
    values: np.ndarray,  # [n_bars, 2]
    output_path: Path,
) -> None:
    import matplotlib.pyplot as plt
    from matplotlib.colors import to_rgba

    n, m = values.shape
    if n == 0 or m == 0:
        raise ValueError("Sem dados para plotar")

    fig, ax = plt.subplots(figsize=(6.8, 4.9), dpi=240)
    fig.patch.set_alpha(0)
    ax.set_facecolor("none")

    if n <= 2:
        bar_slot = 0.22
        width = 0.18
    else:
        bar_slot = 0.24
        width = 0.20
    x = np.arange(n, dtype=float) * bar_slot
    bottom = np.zeros(n, dtype=float)
    segment_centers: list[list[float]] = [[] for _ in range(m)]
    total_label_tops: list[float] = []

    for j in range(m):
        y = np.asarray(values[:, j], dtype=float)
        ax.bar(
            x,
            y,
            width=width,
            bottom=bottom,
            color=STACKED_COLORS[j % len(STACKED_COLORS)],
            edgecolor="none",
            zorder=2,
        )

        rgba = to_rgba(STACKED_COLORS[j % len(STACKED_COLORS)])
        txt_color = _text_color_for_bg_rgba(rgba)

        for i in range(n):
            v = float(y[i])
            if not np.isfinite(v) or abs(v) < 1e-12:
                segment_centers[j].append(float("nan"))
                continue
            yc = float(bottom[i]) + v / 2.0
            segment_centers[j].append(yc)
            total_bar = float(np.nansum(values[i, :]))
            share = (v / total_bar) * 100.0 if np.isfinite(total_bar) and total_bar > 0 else float("nan")
            label = _fmt_value(v)
            if j in _largest_segment_indices(values[i, :]) and np.isfinite(share):
                label = f"{label} ({_fmt_pct(share)})"
            ax.text(
                float(x[i]),
                yc,
                label,
                ha="center",
                va="center",
                fontsize=8.6 * SLIDE14_FONT_SCALE,
                color=txt_color,
                zorder=4,
            )
        bottom = bottom + np.nan_to_num(y, nan=0.0)

    totals = bottom.copy()
    for i, total in enumerate(totals):
        if not np.isfinite(total):
            continue
        y_label = float(total) + max(abs(float(total)) * 0.03, 0.25)
        total_label_tops.append(y_label)
        ax.text(
            float(x[i]),
            y_label,
            _fmt_value(total),
            ha="center",
            va="bottom",
            fontsize=9.8 * SLIDE14_FONT_SCALE,
            fontweight="bold" if i == n - 1 else "normal",
            color="#2f2f2f",
            zorder=5,
            clip_on=False,
        )

    if n >= 2:
        abs_max = float(np.nanmax(np.abs(totals))) if np.isfinite(np.nanmax(np.abs(totals))) else 0.0
        offset_y = max(abs_max * 0.08, 0.20)
        bracket_h = max(abs_max * 0.03, 0.15)
        top_labels_max = max(total_label_tops) if total_label_tops else float(np.nanmax(totals))
        top_base = float(top_labels_max) + max(abs_max * 0.10, 0.35)
        max_text_y: float | None = None

        for i in range(1, n):
            prev = float(totals[i - 1])
            curr = float(totals[i])
            if not np.isfinite(prev) or not np.isfinite(curr) or prev == 0:
                continue
            pct = (curr / prev - 1.0) * 100.0
            label = f"{pct:+.1f}%".replace(".", ",")

            x1 = float(x[i - 1])
            x2 = float(x[i])
            y_anchor = top_base
            ax.plot(
                [x1, x1, x2, x2],
                [y_anchor, y_anchor + bracket_h, y_anchor + bracket_h, y_anchor],
                color="#2f2f2f",
                linewidth=1.2,
                solid_capstyle="round",
                zorder=4,
            )
            text_y = y_anchor + bracket_h + offset_y * 0.25
            ax.text(
                (x1 + x2) / 2.0,
                text_y,
                label,
                ha="center",
                va="bottom",
                fontsize=8.9 * SLIDE14_FONT_SCALE,
                color="#2f2f2f",
                zorder=5,
            )
            max_text_y = text_y if max_text_y is None else max(max_text_y, text_y)

        if max_text_y is not None:
            ymin, ymax = ax.get_ylim()
            ax.set_ylim(ymin, max(ymax, max_text_y + offset_y * 0.9))

    x_text = float(x[0]) - width / 2.0 - 0.05
    for j, name in enumerate(series_names):
        y_ref = segment_centers[j][0] if segment_centers[j] else float("nan")
        if not np.isfinite(y_ref):
            for yc in segment_centers[j]:
                if np.isfinite(yc):
                    y_ref = yc
                    break
        if not np.isfinite(y_ref):
            continue
        ax.text(
            x_text,
            float(y_ref),
            str(name),
            ha="right",
            va="center",
            fontsize=8.6 * SLIDE14_FONT_SCALE,
            color="#2f2f2f",
            zorder=6,
            clip_on=False,
        )

    ax.set_xlim(float(x.min()) - 0.72, float(x.max()) + 0.28)
    ax.set_xticks(x)
    ax.set_xticklabels(xlabels, fontsize=10.0 * SLIDE14_FONT_SCALE)
    ax.tick_params(axis="x", bottom=False, pad=8)
    ax.set_yticks([])
    for spine in ("left", "right", "top", "bottom"):
        ax.spines[spine].set_visible(False)
    ax.grid(False)
    ax.margins(x=0.04, y=0.12)

    fig.tight_layout(pad=0.25)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    fig.savefig(output_path, dpi=450, transparent=True, bbox_inches="tight", pad_inches=0.05)
    close_figure(fig)


def _plot_percent_bars(
    *,
    xlabels: list[str],
    values: list[float],  # fractions
    output_path: Path,
) -> None:
    import matplotlib.pyplot as plt

    vals = np.asarray(values, dtype=float) * 100.0
    x = np.arange(len(vals), dtype=float) * (0.22 if len(vals) <= 2 else 0.24)
    width = 0.18 if len(vals) <= 2 else 0.20

    fig, ax = plt.subplots(figsize=SLIDE14_PERCENT_FIGSIZE, dpi=240)
    fig.patch.set_alpha(0)
    ax.set_facecolor("none")

    bars = ax.bar(x, vals, width=width, color=PERCENT_BAR_COLOR, edgecolor="none", zorder=2)
    top_padding = max(float(np.nanmax(vals)) * 0.10, 3.0) if len(vals) else 3.0
    for i, (rect, v) in enumerate(zip(bars, vals)):
        ax.text(
            rect.get_x() + rect.get_width() / 2.0,
            rect.get_height() + max(abs(float(v)) * 0.03, 1.2),
            _fmt_pct_trunc(v),
            ha="center",
            va="bottom",
            fontsize=10.0 * SLIDE14_FONT_SCALE,
            fontweight="bold" if i == len(vals) - 1 else "normal",
            color="#2f2f2f",
            zorder=4,
            clip_on=False,
        )

    ymax = float(np.nanmax(vals)) + top_padding if len(vals) else top_padding
    ax.set_ylim(0.0, ymax)
    ax.set_xlim(float(x.min()) - 0.30, float(x.max()) + 0.30)
    ax.set_xticks(x)
    ax.set_xticklabels(xlabels, fontsize=10.0 * SLIDE14_FONT_SCALE)
    ax.tick_params(axis="x", bottom=False, pad=8)
    ax.set_yticks([])
    for spine in ("left", "right", "top", "bottom"):
        ax.spines[spine].set_visible(False)
    ax.grid(False)
    ax.margins(x=0.04, y=0.08)

    fig.tight_layout(pad=0.25)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    fig.savefig(output_path, dpi=450, transparent=True, bbox_inches="tight", pad_inches=0.05)
    close_figure(fig)


def generate_slide14_charts(*, xlsx_path: Path, output_dir: Path) -> list[Path]:
    """Slide 14: quatro gráficos a partir da aba Veículos."""

    output_dir.mkdir(parents=True, exist_ok=True)

    wb = load_workbook(filename=xlsx_path, data_only=True)
    sheet_name = _resolve_veiculos_sheet_name(wb)
    ws = wb[sheet_name]

    generated: list[Path] = []

    for chart_spec in STACKED_SPECS:
        xlabels = [
            ("" if value is None else str(value)).strip()
            for value in _read_range_row(ws, chart_spec.labels_range)
        ]
        series_names = [
            ("" if ws[cell].value is None else str(ws[cell].value)).strip()
            for cell in chart_spec.name_cells
        ]
        raw_rows = [
            to_float_list(_read_range_row(ws, values_range))
            for values_range in chart_spec.values_ranges
        ]
        values = np.asarray(raw_rows, dtype=float).T

        output_path = output_dir / chart_spec.output_name
        _plot_stacked_veiculos(
            xlabels=xlabels,
            series_names=series_names,
            values=values,
            output_path=output_path,
        )
        generated.append(output_path)

    for chart_spec in PERCENT_SPECS:
        xlabels = [
            ("" if value is None else str(value)).strip()
            for value in _read_range_row(ws, chart_spec.labels_range)
        ]
        values = _normalize_percent_values(to_float_list(_read_range_row(ws, chart_spec.values_range)))

        output_path = output_dir / chart_spec.output_name
        _plot_percent_bars(
            xlabels=xlabels,
            values=values,
            output_path=output_path,
        )
        generated.append(output_path)

    return generated


if __name__ == "__main__":
    xlsx = Path("testing.xlsx")
    out = Path(".")
    if xlsx.exists():
        files = generate_slide14_charts(xlsx_path=xlsx, output_dir=out)
        print(f"Gerados: {files}")
    else:
        print(f"Arquivo {xlsx} não encontrado")
