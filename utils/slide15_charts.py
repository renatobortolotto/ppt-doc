from __future__ import annotations

from pathlib import Path

import numpy as np
from openpyxl import load_workbook
from openpyxl.utils.cell import range_boundaries

from utils.charts_common import close_figure


SLIDE15_OUTPUT = "15_captacoes_trimestres.png"
SLIDE15_SERIES_ROWS = (5, 6, 10, 11, 12, 13, 14, 15)
SLIDE15_COLORS = (
    "#123A7A",
    "#5B8FF9",
    "#9AA0A6",
    "#2CA02C",
    "#F2C94C",
    "#C2185B",
    "#F8BBD0",
    "#8E44AD",
)
SLIDE15_BAR_SLOT = 0.36
SLIDE15_BAR_WIDTH = 0.30
SLIDE15_INNER_LABEL_FONT_SCALE = 1.0
SLIDE15_BADGE_PAD = 0.10


def _read_range_row(ws, cell_range: str) -> list[object]:
    min_col, min_row, max_col, max_row = range_boundaries(cell_range)
    out: list[object] = []
    for r in range(min_row, max_row + 1):
        for c in range(min_col, max_col + 1):
            out.append(ws.cell(row=r, column=c).value)
    return out


def _to_float_or_nan(v: object) -> float:
    if v is None:
        return float("nan")
    if isinstance(v, str):
        s = v.strip()
        if s == "":
            return float("nan")
        s = s.replace("%", "").replace(" ", "")
        s = s.replace(".", "").replace(",", ".") if ("," in s and "." in s) else s.replace(",", ".")
        try:
            return float(s)
        except Exception:
            return float("nan")
    try:
        return float(v)
    except Exception:
        return float("nan")


def _fmt_value(v: float) -> str:
    value = float(v)
    if abs(value - round(value)) < 1e-9:
        return f"{int(round(value))}"
    return f"{value:.1f}".replace(".", ",")


def _fmt_share_pct(v: float) -> str:
    return f"{float(v):.0f}%".replace(".", ",")


def _text_color_for_bg_rgba(rgba) -> str:
    r, g, b = float(rgba[0]), float(rgba[1]), float(rgba[2])
    lum = 0.2126 * r + 0.7152 * g + 0.0722 * b
    return "#ffffff" if lum < 0.50 else "#2f2f2f"


def _top_share_indices(values_row: np.ndarray, *, top_n: int = 2) -> set[int]:
    row = np.asarray(values_row, dtype=float)
    ordered_positive = [
        idx
        for idx in sorted(
            range(len(row)),
            key=lambda idx: (-float(row[idx]), idx) if np.isfinite(float(row[idx])) else (float("inf"), idx),
        )
        if np.isfinite(float(row[idx])) and float(row[idx]) > 0
    ]
    return set(ordered_positive[: max(0, int(top_n))])


def _should_render_zero_label(series_name: str) -> bool:
    return "fidc" in str(series_name).strip().lower()


def _resolve_captacoes_sheet_name(wb) -> str:
    for candidate in ("Captações", "Captacoes"):
        if candidate in wb.sheetnames:
            return candidate
    raise ValueError(f"Aba não encontrada: 'Captações'. Disponíveis: {wb.sheetnames}")


def _plot_stacked_captacoes(
    *,
    xlabels: list[str],
    series_names: list[str],
    values: np.ndarray,  # [n_bars, n_series]
    output_path: Path,
) -> None:
    import matplotlib.pyplot as plt
    from matplotlib.colors import to_rgba

    n, m = values.shape
    if n == 0 or m == 0:
        raise ValueError("Sem dados para plotar")

    fig, ax = plt.subplots(figsize=(10.2, 5.7), dpi=240)
    fig.patch.set_alpha(0)
    ax.set_facecolor("none")

    bar_slot = SLIDE15_BAR_SLOT
    width = SLIDE15_BAR_WIDTH
    x = np.arange(n, dtype=float) * bar_slot
    bottom = np.zeros(n, dtype=float)
    segment_centers: list[list[float]] = [[] for _ in range(m)]
    totals_per_bar = np.nansum(np.where(np.isfinite(values), values, 0.0), axis=1).astype(float)
    top_share_indices_per_bar = tuple(_top_share_indices(values[i, :], top_n=2) for i in range(n))

    for j in range(m):
        y = np.asarray(values[:, j], dtype=float)
        color = SLIDE15_COLORS[j % len(SLIDE15_COLORS)]
        ax.bar(
            x,
            y,
            width=width,
            bottom=bottom,
            color=color,
            edgecolor="none",
            zorder=2,
        )

        rgba = to_rgba(color)
        txt_color = _text_color_for_bg_rgba(rgba)

        for i in range(n):
            value = float(y[i])
            if not np.isfinite(value):
                segment_centers[j].append(float("nan"))
                continue
            total_bar = float(totals_per_bar[i])
            show_share = j in top_share_indices_per_bar[i] and np.isfinite(total_bar) and total_bar > 0

            if abs(value) < 1e-12:
                if not _should_render_zero_label(series_names[j]):
                    segment_centers[j].append(float("nan"))
                    continue
                # Keep zero-value labels centered on the band boundary where that
                # segment would sit instead of floating above the stack.
                yc = float(bottom[i])
                label = _fmt_value(0.0)
            else:
                yc = float(bottom[i]) + value / 2.0
                label = _fmt_value(value)

            if show_share:
                share = (max(value, 0.0) / total_bar) * 100.0
                label = f"{label}\n({_fmt_share_pct(share)})"

            segment_centers[j].append(yc)
            ax.text(
                float(x[i]),
                yc,
                label,
                ha="center",
                va="center",
                fontsize=8.5 * SLIDE15_INNER_LABEL_FONT_SCALE,
                color=txt_color,
                zorder=4,
                linespacing=0.95,
                bbox={
                    "facecolor": color,
                    "edgecolor": "none",
                    "boxstyle": f"round,pad={SLIDE15_BADGE_PAD}",
                },
                clip_on=False,
            )
        bottom = bottom + np.nan_to_num(y, nan=0.0)

    totals = bottom.copy()
    total_label_tops: list[float] = []
    for i, total in enumerate(totals):
        if not np.isfinite(total):
            continue
        y_label = float(total) + max(abs(float(total)) * 0.03, 0.35)
        total_label_tops.append(y_label)
        ax.text(
            float(x[i]),
            y_label,
            _fmt_value(total),
            ha="center",
            va="bottom",
            fontsize=9.8,
            fontweight="bold" if i == n - 1 else "normal",
            color="#2f2f2f",
            zorder=5,
            clip_on=False,
        )

    if n >= 2:
        abs_max = float(np.nanmax(np.abs(totals))) if np.isfinite(np.nanmax(np.abs(totals))) else 0.0
        offset_y = max(abs_max * 0.08, 0.20)
        bracket_h = max(abs_max * 0.03, 0.15)
        top_labels_max = max(total_label_tops) if total_label_tops else float(np.nanmax(totals))
        top_base = float(top_labels_max) + max(abs_max * 0.10, 0.35)
        max_text_y: float | None = None

        for i in range(1, n):
            prev = float(totals[i - 1])
            curr = float(totals[i])
            if not np.isfinite(prev) or not np.isfinite(curr) or prev == 0:
                continue
            pct = (curr / prev - 1.0) * 100.0
            label = f"{pct:+.1f}%".replace(".", ",")

            x1 = float(x[i - 1])
            x2 = float(x[i])
            y_anchor = top_base
            ax.plot(
                [x1, x1, x2, x2],
                [y_anchor, y_anchor + bracket_h, y_anchor + bracket_h, y_anchor],
                color="#2f2f2f",
                linewidth=1.2,
                solid_capstyle="round",
                zorder=4,
            )
            text_y = y_anchor + bracket_h + offset_y * 0.25
            ax.text(
                (x1 + x2) / 2.0,
                text_y,
                label,
                ha="center",
                va="bottom",
                fontsize=9.0,
                color="#2f2f2f",
                zorder=5,
            )
            max_text_y = text_y if max_text_y is None else max(max_text_y, text_y)

        if max_text_y is not None:
            ymin, ymax = ax.get_ylim()
            ax.set_ylim(ymin, max(ymax, max_text_y + offset_y * 0.9))

    max_name_len = max((len(str(name).strip()) for name in series_names), default=1)
    left_margin = max(1.45, 0.72 + max_name_len * 0.045)
    x_text = float(x[0]) - width / 2.0 - 0.24
    connector_start = x_text + 0.04
    connector_end = float(x[0]) - width / 2.0 - 0.04
    for j, name in enumerate(series_names):
        y_ref = segment_centers[j][0] if segment_centers[j] else float("nan")
        if not np.isfinite(y_ref):
            for yc in segment_centers[j]:
                if np.isfinite(yc):
                    y_ref = yc
                    break
        if not np.isfinite(y_ref):
            continue
        if connector_end > connector_start:
            ax.plot(
                [connector_start, connector_end],
                [float(y_ref), float(y_ref)],
                color=SLIDE15_COLORS[j % len(SLIDE15_COLORS)],
                linewidth=2.2,
                solid_capstyle="round",
                zorder=6,
                clip_on=False,
            )
        ax.text(
            x_text,
            float(y_ref),
            str(name),
            ha="right",
            va="center",
            fontsize=8.9,
            color="#2f2f2f",
            zorder=7,
            clip_on=False,
        )

    ax.set_xlim(float(x.min()) - (left_margin + 0.10), float(x.max()) + 0.42)
    ax.set_xticks(x)
    ax.set_xticklabels(xlabels, fontsize=10.0)
    ax.tick_params(axis="x", bottom=False, pad=8)
    ax.set_yticks([])
    for spine in ("left", "right", "top", "bottom"):
        ax.spines[spine].set_visible(False)
    ax.grid(False)
    ax.margins(x=0.04, y=0.12)

    fig.tight_layout(pad=0.25)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    fig.savefig(output_path, dpi=450, transparent=True, bbox_inches="tight", pad_inches=0.05)
    close_figure(fig)


def generate_slide15_charts(*, xlsx_path: Path, output_dir: Path) -> list[Path]:
    """Slide 15: gráfico empilhado trimestral a partir da aba Captações."""

    output_dir.mkdir(parents=True, exist_ok=True)

    wb = load_workbook(filename=xlsx_path, data_only=True)
    sheet_name = _resolve_captacoes_sheet_name(wb)
    ws = wb[sheet_name]

    xlabels = [("" if v is None else str(v)).strip() for v in _read_range_row(ws, "C3:E3")]
    series_names = [
        ("" if ws[f"B{row}"].value is None else str(ws[f"B{row}"].value)).strip()
        for row in SLIDE15_SERIES_ROWS
    ]
    raw_rows = [
        [_to_float_or_nan(v) for v in _read_range_row(ws, f"C{row}:E{row}")]
        for row in SLIDE15_SERIES_ROWS
    ]
    values = np.asarray(raw_rows, dtype=float).T

    output_path = output_dir / SLIDE15_OUTPUT
    _plot_stacked_captacoes(
        xlabels=xlabels,
        series_names=series_names,
        values=values,
        output_path=output_path,
    )
    return [output_path]


if __name__ == "__main__":
    xlsx = Path("testing.xlsx")
    out = Path(".")
    if xlsx.exists():
        files = generate_slide15_charts(xlsx_path=xlsx, output_dir=out)
        print(f"Gerados: {files}")
    else:
        print(f"Arquivo {xlsx} não encontrado")
