from __future__ import annotations

from pathlib import Path

import numpy as np
from openpyxl import load_workbook
from openpyxl.utils.cell import range_boundaries

from utils.charts_common import close_figure


SLIDE16_OUTPUT = "16_indice_basileia_trimestres.png"
SLIDE16_SHEET_CANDIDATES = ("Basiléia", "Basileia")
SLIDE16_SERIES = (
    ("Nível I Principal", 18, "#123A7A"),
    ("Nível I Complementar", 19, "#5B8FF9"),
    ("Nível II", 20, "#AFC8F5"),
)
SLIDE16_FIGSIZE = (7.4, 3.5)
SLIDE16_BAR_SLOT = 1.0
SLIDE16_BAR_WIDTH = 0.66


def _read_range_row(ws, cell_range: str) -> list[object]:
    min_col, min_row, max_col, max_row = range_boundaries(cell_range)
    out: list[object] = []
    for r in range(min_row, max_row + 1):
        for c in range(min_col, max_col + 1):
            out.append(ws.cell(row=r, column=c).value)
    return out


def _to_float_or_nan(v: object) -> float:
    if v is None:
        return float("nan")
    if isinstance(v, str):
        s = v.strip()
        if s == "":
            return float("nan")
        s = s.replace("%", "").replace(" ", "")
        s = s.replace(".", "").replace(",", ".") if ("," in s and "." in s) else s.replace(",", ".")
        try:
            return float(s)
        except Exception:
            return float("nan")
    try:
        return float(v)
    except Exception:
        return float("nan")


def _fmt_pct(v: float) -> str:
    return f"{float(v):.1f}%".replace(".", ",")


def _fmt_pp(v: float) -> str:
    return f"{float(v):+.1f} p.p.".replace(".", ",")


def _text_color_for_bg_rgba(rgba) -> str:
    r, g, b = float(rgba[0]), float(rgba[1]), float(rgba[2])
    lum = 0.2126 * r + 0.7152 * g + 0.0722 * b
    return "#ffffff" if lum < 0.50 else "#2f2f2f"


def _resolve_basileia_sheet_name(wb) -> str:
    for candidate in SLIDE16_SHEET_CANDIDATES:
        if candidate in wb.sheetnames:
            return candidate
    raise ValueError(f"Aba não encontrada: 'Basiléia'. Disponíveis: {wb.sheetnames}")


def _plot_stacked_basileia(
    *,
    xlabels: list[str],
    series_names: list[str],
    values_pct: np.ndarray,  # [n_bars, n_series]
    output_path: Path,
) -> None:
    import matplotlib.pyplot as plt
    from matplotlib.colors import to_rgba

    n, m = values_pct.shape
    if n == 0 or m == 0:
        raise ValueError("Sem dados para plotar")

    fig, ax = plt.subplots(figsize=SLIDE16_FIGSIZE, dpi=240)
    fig.patch.set_alpha(0)
    ax.set_facecolor("none")

    x = np.arange(n, dtype=float) * SLIDE16_BAR_SLOT
    width = SLIDE16_BAR_WIDTH
    bottom = np.zeros(n, dtype=float)
    segment_centers: list[list[float]] = [[] for _ in range(m)]
    colors = [color for _, _, color in SLIDE16_SERIES]

    for j in range(m):
        y = np.asarray(values_pct[:, j], dtype=float)
        color = colors[j % len(colors)]
        ax.bar(
            x,
            y,
            width=width,
            bottom=bottom,
            color=color,
            edgecolor="none",
            zorder=2,
        )

        txt_color = _text_color_for_bg_rgba(to_rgba(color))
        for i in range(n):
            value = float(y[i])
            if not np.isfinite(value):
                segment_centers[j].append(float("nan"))
                continue
            yc = float(bottom[i]) + value / 2.0
            segment_centers[j].append(yc)
            ax.text(
                float(x[i]),
                yc,
                _fmt_pct(value),
                ha="center",
                va="center",
                fontsize=8.8,
                color=txt_color,
                zorder=4,
                clip_on=False,
            )

        bottom = bottom + np.nan_to_num(y, nan=0.0)

    totals = bottom.copy()
    total_label_tops: list[float] = []
    for i, total in enumerate(totals):
        if not np.isfinite(total):
            continue
        y_label = float(total) + max(abs(float(total)) * 0.03, 0.18)
        total_label_tops.append(y_label)
        ax.text(
            float(x[i]),
            y_label,
            _fmt_pct(total),
            ha="center",
            va="bottom",
            fontsize=9.8,
            fontweight="bold" if i == n - 1 else "normal",
            color="#2f2f2f",
            zorder=5,
            clip_on=False,
        )

    if n >= 2:
        abs_max = float(np.nanmax(np.abs(totals))) if np.isfinite(np.nanmax(np.abs(totals))) else 0.0
        offset_y = max(abs_max * 0.05, 0.12)
        bracket_h = max(abs_max * 0.02, 0.08)
        top_labels_max = max(total_label_tops) if total_label_tops else float(np.nanmax(totals))
        top_base = float(top_labels_max) + max(abs_max * 0.06, 0.16)
        first_total = float(totals[0])
        last_total = float(totals[-1])
        if np.isfinite(first_total) and np.isfinite(last_total):
            x1 = float(x[0])
            x2 = float(x[-1])
            text_y = top_base + bracket_h + offset_y * 0.25
            ax.plot(
                [x1, x1, x2, x2],
                [top_base, top_base + bracket_h, top_base + bracket_h, top_base],
                color="#2f2f2f",
                linewidth=1.2,
                solid_capstyle="round",
                zorder=4,
            )
            ax.text(
                (x1 + x2) / 2.0,
                text_y,
                _fmt_pp(last_total - first_total),
                ha="center",
                va="bottom",
                fontsize=9.0,
                color="#2f2f2f",
                zorder=5,
            )
            ymin, ymax = ax.get_ylim()
            ax.set_ylim(ymin, max(ymax, text_y + offset_y * 0.9))

    max_name_len = max((len(str(name).strip()) for name in series_names), default=1)
    left_margin = max(1.8, 0.85 + max_name_len * 0.043)
    x_text = float(x[0]) - width / 2.0 - 0.20
    connector_start = x_text + 0.06
    connector_end = float(x[0]) - width / 2.0 - 0.05
    for j, name in enumerate(series_names):
        y_ref = segment_centers[j][0] if segment_centers[j] else float("nan")
        if not np.isfinite(y_ref):
            for yc in segment_centers[j]:
                if np.isfinite(yc):
                    y_ref = yc
                    break
        if not np.isfinite(y_ref):
            continue
        if connector_end > connector_start:
            ax.plot(
                [connector_start, connector_end],
                [float(y_ref), float(y_ref)],
                color=colors[j % len(colors)],
                linewidth=2.2,
                solid_capstyle="round",
                zorder=6,
                clip_on=False,
            )
        ax.text(
            x_text,
            float(y_ref),
            str(name),
            ha="right",
            va="center",
            fontsize=8.8,
            color="#2f2f2f",
            zorder=7,
            clip_on=False,
        )

    ax.axhline(0.0, color="#b5b5b5", linewidth=0.9, zorder=1)
    ax.set_xlim(float(x.min()) - (left_margin + 0.1), float(x.max()) + 0.36)
    ax.set_xticks(x)
    ax.set_xticklabels(xlabels, fontsize=10.0)
    ax.tick_params(axis="x", bottom=False, pad=8)
    ax.set_yticks([])
    for spine in ("left", "right", "top", "bottom"):
        ax.spines[spine].set_visible(False)
    ax.grid(False)
    ax.margins(x=0.04, y=0.06)

    fig.tight_layout(pad=0.25)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    fig.savefig(output_path, dpi=450, transparent=True, bbox_inches="tight", pad_inches=0.05)
    close_figure(fig)


def generate_slide16_charts(*, xlsx_path: Path, output_dir: Path) -> list[Path]:
    """Slide 16: índice de Basiléia trimestral."""

    output_dir.mkdir(parents=True, exist_ok=True)

    wb = load_workbook(filename=xlsx_path, data_only=True)
    sheet_name = _resolve_basileia_sheet_name(wb)
    ws = wb[sheet_name]

    xlabels = [("" if v is None else str(v)).strip() for v in _read_range_row(ws, "C3:E3")]
    series_names = [name for name, _, _ in SLIDE16_SERIES]
    raw_rows = [
        [_to_float_or_nan(v) for v in _read_range_row(ws, f"C{row}:E{row}")]
        for _, row, _ in SLIDE16_SERIES
    ]
    values = np.asarray(raw_rows, dtype=float).T * 100.0

    output_path = output_dir / SLIDE16_OUTPUT
    _plot_stacked_basileia(
        xlabels=xlabels,
        series_names=series_names,
        values_pct=values,
        output_path=output_path,
    )
    return [output_path]


if __name__ == "__main__":
    xlsx = Path("testing.xlsx")
    out = Path(".")
    if xlsx.exists():
        files = generate_slide16_charts(xlsx_path=xlsx, output_dir=out)
        print(f"Gerados: {files}")
    else:
        print(f"Arquivo {xlsx} não encontrado")
