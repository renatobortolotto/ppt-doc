from __future__ import annotations

from pathlib import Path

import numpy as np
from openpyxl import load_workbook
from openpyxl.utils.cell import range_boundaries

from utils.charts_common import close_figure


def _read_range_row(ws, cell_range: str) -> list[object]:
    min_col, min_row, max_col, max_row = range_boundaries(cell_range)
    out: list[object] = []
    for r in range(min_row, max_row + 1):
        for c in range(min_col, max_col + 1):
            out.append(ws.cell(row=r, column=c).value)
    return out


def _to_float(v: object) -> float:
    if v is None:
        return float("nan")
    try:
        return float(v)
    except Exception:
        return float("nan")


def _fmt_pct(v: float, decimals: int = 1) -> str:
    return f"{v:.{decimals}f}%".replace(".", ",")


def _fmt_dec(v: float, decimals: int = 1) -> str:
    return f"{v:,.{decimals}f}".replace(",", "X").replace(".", ",").replace("X", ".")


def _plot_stacked_pct_bars(
    *,
    xlabels: list[str],
    series_names: list[str],
    values: np.ndarray,   # shape [n_bars, n_series]
    output_path: Path,
    colors: list[str],
    bracket_pct: bool = True,
    fmt_label=None,  # callable(float) -> str; defaults to _fmt_dec(v, 1)
) -> None:
    """Barra empilhada com labels dentro de cada segmento e bracket no total."""
    import matplotlib.pyplot as plt
    from matplotlib.colors import to_rgba

    n, m = values.shape
    x = np.arange(n, dtype=float)
    width = 0.62

    fig, ax = plt.subplots(figsize=(10, 4.8), dpi=240)
    fig.patch.set_alpha(0)
    ax.set_facecolor("none")

    if fmt_label is None:
        fmt_label = lambda v: _fmt_dec(v, 1)

    bottom = np.zeros(n, dtype=float)
    segment_centers: list[list[float]] = [[] for _ in range(m)]

    for j in range(m):
        y = np.asarray(values[:, j], dtype=float)
        y_safe = np.where(np.isfinite(y), y, 0.0)
        ax.bar(x, y_safe, width=width, bottom=bottom,
               color=colors[j % len(colors)], edgecolor="none", zorder=2)

        rgba = to_rgba(colors[j % len(colors)])
        lum = 0.2126 * rgba[0] + 0.7152 * rgba[1] + 0.0722 * rgba[2]
        txt_color = "#ffffff" if lum < 0.50 else "#2f2f2f"

        for i in range(n):
            v = float(y[i])
            if not np.isfinite(v) or abs(v) < 1e-9:
                segment_centers[j].append(float("nan"))
                continue
            yc = float(bottom[i]) + v / 2.0
            segment_centers[j].append(yc)
            ax.text(float(x[i]), yc, fmt_label(v),
                    ha="center", va="center", fontsize=9.0,
                    color=txt_color, zorder=4)

        bottom = bottom + y_safe

    totals = bottom.copy()

    # Total label above each bar
    total_tops: list[float] = []
    for i, total in enumerate(totals):
        y_lbl = float(total) + max(abs(float(total)) * 0.02, 0.3)
        total_tops.append(y_lbl)
        ax.text(float(x[i]), y_lbl, fmt_label(total),
                ha="center", va="bottom", fontsize=10.0,
                fontweight="bold" if i == n - 1 else "normal",
                color="#2f2f2f", zorder=5, clip_on=False)

    # Brackets between consecutive bars
    if n >= 2:
        abs_max = float(np.nanmax(np.abs(totals)))
        offset_y = max(abs_max * 0.12, 0.5)
        bracket_h = max(abs_max * 0.035, 0.3)
        top_base = max(total_tops) + max(abs_max * 0.22, 1.0)
        max_text_y: float | None = None

        for i in range(1, n):
            prev, curr = float(totals[i - 1]), float(totals[i])
            if not np.isfinite(prev) or not np.isfinite(curr):
                continue
            if bracket_pct:
                if prev == 0:
                    continue
                lbl = f"{(curr / prev - 1.0) * 100.0:+.1f}%".replace(".", ",")
            else:
                lbl = f"{(curr - prev):+.1f} p.p.".replace(".", ",")

            x1, x2 = float(x[i - 1]), float(x[i])
            ax.plot([x1, x1, x2, x2],
                    [top_base, top_base + bracket_h, top_base + bracket_h, top_base],
                    color="#2f2f2f", linewidth=1.2, zorder=4)
            ty = top_base + bracket_h + offset_y * 0.25
            ax.text((x1 + x2) / 2.0, ty, lbl,
                    ha="center", va="bottom", fontsize=9.0, color="#2f2f2f", zorder=5)
            max_text_y = ty if max_text_y is None else max(max_text_y, ty)

        if max_text_y is not None:
            ymin, ymax = ax.get_ylim()
            ax.set_ylim(ymin, max(ymax, max_text_y + offset_y * 1.4))

    # Inline legend on left
    x_leg = float(x.min()) - 0.95
    for j, name in enumerate(series_names):
        y_ref = next((yc for yc in segment_centers[j] if np.isfinite(yc)), float("nan"))
        if not np.isfinite(y_ref):
            continue
        ax.scatter([x_leg], [y_ref], s=90.0, marker="s",
                   color=colors[j % len(colors)], edgecolors="none", zorder=6)
        ax.text(x_leg + 0.12, y_ref, str(name),
                ha="left", va="center", fontsize=9.0, color="#2f2f2f", zorder=6, clip_on=False)

    ax.set_xlim(float(x.min()) - 1.25, float(x.max()) + 0.65)
    ax.set_xticks(x)
    ax.set_xticklabels(xlabels, fontsize=10.0)
    ax.tick_params(axis="x", bottom=False, pad=8)
    ax.set_yticks([])
    for s in ("left", "right", "top", "bottom"):
        ax.spines[s].set_visible(False)
    ax.grid(False)
    ax.margins(x=0.05, y=0.12)

    fig.tight_layout(pad=0.2)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    fig.savefig(output_path, dpi=450, transparent=True, bbox_inches="tight", pad_inches=0.05)
    close_figure(fig)


def _plot_simple_bars(
    *,
    xlabels: list[str],
    values: list[float],
    output_path: Path,
    fmt_label,
    bar_color: str = "#123a7a",
    bracket_pct: bool = True,
) -> None:
    import matplotlib.pyplot as plt

    vals = np.asarray(values, dtype=float)
    n = len(vals)
    x = np.arange(n, dtype=float)

    fig, ax = plt.subplots(figsize=(10, 4.8), dpi=240)
    fig.patch.set_alpha(0)
    ax.set_facecolor("none")

    bars = ax.bar(x, vals, width=0.62, color=bar_color, edgecolor="none", zorder=2)

    for i, (rect, v) in enumerate(zip(bars, vals)):
        if not np.isfinite(v):
            continue
        ax.text(
            rect.get_x() + rect.get_width() / 2, rect.get_height(),
            fmt_label(v), ha="center", va="bottom", fontsize=10.0,
            fontweight="bold" if i == n - 1 else "normal",
            color="#2f2f2f", zorder=4, clip_on=False,
        )

    if n >= 2:
        abs_max = float(np.nanmax(np.abs(vals)))
        offset_y = max(abs_max * 0.12, 0.5)
        bracket_h = max(abs_max * 0.035, 0.3)
        top_base = float(np.nanmax(vals)) + max(abs_max * 0.25, 1.0)
        max_text_y: float | None = None

        for i in range(1, n):
            prev, curr = float(vals[i - 1]), float(vals[i])
            if not np.isfinite(prev) or not np.isfinite(curr):
                continue
            if bracket_pct:
                if prev == 0:
                    continue
                lbl = f"{(curr / prev - 1.0) * 100.0:+.1f}%".replace(".", ",")
            else:
                lbl = f"{(curr - prev):+.1f} p.p.".replace(".", ",")

            x1, x2 = float(x[i - 1]), float(x[i])
            ax.plot([x1, x1, x2, x2],
                    [top_base, top_base + bracket_h, top_base + bracket_h, top_base],
                    color="#2f2f2f", linewidth=1.2, zorder=4)
            ty = top_base + bracket_h + offset_y * 0.25
            ax.text((x1 + x2) / 2.0, ty, lbl,
                    ha="center", va="bottom", fontsize=9.0, color="#2f2f2f", zorder=5)
            max_text_y = ty if max_text_y is None else max(max_text_y, ty)

        if max_text_y is not None:
            ymin, ymax = ax.get_ylim()
            ax.set_ylim(ymin, max(ymax, max_text_y + offset_y * 1.4))

    ax.set_xticks(x)
    ax.set_xticklabels(xlabels, fontsize=10.0)
    ax.tick_params(axis="x", bottom=False, pad=8)
    ax.set_yticks([])
    for s in ("left", "right", "top", "bottom"):
        ax.spines[s].set_visible(False)
    ax.grid(False)
    ax.margins(x=0.12, y=0.15)

    fig.tight_layout(pad=0.2)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    fig.savefig(output_path, dpi=450, transparent=True, bbox_inches="tight", pad_inches=0.05)
    close_figure(fig)


def generate_slide18_charts(*, xlsx_path: Path, output_dir: Path) -> list[Path]:
    """Slide 18:
    - A3:D5 → 1 gráfico de barra empilhada (Outros Veiculos + Leves Usados), % labels, bracket %.
    - I3:N4 → 2 gráficos separados: trimestral (J:L) e 9M (M:N), brackets em %.
    """

    output_dir.mkdir(parents=True, exist_ok=True)

    wb = load_workbook(filename=xlsx_path, data_only=True)
    sheet_name = "slide_18"
    if sheet_name not in wb.sheetnames:
        raise ValueError(f"Aba não encontrada: {sheet_name!r}. Disponíveis: {wb.sheetnames}")
    ws = wb[sheet_name]

    generated: list[Path] = []

    # ── Block A3:D5 – barra empilhada ───────────────────────────────────────
    labels_a = [("" if v is None else str(v)).strip() for v in _read_range_row(ws, "B3:D3")]
    series_names_a: list[str] = []
    rows_a: list[list[float]] = []
    for row in [4, 5]:
        raw = _read_range_row(ws, f"A{row}:D{row}")
        name = ("" if raw[0] is None else str(raw[0])).strip()
        vals = [_to_float(v) for v in raw[1:]]  # B, C, D
        series_names_a.append(name)
        rows_a.append(vals)

    # values shape: [n_bars=3, n_series=2]
    values_a = np.asarray(rows_a, dtype=float).T

    out33 = output_dir / "33_veiculos_empilhado.png"
    _plot_stacked_pct_bars(
        xlabels=labels_a,
        series_names=series_names_a,
        values=values_a,
        output_path=out33,
        colors=["#123a7a", "#5B8FF9"],
        bracket_pct=True,
    )
    generated.append(out33)

    # ── Block I3:N4 ─────────────────────────────────────────────────────────
    labels_i = [("" if v is None else str(v)).strip() for v in _read_range_row(ws, "J3:N3")]
    raw_i = _read_range_row(ws, "I4:N4")
    vals_i = [_to_float(v) for v in raw_i[1:]]  # J, K, L, M, N

    # 34 – Trimestral: J, K, L
    out34 = output_dir / "34_premios_seguros_trimestres.png"
    _plot_simple_bars(
        xlabels=labels_i[:3],
        values=vals_i[:3],
        output_path=out34,
        fmt_label=lambda v: _fmt_dec(v, 1),
        bracket_pct=True,
    )
    generated.append(out34)

    # 35 – 9M: M, N
    out35 = output_dir / "35_premios_seguros_9m.png"
    _plot_simple_bars(
        xlabels=labels_i[3:],
        values=vals_i[3:],
        output_path=out35,
        fmt_label=lambda v: _fmt_dec(v, 1),
        bracket_pct=True,
    )
    generated.append(out35)

    return generated


if __name__ == "__main__":
    xlsx = Path("testing.xlsx")
    out = Path(".")
    if xlsx.exists():
        files = generate_slide18_charts(xlsx_path=xlsx, output_dir=out)
        print(f"Gerados: {files}")
    else:
        print(f"Arquivo {xlsx} não encontrado")
