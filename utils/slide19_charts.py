from __future__ import annotations

from pathlib import Path

import numpy as np
from openpyxl import load_workbook
from openpyxl.utils.cell import range_boundaries

from utils.charts_common import close_figure


SLIDE19_VEICULOS_OUTPUT = "19_veiculos_empilhado.png"
SLIDE19_SEGUROS_TRIMESTRES_OUTPUT = "19_seguros_cartoes_trimestres.png"
SLIDE19_SEGUROS_ANOS_OUTPUT = "19_seguros_cartoes_anos.png"
SLIDE19_SHEET_CANDIDATES = ("Veículos", "Veiculos")
SLIDE19_SEGUROS_SHEET_CANDIDATES = ("Seguros e Cartões", "Seguros e Cartoes")
SLIDE19_SERIES = (
    ("Leves Usados", "#123A7A"),
    ("Outros Veículos", "#8FB0E8"),
)
SLIDE19_FIGSIZE = (7.4, 3.5)
SLIDE19_BAR_SLOT = 1.0
SLIDE19_BAR_WIDTH = 0.66
SLIDE19_SIMPLE_BAR_COLOR = "#123A7A"


def _read_range_row(ws, cell_range: str) -> list[object]:
    min_col, min_row, max_col, max_row = range_boundaries(cell_range)
    out: list[object] = []
    for r in range(min_row, max_row + 1):
        for c in range(min_col, max_col + 1):
            out.append(ws.cell(row=r, column=c).value)
    return out


def _to_float_or_nan(v: object) -> float:
    if v is None:
        return float("nan")
    if isinstance(v, str):
        s = v.strip()
        if s == "":
            return float("nan")
        s = s.replace("%", "").replace(" ", "")
        s = s.replace(".", "").replace(",", ".") if ("," in s and "." in s) else s.replace(",", ".")
        try:
            return float(s)
        except Exception:
            return float("nan")
    try:
        return float(v)
    except Exception:
        return float("nan")


def _fmt_value(v: float) -> str:
    return f"{float(v):.1f}".replace(".", ",")


def _fmt_value_auto(v: float) -> str:
    value = float(v)
    if abs(value - round(value)) < 1e-9:
        return f"{int(round(value))}"
    return _fmt_value(value)


def _fmt_share_pct(v: float) -> str:
    return f"{float(v):.0f}%".replace(".", ",")


def _fmt_bracket_pct(curr: float, prev: float) -> str:
    return f"{((curr / prev) - 1.0) * 100.0:+.1f}%".replace(".", ",")


def _text_color_for_bg_rgba(rgba) -> str:
    r, g, b = float(rgba[0]), float(rgba[1]), float(rgba[2])
    lum = 0.2126 * r + 0.7152 * g + 0.0722 * b
    return "#ffffff" if lum < 0.50 else "#2f2f2f"


def _resolve_veiculos_sheet_name(wb) -> str:
    for candidate in SLIDE19_SHEET_CANDIDATES:
        if candidate in wb.sheetnames:
            return candidate
    raise ValueError(f"Aba não encontrada: 'Veículos'. Disponíveis: {wb.sheetnames}")


def _resolve_seguros_sheet_name(wb) -> str:
    for candidate in SLIDE19_SEGUROS_SHEET_CANDIDATES:
        if candidate in wb.sheetnames:
            return candidate
    raise ValueError(f"Aba não encontrada: 'Seguros e Cartões'. Disponíveis: {wb.sheetnames}")


def _major_series_indices(values: np.ndarray) -> tuple[int, ...]:
    result: list[int] = []
    for row in np.asarray(values, dtype=float):
        valid = [
            (idx, float(value))
            for idx, value in enumerate(row)
            if np.isfinite(float(value))
        ]
        if not valid:
            result.append(-1)
            continue
        result.append(max(valid, key=lambda item: (item[1], -item[0]))[0])
    return tuple(result)


def _plot_stacked_veiculos(
    *,
    xlabels: list[str],
    series_names: list[str],
    values: np.ndarray,  # [n_bars, n_series]
    output_path: Path,
) -> None:
    import matplotlib.pyplot as plt
    from matplotlib.colors import to_rgba

    n, m = values.shape
    if n == 0 or m == 0:
        raise ValueError("Sem dados para plotar")

    fig, ax = plt.subplots(figsize=SLIDE19_FIGSIZE, dpi=240)
    fig.patch.set_alpha(0)
    ax.set_facecolor("none")

    x = np.arange(n, dtype=float) * SLIDE19_BAR_SLOT
    width = SLIDE19_BAR_WIDTH
    bottom = np.zeros(n, dtype=float)
    segment_centers: list[list[float]] = [[] for _ in range(m)]
    totals = np.nansum(np.where(np.isfinite(values), values, 0.0), axis=1).astype(float)
    major_indices = _major_series_indices(values)
    colors = [color for _, color in SLIDE19_SERIES]

    for j in range(m):
        y = np.asarray(values[:, j], dtype=float)
        y_safe = np.where(np.isfinite(y), y, 0.0)
        color = colors[j % len(colors)]
        ax.bar(
            x,
            y_safe,
            width=width,
            bottom=bottom,
            color=color,
            edgecolor="none",
            zorder=2,
        )

        txt_color = _text_color_for_bg_rgba(to_rgba(color))
        for i in range(n):
            value = float(y[i])
            if not np.isfinite(value) or abs(value) < 1e-9:
                segment_centers[j].append(float("nan"))
                continue
            yc = float(bottom[i]) + value / 2.0
            label = _fmt_value(value)
            if major_indices[i] == j and np.isfinite(totals[i]) and totals[i] > 0:
                share = (value / float(totals[i])) * 100.0
                label = f"{label}\n({_fmt_share_pct(share)})"
            segment_centers[j].append(yc)
            ax.text(
                float(x[i]),
                yc,
                label,
                ha="center",
                va="center",
                fontsize=8.8,
                color=txt_color,
                zorder=4,
                linespacing=0.95,
                clip_on=False,
            )

        bottom = bottom + y_safe

    total_label_tops: list[float] = []
    for i, total in enumerate(totals):
        if not np.isfinite(total):
            continue
        y_label = float(total) + max(abs(float(total)) * 0.03, 0.28)
        total_label_tops.append(y_label)
        ax.text(
            float(x[i]),
            y_label,
            _fmt_value(total),
            ha="center",
            va="bottom",
            fontsize=9.8,
            fontweight="bold" if i == n - 1 else "normal",
            color="#2f2f2f",
            zorder=5,
            clip_on=False,
        )

    if n >= 2:
        abs_max = float(np.nanmax(np.abs(totals))) if np.isfinite(np.nanmax(np.abs(totals))) else 0.0
        offset_y = max(abs_max * 0.06, 0.18)
        bracket_h = max(abs_max * 0.022, 0.10)
        top_labels_max = max(total_label_tops) if total_label_tops else float(np.nanmax(totals))
        top_base = float(top_labels_max) + max(abs_max * 0.07, 0.22)
        max_text_y: float | None = None

        for i in range(1, n):
            prev = float(totals[i - 1])
            curr = float(totals[i])
            if not np.isfinite(prev) or not np.isfinite(curr) or abs(prev) < 1e-12:
                continue
            x1 = float(x[i - 1])
            x2 = float(x[i])
            text_y = top_base + bracket_h + offset_y * 0.25
            ax.plot(
                [x1, x1, x2, x2],
                [top_base, top_base + bracket_h, top_base + bracket_h, top_base],
                color="#2f2f2f",
                linewidth=1.2,
                solid_capstyle="round",
                zorder=4,
            )
            ax.text(
                (x1 + x2) / 2.0,
                text_y,
                _fmt_bracket_pct(curr, prev),
                ha="center",
                va="bottom",
                fontsize=9.0,
                color="#2f2f2f",
                zorder=5,
            )
            max_text_y = text_y if max_text_y is None else max(max_text_y, text_y)

        if max_text_y is not None:
            ymin, ymax = ax.get_ylim()
            ax.set_ylim(ymin, max(ymax, max_text_y + offset_y * 0.9))

    max_name_len = max((len(str(name).strip()) for name in series_names), default=1)
    left_margin = max(1.55, 0.78 + max_name_len * 0.043)
    x_text = float(x[0]) - width / 2.0 - 0.20
    connector_start = x_text + 0.06
    connector_end = float(x[0]) - width / 2.0 - 0.05
    for j, name in enumerate(series_names):
        y_ref = segment_centers[j][0] if segment_centers[j] else float("nan")
        if not np.isfinite(y_ref):
            for yc in segment_centers[j]:
                if np.isfinite(yc):
                    y_ref = yc
                    break
        if not np.isfinite(y_ref):
            continue
        if connector_end > connector_start:
            ax.plot(
                [connector_start, connector_end],
                [float(y_ref), float(y_ref)],
                color=colors[j % len(colors)],
                linewidth=2.2,
                solid_capstyle="round",
                zorder=6,
                clip_on=False,
            )
        ax.text(
            x_text,
            float(y_ref),
            str(name),
            ha="right",
            va="center",
            fontsize=8.8,
            color="#2f2f2f",
            zorder=7,
            clip_on=False,
        )

    ax.axhline(0.0, color="#b5b5b5", linewidth=0.9, zorder=1)
    ax.set_xlim(float(x.min()) - (left_margin + 0.10), float(x.max()) + 0.36)
    ax.set_xticks(x)
    ax.set_xticklabels(xlabels, fontsize=10.0)
    ax.tick_params(axis="x", bottom=False, pad=8)
    ax.set_yticks([])
    for spine in ("left", "right", "top", "bottom"):
        ax.spines[spine].set_visible(False)
    ax.grid(False)
    ax.margins(x=0.04, y=0.06)

    fig.tight_layout(pad=0.25)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    fig.savefig(output_path, dpi=450, transparent=True, bbox_inches="tight", pad_inches=0.05)
    close_figure(fig)


def _plot_simple_bars(
    *,
    xlabels: list[str],
    values: list[float] | np.ndarray,
    output_path: Path,
    bar_color: str = SLIDE19_SIMPLE_BAR_COLOR,
) -> None:
    import matplotlib.pyplot as plt

    vals = np.asarray(values, dtype=float)
    n = len(vals)
    if n == 0:
        raise ValueError("Sem dados para plotar")

    fig, ax = plt.subplots(figsize=(6.5, 3.1), dpi=240)
    fig.patch.set_alpha(0)
    ax.set_facecolor("none")

    x = np.arange(n, dtype=float)
    bars = ax.bar(x, vals, width=0.62, color=bar_color, edgecolor="none", zorder=2)

    label_tops: list[float] = []
    for i, (rect, value) in enumerate(zip(bars, vals)):
        if not np.isfinite(value):
            continue
        y_label = float(rect.get_height()) + max(abs(float(value)) * 0.03, 8.0)
        label_tops.append(y_label)
        ax.text(
            rect.get_x() + rect.get_width() / 2.0,
            y_label,
            _fmt_value(value),
            ha="center",
            va="bottom",
            fontsize=9.8,
            fontweight="bold" if i == n - 1 else "normal",
            color="#2f2f2f",
            zorder=4,
            clip_on=False,
        )

    if n >= 2:
        abs_max = float(np.nanmax(np.abs(vals))) if np.isfinite(np.nanmax(np.abs(vals))) else 0.0
        offset_y = max(abs_max * 0.08, 10.0)
        bracket_h = max(abs_max * 0.03, 8.0)
        top_labels_max = max(label_tops) if label_tops else float(np.nanmax(vals))
        top_base = float(top_labels_max) + max(abs_max * 0.08, 12.0)
        max_text_y: float | None = None

        for i in range(1, n):
            prev = float(vals[i - 1])
            curr = float(vals[i])
            if not np.isfinite(prev) or not np.isfinite(curr) or abs(prev) < 1e-12:
                continue
            x1 = float(x[i - 1])
            x2 = float(x[i])
            text_y = top_base + bracket_h + offset_y * 0.25
            ax.plot(
                [x1, x1, x2, x2],
                [top_base, top_base + bracket_h, top_base + bracket_h, top_base],
                color="#2f2f2f",
                linewidth=1.2,
                solid_capstyle="round",
                zorder=4,
            )
            ax.text(
                (x1 + x2) / 2.0,
                text_y,
                _fmt_bracket_pct(curr, prev),
                ha="center",
                va="bottom",
                fontsize=9.0,
                color="#2f2f2f",
                zorder=5,
            )
            max_text_y = text_y if max_text_y is None else max(max_text_y, text_y)

        if max_text_y is not None:
            ymin, ymax = ax.get_ylim()
            ax.set_ylim(ymin, max(ymax, max_text_y + offset_y * 0.9))

    ax.axhline(0.0, color="#b5b5b5", linewidth=0.9, zorder=1)
    ax.set_xticks(x)
    ax.set_xticklabels(xlabels, fontsize=10.0)
    ax.tick_params(axis="x", bottom=False, pad=8)
    ax.set_yticks([])
    for spine in ("left", "right", "top", "bottom"):
        ax.spines[spine].set_visible(False)
    ax.grid(False)
    ax.margins(x=0.18, y=0.06)

    fig.tight_layout(pad=0.25)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    fig.savefig(output_path, dpi=450, transparent=True, bbox_inches="tight", pad_inches=0.05)
    close_figure(fig)


def generate_slide19_charts(*, xlsx_path: Path, output_dir: Path) -> list[Path]:
    """Slide 19: empilhado de veículos por trimestre."""

    output_dir.mkdir(parents=True, exist_ok=True)

    wb = load_workbook(filename=xlsx_path, data_only=True)
    ws_veiculos = wb[_resolve_veiculos_sheet_name(wb)]
    ws_seguros = wb[_resolve_seguros_sheet_name(wb)]

    xlabels = [("" if v is None else str(v)).strip() for v in _read_range_row(ws_veiculos, "D14:F14")]

    leves_usados = np.asarray([_to_float_or_nan(v) for v in _read_range_row(ws_veiculos, "D21:F21")], dtype=float)
    outros_rows = np.asarray(
        [[_to_float_or_nan(v) for v in _read_range_row(ws_veiculos, f"D{row}:F{row}")] for row in (22, 23)],
        dtype=float,
    )
    outros_veiculos = np.nansum(np.where(np.isfinite(outros_rows), outros_rows, 0.0), axis=0)

    values = np.column_stack([leves_usados, outros_veiculos]).astype(float)
    output_path = output_dir / SLIDE19_VEICULOS_OUTPUT
    _plot_stacked_veiculos(
        xlabels=xlabels,
        series_names=[name for name, _ in SLIDE19_SERIES],
        values=values,
        output_path=output_path,
    )

    seguros_trim_labels = [("" if v is None else str(v)).strip() for v in _read_range_row(ws_seguros, "D3:F3")]
    seguros_trim_values = [_to_float_or_nan(v) for v in _read_range_row(ws_seguros, "D8:F8")]
    seguros_trim_output = output_dir / SLIDE19_SEGUROS_TRIMESTRES_OUTPUT
    _plot_simple_bars(
        xlabels=seguros_trim_labels,
        values=seguros_trim_values,
        output_path=seguros_trim_output,
    )

    seguros_ano_labels = [("" if v is None else str(v)).strip() for v in _read_range_row(ws_seguros, "G3:H3")]
    seguros_ano_values = [_to_float_or_nan(v) for v in _read_range_row(ws_seguros, "G8:H8")]
    seguros_ano_output = output_dir / SLIDE19_SEGUROS_ANOS_OUTPUT
    _plot_simple_bars(
        xlabels=seguros_ano_labels,
        values=seguros_ano_values,
        output_path=seguros_ano_output,
    )

    return [output_path, seguros_trim_output, seguros_ano_output]


if __name__ == "__main__":
    xlsx = Path("testing.xlsx")
    out = Path(".")
    if xlsx.exists():
        files = generate_slide19_charts(xlsx_path=xlsx, output_dir=out)
        print(f"Gerados: {files}")
    else:
        print(f"Arquivo {xlsx} não encontrado")
