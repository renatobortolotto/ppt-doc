from __future__ import annotations

from pathlib import Path

import numpy as np
from openpyxl import load_workbook
from openpyxl.utils.cell import range_boundaries

from utils.charts_common import close_figure


SLIDE21_STACKED_OUTPUT = "21_carteira_atacado_empilhado.png"
SLIDE21_COMPARATIVE_OUTPUT = "21_carteira_atacado_comparativo.png"
SLIDE21_CARTEIRA_SHEET_CANDIDATES = ("Carteira",)
SLIDE21_ATAC_SETOR_SHEET_CANDIDATES = ("Carteira Atac Setor",)
SLIDE21_STACKED_FIGSIZE = (7.4, 3.0)
SLIDE21_STACKED_BAR_SLOT = 0.24
SLIDE21_STACKED_BAR_WIDTH = 0.20
SLIDE21_STACKED_FONT_SCALE = 1.2
SLIDE21_COMPARATIVE_FIGSIZE = (13.6, 4.8)
SLIDE21_COMPARATIVE_BAR_WIDTH = 0.36
SLIDE21_COMPARATIVE_FONT_SCALE = 1.0
SLIDE21_COMPARATIVE_COLORS = ("#1E4588", "#5B84E8")
SLIDE21_STACKED_COLOR_BY_KEY = {
    "pme": "#9AA0A6",
    "corporate": "#123A7A",
    "if": "#5B84E8",
    "large": "#5B84E8",
}


def _read_range_row(ws, cell_range: str) -> list[object]:
    min_col, min_row, max_col, max_row = range_boundaries(cell_range)
    out: list[object] = []
    for r in range(min_row, max_row + 1):
        for c in range(min_col, max_col + 1):
            out.append(ws.cell(row=r, column=c).value)
    return out


def _read_range_column(ws, cell_range: str) -> list[object]:
    min_col, min_row, max_col, max_row = range_boundaries(cell_range)
    out: list[object] = []
    for c in range(min_col, max_col + 1):
        for r in range(min_row, max_row + 1):
            out.append(ws.cell(row=r, column=c).value)
    return out


def _to_float_or_nan(v: object) -> float:
    if v is None:
        return float("nan")
    if isinstance(v, str):
        s = v.strip()
        if not s:
            return float("nan")
        s = s.replace("%", "").replace(" ", "")
        s = s.replace(".", "").replace(",", ".") if ("," in s and "." in s) else s.replace(",", ".")
        try:
            return float(s)
        except Exception:
            return float("nan")
    try:
        return float(v)
    except Exception:
        return float("nan")


def _fmt_num(v: float) -> str:
    value = float(v)
    decimals = 2 if 0 < abs(value) < 0.1 else 1
    return f"{value:.{decimals}f}".replace(".", ",")


def _fmt_pct(v: float) -> str:
    return f"{float(v):.0f}%".replace(".", ",")


def _fmt_bracket_pct(curr: float, prev: float) -> str:
    return f"{((curr / prev) - 1.0) * 100.0:+.1f}%".replace(".", ",")


def _fmt_comparative_value(v: float) -> str:
    value = float(v)
    if abs(value - round(value)) < 1e-9:
        return f"{int(round(value))}%"
    return f"{value:.1f}%".replace(".", ",")


def _text_color_for_bg_rgba(rgba) -> str:
    r, g, b = float(rgba[0]), float(rgba[1]), float(rgba[2])
    lum = 0.2126 * r + 0.7152 * g + 0.0722 * b
    return "#ffffff" if lum < 0.50 else "#2f2f2f"


def _resolve_sheet_name(wb, candidates: tuple[str, ...], label: str) -> str:
    for candidate in candidates:
        if candidate in wb.sheetnames:
            return candidate
    raise ValueError(f"Aba não encontrada: {label!r}. Disponíveis: {wb.sheetnames}")


def _normalize_series_name(value: object) -> str:
    text = ("" if value is None else str(value)).strip()
    if not text:
        return ""
    lowered = text.lower()
    if lowered == "corporate":
        return "Corporate"
    if lowered in {"large + if", "large e if", "large + ifs"}:
        return "Large + IF"
    if lowered in {"pme", "pmes"}:
        return "PMEs"
    return text


def _stack_order_for_bar(values_row: np.ndarray) -> tuple[int, ...]:
    row = np.asarray(values_row, dtype=float)
    return tuple(
        sorted(
            range(len(row)),
            key=lambda idx: (
                1 if not np.isfinite(float(row[idx])) else 0,
                -float(row[idx]) if np.isfinite(float(row[idx])) else 0.0,
                idx,
            ),
        )
    )


def _share_label_indices(values_row: np.ndarray) -> set[int]:
    ordered_positive = [
        idx
        for idx in _stack_order_for_bar(values_row)
        if np.isfinite(float(values_row[idx])) and float(values_row[idx]) > 0
    ]
    if len(ordered_positive) >= 4:
        top_n = 3
    elif len(ordered_positive) >= 3:
        top_n = 2
    else:
        top_n = len(ordered_positive)
    return set(ordered_positive[:top_n])


def _top_badge_label_indices(values_row: np.ndarray) -> set[int]:
    row = np.asarray(values_row, dtype=float)
    ordered_positive = [
        idx for idx in _stack_order_for_bar(row) if np.isfinite(float(row[idx])) and float(row[idx]) > 0
    ]
    if not ordered_positive:
        return set()

    total = float(np.nansum(row))
    if not np.isfinite(total) or total <= 0:
        return set()

    candidates = ordered_positive[-2:]
    return {
        idx
        for idx in candidates
        if ((float(row[idx]) / total) * 100.0) <= 12.0
    }


def _stacked_colors(series_names: list[str]) -> tuple[str, ...]:
    colors: list[str] = []
    for name in series_names:
        key = name.lower()
        color = "#9AA0A6"
        for token, token_color in SLIDE21_STACKED_COLOR_BY_KEY.items():
            if token in key:
                color = token_color
                break
        colors.append(color)
    return tuple(colors)


def _plot_stacked_atacado(
    *,
    xlabels: list[str],
    series_names: list[str],
    values: np.ndarray,
    output_path: Path,
) -> None:
    import matplotlib.pyplot as plt
    from matplotlib.colors import to_rgba

    values_arr = np.asarray(values, dtype=float)
    n, m = values_arr.shape
    if n == 0 or m == 0:
        raise ValueError("Slide 21 sem dados para plotar")

    colors = _stacked_colors(series_names)
    fig, ax = plt.subplots(figsize=SLIDE21_STACKED_FIGSIZE, dpi=240)
    fig.patch.set_alpha(0)
    ax.set_facecolor("none")

    x = np.arange(n, dtype=float) * SLIDE21_STACKED_BAR_SLOT
    width = SLIDE21_STACKED_BAR_WIDTH
    segment_centers = np.full_like(values_arr, np.nan, dtype=float)
    totals = np.zeros(n, dtype=float)
    orders = tuple(_stack_order_for_bar(values_arr[i, :]) for i in range(n))

    for i in range(n):
        bottom = 0.0
        for idx in orders[i]:
            value = float(values_arr[i, idx])
            if not np.isfinite(value) or value <= 0:
                continue
            color = colors[idx % len(colors)]
            ax.bar(
                float(x[i]),
                value,
                width=width,
                bottom=bottom,
                color=color,
                edgecolor="none",
                zorder=2,
            )
            segment_centers[i, idx] = bottom + value / 2.0
            bottom += value
        totals[i] = bottom

    finite_totals = totals[np.isfinite(totals)]
    abs_max = float(np.nanmax(np.abs(finite_totals))) if finite_totals.size else 0.0
    total_label_gap = max(abs_max * 0.02, 0.15)
    total_label_tops: list[float] = []
    for i, total in enumerate(totals):
        if not np.isfinite(total):
            continue
        y_label = float(total) + total_label_gap
        total_label_tops.append(y_label)
        ax.text(
            float(x[i]),
            y_label,
            _fmt_num(total),
            ha="center",
            va="bottom",
            fontsize=10.0 * SLIDE21_STACKED_FONT_SCALE,
            fontweight="bold" if i == n - 1 else "normal",
            color="#2f2f2f",
            zorder=5,
            clip_on=False,
        )

    for i in range(n):
        total = float(totals[i])
        if not np.isfinite(total) or total <= 0:
            continue
        pct_indices = _share_label_indices(values_arr[i, :])
        badge_indices = _top_badge_label_indices(values_arr[i, :])
        for idx in orders[i]:
            value = float(values_arr[i, idx])
            if not np.isfinite(value) or value <= 0:
                continue
            share = (value / total) * 100.0
            label = _fmt_num(value)
            if idx in pct_indices:
                label = f"{label} ({_fmt_pct(share)})"

            rgba = to_rgba(colors[idx % len(colors)])
            txt_color = _text_color_for_bg_rgba(rgba)
            use_badge = idx in badge_indices
            ax.text(
                float(x[i]),
                float(segment_centers[i, idx]),
                label,
                ha="center",
                va="center",
                fontsize=(8.6 if idx in pct_indices else 7.8) * SLIDE21_STACKED_FONT_SCALE,
                color=txt_color,
                fontweight="bold" if idx in pct_indices else "normal",
                zorder=7 if use_badge else 6,
                clip_on=not use_badge,
                bbox=(
                    {
                        "facecolor": colors[idx % len(colors)],
                        "edgecolor": "none",
                        "boxstyle": "round,pad=0.16",
                    }
                    if use_badge
                    else None
                ),
            )

    if n >= 2:
        offset_y = max(abs_max * 0.07, 0.24)
        bracket_h = max(abs_max * 0.028, 0.15)
        top_labels_max = max(total_label_tops) if total_label_tops else float(np.nanmax(totals))
        top_base = float(top_labels_max) + max(abs_max * 0.08, 0.32)
        max_text_y: float | None = None

        for i in range(1, n):
            prev = float(totals[i - 1])
            curr = float(totals[i])
            if not np.isfinite(prev) or not np.isfinite(curr) or prev == 0:
                continue
            x1 = float(x[i - 1])
            x2 = float(x[i])
            ax.plot(
                [x1, x1, x2, x2],
                [top_base, top_base + bracket_h, top_base + bracket_h, top_base],
                color="#2f2f2f",
                linewidth=1.2,
                solid_capstyle="round",
                zorder=4,
            )
            text_y = top_base + bracket_h + offset_y * 0.25
            ax.text(
                (x1 + x2) / 2.0,
                text_y,
                _fmt_bracket_pct(curr, prev),
                ha="center",
                va="bottom",
                fontsize=9.0 * SLIDE21_STACKED_FONT_SCALE,
                color="#2f2f2f",
                zorder=5,
            )
            max_text_y = text_y if max_text_y is None else max(max_text_y, text_y)

        if max_text_y is not None:
            ymin, ymax = ax.get_ylim()
            ax.set_ylim(ymin, max(ymax, max_text_y + offset_y * 0.9))

    x_text = float(x[0]) - width / 2.0 - 0.06
    for idx in orders[0]:
        y_ref = float(segment_centers[0, idx])
        if not np.isfinite(y_ref):
            for row_idx in range(n):
                candidate = float(segment_centers[row_idx, idx])
                if np.isfinite(candidate):
                    y_ref = candidate
                    break
        if not np.isfinite(y_ref):
            continue
        ax.text(
            x_text,
            y_ref,
            str(series_names[idx]),
            ha="right",
            va="center",
            fontsize=8.4 * SLIDE21_STACKED_FONT_SCALE,
            color="#2f2f2f",
            zorder=7,
            clip_on=False,
        )

    ax.set_xlim(float(x.min()) - 0.72, float(x.max()) + 0.28)
    ax.set_xticks(x)
    ax.set_xticklabels(xlabels, fontsize=10.0 * SLIDE21_STACKED_FONT_SCALE)
    ax.tick_params(axis="x", bottom=False, pad=8)
    ax.set_yticks([])
    for spine in ("left", "right", "top", "bottom"):
        ax.spines[spine].set_visible(False)
    ax.grid(False)
    ax.margins(x=0.04, y=0.12)

    fig.tight_layout(pad=0.3)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    fig.savefig(output_path, dpi=450, transparent=True, bbox_inches="tight", pad_inches=0.06)
    close_figure(fig)


def _plot_comparative_bars(
    *,
    categories: list[str],
    label_left: str,
    label_right: str,
    values_left: np.ndarray,
    values_right: np.ndarray,
    output_path: Path,
) -> None:
    import matplotlib.pyplot as plt

    left = np.asarray(values_left, dtype=float)
    right = np.asarray(values_right, dtype=float)
    valid_mask = [bool(str(cat).strip()) for cat in categories]
    categories = [str(cat).strip() for cat, keep in zip(categories, valid_mask) if keep]
    left = left[np.asarray(valid_mask, dtype=bool)]
    right = right[np.asarray(valid_mask, dtype=bool)]

    if not categories:
        raise ValueError("Sem categorias para plotar")

    fig, ax = plt.subplots(figsize=SLIDE21_COMPARATIVE_FIGSIZE, dpi=240)
    fig.patch.set_alpha(0)
    ax.set_facecolor("none")

    x = np.arange(len(categories), dtype=float)
    width = SLIDE21_COMPARATIVE_BAR_WIDTH
    left_x = x - width / 2.0
    right_x = x + width / 2.0

    bars_left = ax.bar(left_x, left, width=width * 0.92, color=SLIDE21_COMPARATIVE_COLORS[0], edgecolor="none", zorder=2)
    bars_right = ax.bar(right_x, right, width=width * 0.92, color=SLIDE21_COMPARATIVE_COLORS[1], edgecolor="none", zorder=2)

    max_val = max(
        float(np.nanmax(left)) if left.size else 0.0,
        float(np.nanmax(right)) if right.size else 0.0,
    )
    label_gap = max(max_val * 0.02, 0.35)
    for bars, values in ((bars_left, left), (bars_right, right)):
        for rect, value in zip(bars, values):
            if not np.isfinite(value):
                continue
            ax.text(
                rect.get_x() + rect.get_width() / 2.0,
                float(rect.get_height()) + label_gap,
                _fmt_comparative_value(value),
                ha="center",
                va="bottom",
                fontsize=9.4 * SLIDE21_COMPARATIVE_FONT_SCALE,
                fontweight="normal",
                color="#3f3f3f",
                zorder=4,
                clip_on=False,
            )

    ax.set_xticks(x)
    ax.set_xticklabels(
        categories,
        rotation=90,
        ha="center",
        va="top",
        fontsize=9.2 * SLIDE21_COMPARATIVE_FONT_SCALE,
    )
    ax.tick_params(axis="x", bottom=False, pad=6)
    ax.set_yticks([])
    for spine in ("left", "right", "top", "bottom"):
        ax.spines[spine].set_visible(False)
    ax.grid(False)
    ax.margins(x=0.02, y=0.12)
    ax.legend(
        [bars_left[0], bars_right[0]],
        [label_left, label_right],
        loc="lower center",
        bbox_to_anchor=(0.5, 1.02),
        ncol=2,
        frameon=False,
        fontsize=10.0 * SLIDE21_COMPARATIVE_FONT_SCALE,
    )

    ylim_top = max_val + label_gap + max(max_val * 0.06, 0.8)
    ax.set_ylim(0.0, ylim_top)

    fig.tight_layout(pad=0.35)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    fig.savefig(output_path, dpi=450, transparent=True, bbox_inches="tight", pad_inches=0.08)
    close_figure(fig)


def generate_slide21_charts(*, xlsx_path: Path, output_dir: Path) -> list[Path]:
    """Slide 21: atacado empilhado + comparativo por setor."""

    output_dir.mkdir(parents=True, exist_ok=True)

    wb = load_workbook(filename=xlsx_path, data_only=True)
    ws_carteira = wb[_resolve_sheet_name(wb, SLIDE21_CARTEIRA_SHEET_CANDIDATES, "Carteira")]
    ws_setor = wb[_resolve_sheet_name(wb, SLIDE21_ATAC_SETOR_SHEET_CANDIDATES, "Carteira Atac Setor")]

    stacked_xlabels = [("" if v is None else str(v)).strip() for v in _read_range_row(ws_carteira, "D115:F115")]
    stacked_series_names = [
        _normalize_series_name(value)
        for value in _read_range_column(ws_carteira, "C117:C119")
    ]
    stacked_values = np.asarray(
        [
            [_to_float_or_nan(v) for v in _read_range_row(ws_carteira, f"D{row}:F{row}")]
            for row in range(117, 120)
        ],
        dtype=float,
    ).T / 1000.0

    stacked_output = output_dir / SLIDE21_STACKED_OUTPUT
    _plot_stacked_atacado(
        xlabels=stacked_xlabels,
        series_names=stacked_series_names,
        values=stacked_values,
        output_path=stacked_output,
    )

    comparative_categories = [
        ("" if value is None else str(value)).strip()
        for value in _read_range_column(ws_setor, "B4:B23")
    ]
    comparative_label_left = ("" if ws_setor["C2"].value is None else str(ws_setor["C2"].value)).strip()
    comparative_label_right = ("" if ws_setor["E2"].value is None else str(ws_setor["E2"].value)).strip()
    comparative_left = np.asarray(
        [_to_float_or_nan(v) for v in _read_range_column(ws_setor, "D4:D23")],
        dtype=float,
    )
    comparative_right = np.asarray(
        [_to_float_or_nan(v) for v in _read_range_column(ws_setor, "F4:F23")],
        dtype=float,
    )
    valid_mask = np.asarray([bool(category) for category in comparative_categories], dtype=bool)
    comparative_categories = [category for category, keep in zip(comparative_categories, valid_mask) if keep]
    comparative_left = comparative_left[valid_mask]
    comparative_right = comparative_right[valid_mask]

    max_abs_left = float(np.nanmax(np.abs(comparative_left))) if comparative_left.size else 0.0
    max_abs_right = float(np.nanmax(np.abs(comparative_right))) if comparative_right.size else 0.0
    scale_factor = 100.0 if max(max_abs_left, max_abs_right) <= 1.5 else 1.0
    comparative_output = output_dir / SLIDE21_COMPARATIVE_OUTPUT
    _plot_comparative_bars(
        categories=comparative_categories,
        label_left=comparative_label_left,
        label_right=comparative_label_right,
        values_left=comparative_left * scale_factor,
        values_right=comparative_right * scale_factor,
        output_path=comparative_output,
    )

    return [stacked_output, comparative_output]


if __name__ == "__main__":
    xlsx = Path("testing.xlsx")
    out = Path(".")
    if xlsx.exists():
        files = generate_slide21_charts(xlsx_path=xlsx, output_dir=out)
        print(f"Gerados: {files}")
    else:
        print(f"Arquivo {xlsx} não encontrado")
