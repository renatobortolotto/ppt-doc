from __future__ import annotations

from pathlib import Path
from typing import Iterable

import numpy as np
from openpyxl import load_workbook

from utils.charts_common import close_figure, pchip_interpolate, read_range_col, to_float_list


def _sanitize_filename(s: str) -> str:
    import re

    s = (s or "").strip()
    s = re.sub(r"\s+", "_", s)
    s = re.sub(r"[^A-Za-z0-9_.-]", "", s)
    return s or "serie"


def _find_title_index(titles: list[str], *keywords: str) -> int | None:
    kws = [k.lower() for k in keywords if k]
    for idx, t in enumerate(titles):
        tl = (t or "").lower()
        if any(k in tl for k in kws):
            return idx
    return None


def plot_multi_line(
    *,
    xlabels: list[str],
    series: list[tuple[str, list[float], dict]],
    output_path: Path,
    fmt_as_percent: bool = False,
    smooth: bool = True,
    smooth_points: int = 1400,
    y_pad_multiplier: float = 1.6,
):
    import matplotlib

    matplotlib.use("Agg", force=True)
    import matplotlib.pyplot as plt

    x = np.arange(len(xlabels), dtype=float)
    fig, ax = plt.subplots(figsize=(10, 4.2), dpi=200)
    fig.patch.set_alpha(0)
    ax.set_facecolor("none")

    y_samples: list[float] = []

    for _name, values, style in series:
        y = np.asarray(values, dtype=float)
        finite = np.isfinite(y)
        if finite.any():
            y_samples.extend([float(v) for v in y[finite]])

        color = style.get("color", "#2f2f2f")
        linestyle = style.get("linestyle", "-")
        line_width = float(style.get("line_width", 2.8))
        marker_color = style.get("marker_color", color)
        marker_size = float(style.get("marker_size", 36.0))
        show_markers = bool(style.get("show_markers", True))
        label_fontsize = float(style.get("label_fontsize", 10.0))
        label_offset_pts = float(style.get("label_offset_pts", 10.0))

        is_dashed = linestyle in ("--", "dashed")
        dashes = style.get("dashes", (1.2, 1.2)) if is_dashed else None
        label_below = bool(style.get("label_below", False)) or is_dashed

        if smooth and len(values) >= 3:
            xs = np.linspace(x.min(), x.max(), num=max(int(smooth_points), len(values) * 120))
            ys = pchip_interpolate(x, y, xs)
            (line,) = ax.plot(
                xs,
                ys,
                linewidth=line_width,
                color=color,
                linestyle=linestyle,
                zorder=2,
            )
        else:
            (line,) = ax.plot(
                x,
                y,
                linewidth=line_width,
                color=color,
                linestyle=linestyle,
                zorder=2,
            )

        line.set_antialiased(True)
        line.set_solid_joinstyle("round")
        line.set_solid_capstyle("round")
        try:
            line.set_dash_joinstyle("round")
            line.set_dash_capstyle("round")
        except Exception:
            pass
        try:
            line.set_snap(False)
        except Exception:
            pass

        if dashes is not None:
            try:
                line.set_dashes(dashes)
            except Exception:
                line.set_linestyle((0, dashes))

        if show_markers:
            ax.scatter(x, y, s=marker_size, color=marker_color, zorder=3)

        last_idx = len(values) - 1
        for i, (xi, yi) in enumerate(zip(x, y)):
            label = (
                f"{yi:.1f}%".replace(".", ",")
                if fmt_as_percent
                else str(yi).replace(".", ",")
            )
            ax.annotate(
                label,
                (xi, yi),
                textcoords="offset points",
                xytext=(0, -abs(label_offset_pts) if label_below else abs(label_offset_pts)),
                ha="center",
                va="top" if label_below else "bottom",
                fontsize=label_fontsize,
                color=color,
                fontweight="bold" if i == last_idx else "normal",
                clip_on=False,
            )

    if y_samples:
        y_min = min(y_samples)
        y_max = max(y_samples)
        y_range = max(1e-9, y_max - y_min)
        pad = y_range * float(y_pad_multiplier)
        ax.set_ylim(y_min - pad, y_max + pad)

    for s in ("left", "right", "top", "bottom"):
        ax.spines[s].set_visible(False)
    ax.set_xticks([])
    ax.set_yticks([])
    ax.tick_params(left=False, bottom=False, labelleft=False, labelbottom=False)
    ax.grid(False)
    ax.margins(x=0.03)
    fig.tight_layout(pad=0.2)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    fig.savefig(output_path, dpi=450, transparent=True, bbox_inches="tight", pad_inches=0.08)
    close_figure(fig)


def _pick_remaining(indices: Iterable[int], size: int) -> list[int]:
    s = set(indices)
    return [i for i in range(size) if i not in s]


def generate_slide2_charts(*, xlsx_path: Path, output_dir: Path) -> list[Path]:
    """Slide 2: gera 05, 06, 07 a partir da aba 'Qualidade Cart 2682'."""

    output_dir.mkdir(parents=True, exist_ok=True)

    wb = load_workbook(filename=xlsx_path, data_only=True)
    sheet_name = "Qualidade Cart 2682"
    if sheet_name not in wb.sheetnames:
        raise ValueError(f"Aba não encontrada: {sheet_name!r}. Disponíveis: {wb.sheetnames}")
    ws = wb[sheet_name]

    raw_titles = read_range_col(ws, "B7:B10")
    titles = [("" if v is None else str(v)).strip() for v in raw_titles]

    raw_x = []
    min_col, min_row, max_col, max_row = (3, 6, 16, 6)  # C6:P6
    for c in range(min_col, max_col + 1):
        raw_x.append(ws.cell(row=min_row, column=c).value)

    xlabels = [("" if v is None else str(v)).strip() for v in raw_x]
    if not any(xlabels):
        xlabels = [str(i + 1) for i in range(len(raw_x))]

    values_rows: list[list[float]] = []
    for r in range(7, 11):
        row_vals = []
        for c in range(3, 16 + 1):
            row_vals.append(ws.cell(row=r, column=c).value)
        values_rows.append(to_float_list(row_vals))

    def _name(i: int) -> str:
        t = titles[i] if i < len(titles) else ""
        return t if t else f"serie{i+1}"

    idx_varejo = _find_title_index(titles, "varejo")
    idx_veiculos = _find_title_index(titles, "veic")

    if idx_varejo is None or idx_veiculos is None or idx_varejo == idx_veiculos:
        idx_varejo, idx_veiculos = 0, 1

    used = {idx_varejo, idx_veiculos}

    # 05) Varejo + Veículos (Veículos pontilhado)
    out05 = output_dir / "05_qualidade_varejo_veiculos.png"
    plot_multi_line(
        xlabels=xlabels,
        series=[
            (_name(idx_varejo), values_rows[idx_varejo], {"color": "#123a7a", "linestyle": "-", "line_width": 3.2}),
            (
                _name(idx_veiculos),
                values_rows[idx_veiculos],
                {"color": "#2f2f2f", "linestyle": "--", "line_width": 3.2, "dashes": (1.2, 1.2)},
            ),
        ],
        output_path=out05,
        fmt_as_percent=False,
        smooth=True,
    )

    # 06/07) Total e Atacado (se achar pelo título; senão pega as restantes)
    idx_total = _find_title_index(titles, "total")
    idx_atacado = _find_title_index(titles, "atacad")

    remaining = _pick_remaining(used, len(values_rows))

    if idx_total is None or idx_total in used:
        idx_total = remaining[0] if remaining else None
    if idx_atacado is None or idx_atacado in used or idx_atacado == idx_total:
        rem2 = [i for i in remaining if i != idx_total]
        idx_atacado = rem2[0] if rem2 else None

    generated = [out05]

    if idx_total is not None:
        out06 = output_dir / "06_qualidade_total.png"
        plot_multi_line(
            xlabels=xlabels,
            series=[(_name(idx_total), values_rows[idx_total], {"color": "#123a7a", "linestyle": "-", "line_width": 3.2})],
            output_path=out06,
            fmt_as_percent=False,
            smooth=True,
        )
        generated.append(out06)

    if idx_atacado is not None:
        out07 = output_dir / "07_qualidade_atacado.png"
        plot_multi_line(
            xlabels=xlabels,
            series=[(_name(idx_atacado), values_rows[idx_atacado], {"color": "#123a7a", "linestyle": "-", "line_width": 3.2})],
            output_path=out07,
            fmt_as_percent=False,
            smooth=True,
        )
        generated.append(out07)

    return generated
