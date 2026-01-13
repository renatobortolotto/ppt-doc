from __future__ import annotations

from pathlib import Path

import numpy as np
from openpyxl import load_workbook
from openpyxl.utils.cell import range_boundaries

from utils.charts_common import ExcelBarChartSpec, close_figure, to_float_list, plot_bar_from_excel


def _read_range_row(ws, cell_range: str) -> list[object]:
    min_col, min_row, max_col, max_row = range_boundaries(cell_range)
    out: list[object] = []
    for r in range(min_row, max_row + 1):
        for c in range(min_col, max_col + 1):
            out.append(ws.cell(row=r, column=c).value)
    return out


def _read_range_col(ws, cell_range: str) -> list[object]:
    min_col, min_row, max_col, max_row = range_boundaries(cell_range)
    out: list[object] = []
    for r in range(min_row, max_row + 1):
        for c in range(min_col, max_col + 1):
            out.append(ws.cell(row=r, column=c).value)
    return out


def _normalize_label(s: str) -> str:
    return (s or "").strip().lower()


def _is_total_row(label: str) -> bool:
    key = _normalize_label(label)
    return key == "total" or key.startswith("total ")


def _fmt_number(v: float, *, decimals: int = 1) -> str:
    try:
        return f"{float(v):.{int(decimals)}f}"
    except Exception:
        return str(v)


def _text_color_for_bg_rgba(rgba) -> str:
    r, g, b = float(rgba[0]), float(rgba[1]), float(rgba[2])
    lum = 0.2126 * r + 0.7152 * g + 0.0722 * b
    return "#ffffff" if lum < 0.50 else "#2f2f2f"


def _read_emprestimos_table(*, xlsx_path: Path):
    wb = load_workbook(filename=xlsx_path, data_only=True)
    sheet_name = "Emprestimos"
    if sheet_name not in wb.sheetnames:
        raise ValueError(f"Aba não encontrada: {sheet_name!r}. Disponíveis: {wb.sheetnames}")
    ws = wb[sheet_name]

    raw_x = _read_range_row(ws, "D4:F4")
    xlabels = [("" if v is None else str(v)).strip() for v in raw_x]
    if not any(xlabels):
        xlabels = ["1", "2", "3"]

    raw_cats = _read_range_col(ws, "C5:C9")
    cat_labels = [("" if v is None else str(v)).strip() for v in raw_cats]

    rows: list[tuple[str, list[float]]] = []
    for r, label in zip(range(5, 10), cat_labels):
        if _is_total_row(label):
            continue
        vals = to_float_list(_read_range_row(ws, f"D{r}:F{r}"))
        rows.append((label, vals))

    return xlabels, rows


def _combine_consignado_demais(rows: list[tuple[str, list[float]]]) -> list[tuple[str, list[float]]]:
    consignado: list[float] | None = None
    demais: list[float] | None = None
    out: list[tuple[str, list[float]]] = []

    for label, vals in rows:
        key = _normalize_label(label)
        if "consign" in key:
            consignado = vals
            continue
        if key == "demais" or "demais" in key:
            demais = vals
            continue
        out.append((label, vals))

    if consignado is None and demais is None:
        return rows

    if demais is None:
        demais = [0.0] * len(consignado or [])
    if consignado is None:
        consignado = [0.0] * len(demais or [])

    merged = [float(d) + float(c) for d, c in zip(demais, consignado)]
    out.append(("Demais", merged))
    return out


def plot_emprestimos_stacked(
    *,
    xlabels: list[str],
    rows: list[tuple[str, list[float]]],
    output_path: Path,
    show_segment_labels_left: bool = True,
    show_values_inside: bool = True,
    left_label_margin_slots: float = 1.15,
    show_delta_pct: bool = True,
    show_delta_bracket: bool = True,
    font_scale: float = 1.0,
    figsize=(10, 4.8),
    dpi: int = 240,
):
    import matplotlib

    matplotlib.use("Agg", force=True)
    import matplotlib.pyplot as plt

    n = len(xlabels)
    x = np.arange(n, dtype=float)

    fig, ax = plt.subplots(figsize=figsize, dpi=int(dpi))

    # Real transparency (figure + axes)
    fig.patch.set_facecolor("none")
    fig.patch.set_alpha(0)
    ax.set_facecolor("none")
    try:
        ax.patch.set_alpha(0)
    except Exception:
        pass

    cmap = plt.cm.Blues
    colors = [cmap(v) for v in np.linspace(0.35, 0.85, num=max(1, len(rows)))]

    bar_width = 0.62
    bottom = np.zeros(n, dtype=float)
    segment_bottoms: list[np.ndarray] = []
    segment_values: list[np.ndarray] = []
    segment_labels: list[str] = []

    for (label, vals), color in zip(rows, colors):
        y = np.asarray(vals, dtype=float)
        segment_labels.append(label)
        segment_bottoms.append(bottom.copy())
        segment_values.append(y.copy())
        ax.bar(x, y, bottom=bottom, color=color, edgecolor="none", width=bar_width)
        bottom = bottom + np.nan_to_num(y, nan=0.0)

    if show_segment_labels_left and rows:
        ax.set_xlim(x.min() - float(left_label_margin_slots), x.max() + 0.9)
        x_label = x.min() - (bar_width / 2.0 + 0.25)
        for j, label in enumerate(segment_labels):
            y_center_ref = None
            for i in range(n):
                val = float(segment_values[j][i]) if i < segment_values[j].size else 0.0
                if np.isfinite(val) and abs(val) > 1e-12:
                    y_center_ref = float(segment_bottoms[j][i]) + val / 2.0
                    break
            if y_center_ref is None:
                continue
            ax.text(
                x_label,
                y_center_ref,
                str(label),
                ha="right",
                va="center",
                fontsize=9 * float(font_scale),
                color="#2f2f2f",
                clip_on=False,
            )

    if show_values_inside and rows:
        for i in range(n):
            for j, color in enumerate(colors):
                val = float(segment_values[j][i]) if i < segment_values[j].size else 0.0
                if not np.isfinite(val) or abs(val) < 1e-12:
                    continue
                y_center = float(segment_bottoms[j][i]) + val / 2.0
                ax.text(
                    float(x[i]),
                    y_center,
                    _fmt_number(val, decimals=1),
                    ha="center",
                    va="center",
                    fontsize=9 * float(font_scale),
                    color=_text_color_for_bg_rgba(color),
                )

    totals = bottom.copy()
    for i in range(n):
        total = float(totals[i])
        if not np.isfinite(total):
            continue
        ax.annotate(
            _fmt_number(total, decimals=1),
            (x[i], total),
            textcoords="offset points",
            xytext=(0, 6),
            ha="center",
            va="bottom",
            fontsize=10 * float(font_scale),
            fontweight="bold",
            color="#2f2f2f",
            clip_on=False,
            zorder=6,
        )

    delta_label_top: float | None = None
    if show_delta_pct and n >= 2:
        vals = np.asarray(totals, dtype=float)
        abs_max = float(np.nanmax(np.abs(vals))) if np.isfinite(np.nanmax(np.abs(vals))) else 0.0
        offset_y = max(abs_max * 0.06, 0.5)
        bracket_h = max(abs_max * 0.03, 0.5)

        for level, i in enumerate(range(1, n)):
            prev = float(vals[i - 1])
            curr = float(vals[i])
            if not np.isfinite(prev) or not np.isfinite(curr) or prev == 0:
                continue
            pct = (curr / prev - 1.0) * 100.0
            label = f"{pct:+.1f}%"

            x1 = float(x[i - 1])
            x2 = float(x[i])
            top = max(prev, curr)
            y_anchor = top + offset_y + level * (bracket_h + offset_y * 0.9)

            if show_delta_bracket:
                ax.plot(
                    [x1, x1, x2, x2],
                    [y_anchor, y_anchor + bracket_h, y_anchor + bracket_h, y_anchor],
                    color="#2f2f2f",
                    linewidth=1.2,
                    solid_capstyle="round",
                    zorder=4,
                )
                text_y = y_anchor + bracket_h + offset_y * 0.25
            else:
                ax.plot(
                    [x1, x2],
                    [prev + offset_y, curr + offset_y],
                    color="#2f2f2f",
                    linewidth=1.2,
                    solid_capstyle="round",
                    zorder=4,
                )
                text_y = top + offset_y * 1.15

            ax.text(
                (x1 + x2) / 2.0,
                text_y,
                label,
                ha="center",
                va="bottom",
                fontsize=9 * float(font_scale),
                color="#2f2f2f",
                zorder=5,
                clip_on=False,
            )
            delta_label_top = text_y if delta_label_top is None else max(delta_label_top, text_y)

        if delta_label_top is not None:
            cur_ymin, cur_ymax = ax.get_ylim()
            ax.set_ylim(cur_ymin, max(cur_ymax, delta_label_top + offset_y))

    ax.set_xticks(x)
    ax.set_xticklabels(xlabels, rotation=0, fontsize=10 * float(font_scale))
    ax.set_yticks([])
    for s in ("left", "right", "top"):
        ax.spines[s].set_visible(False)
    ax.spines["bottom"].set_visible(True)
    ax.tick_params(axis="y", left=False, labelleft=False)
    ax.grid(False)
    ax.margins(x=0.02, y=0.10)

    fig.tight_layout(pad=0.2)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    fig.savefig(
        output_path,
        dpi=int(dpi),
        transparent=True,
        facecolor="none",
        edgecolor="none",
        bbox_inches="tight",
        pad_inches=0.06,
    )
    close_figure(fig)


def generate_slide3_charts(*, xlsx_path: Path, output_dir: Path) -> list[Path]:
    """Slide 3: gera 08, 09."""

    output_dir.mkdir(parents=True, exist_ok=True)

    generated: list[Path] = []

    # 08) Empréstimos - Empilhado
    xlabels, rows = _read_emprestimos_table(xlsx_path=xlsx_path)
    rows2 = _combine_consignado_demais(rows)
    out08 = output_dir / "08_emprestimos_empilhado.png"
    plot_emprestimos_stacked(
        xlabels=xlabels,
        rows=rows2,
        output_path=out08,
        show_segment_labels_left=True,
        show_values_inside=True,
        left_label_margin_slots=1.15,
        show_delta_pct=True,
        show_delta_bracket=True,
        font_scale=1.5,
        dpi=240,
        figsize=(10, 4.8),
    )
    generated.append(out08)

    # 09) Seguros e Cartões - Total
    out09 = output_dir / "09_seguros_cartoes_total.png"
    fig, _ax = plot_bar_from_excel(
        ExcelBarChartSpec(
            file_path=xlsx_path,
            sheet_name="Seguros e Cartoes",
            values_range="D15:F15",
            xlabels_range="D14:F14",
            bar_color="#123a7a",
            show_delta_pct=True,
            show_delta_bracket=True,
            value_decimals=1,
            bar_width_scale=0.70,
            font_scale=1.5,
            output_path=out09,
        )
    )
    close_figure(fig)
    generated.append(out09)

    return generated
