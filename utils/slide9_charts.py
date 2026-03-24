from __future__ import annotations

from pathlib import Path

import numpy as np
from openpyxl import load_workbook
from openpyxl.utils.cell import range_boundaries

from utils.charts_common import close_figure, plot_line_from_excel, to_float_list


def _read_range_row(ws, cell_range: str) -> list[object]:
    min_col, min_row, max_col, max_row = range_boundaries(cell_range)
    out: list[object] = []
    for r in range(min_row, max_row + 1):
        for c in range(min_col, max_col + 1):
            out.append(ws.cell(row=r, column=c).value)
    return out


def _fmt_int(v: float) -> str:
    return f"{float(v):,.0f}".replace(",", ".")


def _as_negative(values: list[float]) -> list[float]:
    return [-abs(float(value)) for value in values]


def _as_positive(values: list[float]) -> list[float]:
    return [abs(float(value)) for value in values]


def _wrap_words(text: str, *, max_line_len: int = 15) -> str:
    s = (text or "").strip()
    if not s:
        return ""

    words = s.split()
    lines: list[str] = []
    current: list[str] = []
    current_len = 0

    for word in words:
        word_len = len(word)
        if not current:
            current = [word]
            current_len = word_len
            continue

        if current_len + 1 + word_len <= max_line_len:
            current.append(word)
            current_len += 1 + word_len
            continue

        lines.append(" ".join(current))
        current = [word]
        current_len = word_len

    if current:
        lines.append(" ".join(current))

    return "\n".join(lines)


def _plot_stacked_bars_with_total(
    *,
    xlabels: list[str],
    series_names: list[str],
    values: np.ndarray,  # shape: [n_bars, n_series]
    output_path: Path,
    colors: list[str],
) -> None:
    import matplotlib.pyplot as plt
    from matplotlib.colors import to_rgba

    if not xlabels or values.size == 0:
        raise ValueError("Sem dados para plotar")
    if values.shape[0] != len(xlabels) or values.shape[1] != len(series_names):
        raise ValueError("Shape de values incompatível com labels/séries")

    n, m = values.shape

    fig, ax = plt.subplots(figsize=(10, 4.8), dpi=240)
    fig.patch.set_alpha(0)
    ax.set_facecolor("none")

    x = np.arange(n, dtype=float)
    width = 0.62

    # Stacking com suporte a segmentos positivos e negativos.
    pos_bottom = np.zeros(n, dtype=float)
    neg_bottom = np.zeros(n, dtype=float)
    segment_centers: list[list[float]] = [[] for _ in range(m)]

    for j in range(m):
        y = np.asarray(values[:, j], dtype=float)
        y_pos = np.where(y > 0, y, 0.0)
        y_neg = np.where(y < 0, y, 0.0)

        cont_pos = ax.bar(
            x,
            y_pos,
            width=width,
            bottom=pos_bottom,
            color=colors[j % len(colors)],
            edgecolor="none",
            zorder=2,
        )
        cont_neg = ax.bar(
            x,
            y_neg,
            width=width,
            bottom=neg_bottom,
            color=colors[j % len(colors)],
            edgecolor="none",
            zorder=2,
        )

        # Label interno por segmento (valor absoluto respeitando sinal).
        rgba = to_rgba(colors[j % len(colors)])
        lum = 0.2126 * rgba[0] + 0.7152 * rgba[1] + 0.0722 * rgba[2]
        txt_color = "#ffffff" if lum < 0.50 else "#2f2f2f"

        for i in range(n):
            v = float(y[i])
            if not np.isfinite(v) or abs(v) < 1e-12:
                segment_centers[j].append(float("nan"))
                continue
            if v > 0:
                yc = float(pos_bottom[i]) + v / 2.0
            else:
                yc = float(neg_bottom[i]) + v / 2.0
            segment_centers[j].append(yc)
            ax.text(
                float(x[i]),
                yc,
                _fmt_int(v),
                ha="center",
                va="center",
                fontsize=9.0,
                color=txt_color,
                zorder=4,
            )

        pos_bottom = pos_bottom + y_pos
        neg_bottom = neg_bottom + y_neg

        # Avoid linter warnings about unused containers in some environments.
        _ = (cont_pos, cont_neg)

    totals = np.nansum(values, axis=1).astype(float)
    # Visual top of each bar (ignoring negative stack) to place "total" label above bars.
    visual_tops = pos_bottom.copy()
    total_label_tops: list[float] = []
    for i, total in enumerate(totals):
        if not np.isfinite(total):
            continue
        y_top = float(visual_tops[i])
        y_label = y_top + max(abs(y_top) * 0.02, 0.6)
        total_label_tops.append(y_label)
        ax.text(
            float(x[i]),
            y_label,
            _fmt_int(total),
            ha="center",
            va="bottom",
            fontsize=10.0,
            fontweight="bold" if i == n - 1 else "normal",
            color="#2f2f2f",
            zorder=5,
            clip_on=False,
        )

    # Brackets e diferenciais no topo.
    # Usa um topo-base único para manter alinhamento visual entre comparações.
    if n >= 2:
        abs_max = float(np.nanmax(np.abs(totals))) if np.isfinite(np.nanmax(np.abs(totals))) else 0.0
        offset_y = max(abs_max * 0.12, 1.0)
        bracket_h = max(abs_max * 0.035, 0.8)
        max_text_y: float | None = None
        # Keep brackets clearly above bar-top totals and internal labels.
        headroom = max(abs_max * 0.22, 3.0)
        top_labels_max = max(total_label_tops) if total_label_tops else float(np.nanmax(visual_tops))
        top_base = float(top_labels_max) + headroom

        for i in range(1, n):
            prev = float(totals[i - 1])
            curr = float(totals[i])
            if not np.isfinite(prev) or not np.isfinite(curr) or prev == 0:
                continue

            pct = (curr / prev - 1.0) * 100.0
            label = f"{pct:+.1f}%".replace(".", ",")

            x1 = float(x[i - 1])
            x2 = float(x[i])
            y_anchor = top_base

            ax.plot(
                [x1, x1, x2, x2],
                [y_anchor, y_anchor + bracket_h, y_anchor + bracket_h, y_anchor],
                color="#2f2f2f",
                linewidth=1.2,
                solid_capstyle="round",
                zorder=4,
            )
            text_y = y_anchor + bracket_h + offset_y * 0.25
            ax.text(
                (x1 + x2) / 2.0,
                text_y,
                label,
                ha="center",
                va="bottom",
                fontsize=9.0,
                color="#2f2f2f",
                zorder=5,
            )
            max_text_y = text_y if max_text_y is None else max(max_text_y, text_y)

        if max_text_y is not None:
            ymin, ymax = ax.get_ylim()
            ax.set_ylim(ymin, max(ymax, max_text_y + offset_y * 1.4))

    # Linha de base explícita para deixar os negativos visualmente abaixo de Y=0.
    ax.axhline(0.0, color="#2f2f2f", linewidth=1.2, zorder=1)

    # Legenda inline: marcador de cor + texto ao lado da cor correspondente.
    x_leg = float(x.min()) - 1.28
    for j, name in enumerate(series_names):
        y_ref = segment_centers[j][-1] if segment_centers[j] else float("nan")
        if not np.isfinite(y_ref):
            for yc in segment_centers[j]:
                if np.isfinite(yc):
                    y_ref = yc
                    break
        if not np.isfinite(y_ref):
            continue
        ax.scatter([x_leg], [float(y_ref)], s=90.0, marker="s", color=colors[j % len(colors)], edgecolors="none", zorder=6)
        ax.text(
            x_leg + 0.16,
            float(y_ref),
            _wrap_words(str(name), max_line_len=15),
            ha="left",
            va="center",
            fontsize=9.0,
            color="#2f2f2f",
            zorder=6,
            clip_on=False,
        )

    ax.set_xlim(float(x.min()) - 1.75, float(x.max()) + 0.65)
    ax.set_xticks(x)
    ax.set_xticklabels(xlabels, fontsize=10.0)
    ax.tick_params(axis="x", bottom=False, pad=8)
    ax.set_yticks([])
    for s in ("left", "right", "top", "bottom"):
        ax.spines[s].set_visible(False)
    ax.tick_params(axis="y", left=False, labelleft=False)
    ax.grid(False)
    ax.margins(x=0.05, y=0.12)

    fig.tight_layout(pad=0.2)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    fig.savefig(output_path, dpi=450, transparent=True, bbox_inches="tight", pad_inches=0.05)
    close_figure(fig)


def _plot_indice_cobertura_percent(
    *,
    xlabels: list[str],
    values: list[float],
    output_path: Path,
    highlight_last_count: int = 3,
) -> None:
    import matplotlib.pyplot as plt
    from matplotlib.patches import FancyBboxPatch

    if len(xlabels) != len(values):
        raise ValueError("xlabels e values com tamanhos diferentes")

    # No XLSX esses valores estão em fração (ex.: 1.72 = 172%).
    vals_pct = np.asarray(values, dtype=float) * 100.0
    x = np.arange(len(vals_pct), dtype=float)

    fig, ax = plt.subplots(figsize=(10, 4.8), dpi=240)
    fig.patch.set_alpha(0)
    ax.set_facecolor("none")

    # All bars in the same dark blue as the last bar style.
    colors = ["#123a7a"] * len(vals_pct)

    bars = ax.bar(x, vals_pct, width=0.62, color=colors, edgecolor="none", zorder=2)

    # Highlight the rightmost bars with a rounded rectangle.
    if highlight_last_count > 0 and len(vals_pct) >= highlight_last_count:
        i0 = len(vals_pct) - highlight_last_count
        i1 = len(vals_pct) - 1
        x_left = float(x[i0] - 0.62 / 2.0 - 0.18)
        x_right = float(x[i1] + 0.62 / 2.0 + 0.18)
        y_top = float(np.nanmax(vals_pct[i0:i1 + 1]))
        y_bottom = 0.0
        pad_y = max(y_top * 0.10, 8.0)
        box = FancyBboxPatch(
            (x_left, y_bottom - pad_y * 0.22),
            x_right - x_left,
            (y_top - y_bottom) + pad_y,
            boxstyle="round,pad=0.02,rounding_size=0.14",
            linewidth=2.0,
            edgecolor="#123a7a",
            facecolor="none",
            zorder=1.5,
        )
        ax.add_patch(box)

    for i, (rect, v) in enumerate(zip(bars, vals_pct)):
        label = f"{v:.0f}%"
        ax.text(
            rect.get_x() + rect.get_width() / 2,
            rect.get_height(),
            label,
            ha="center",
            va="bottom",
            fontsize=10.0,
            fontweight="bold" if i == len(vals_pct) - 1 else "normal",
            color="#2f2f2f",
            zorder=4,
            clip_on=False,
        )

    ax.set_xticks(x)
    ax.set_xticklabels(xlabels, fontsize=10.0)
    ax.tick_params(axis="x", bottom=False, pad=8)
    ax.set_yticks([])
    for s in ("left", "right", "top", "bottom"):
        ax.spines[s].set_visible(False)
    ax.grid(False)
    ax.margins(x=0.05, y=0.22)
    fig.tight_layout(pad=0.2)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    fig.savefig(output_path, dpi=450, transparent=True, bbox_inches="tight", pad_inches=0.05)
    close_figure(fig)


def generate_slide9_charts(*, xlsx_path: Path, output_dir: Path) -> list[Path]:
    """Slide 9: gera custo de crédito (trimestres/9M) e índice de cobertura."""

    output_dir.mkdir(parents=True, exist_ok=True)

    wb = load_workbook(filename=xlsx_path, data_only=True)
    custo_sheet_name = "Tabelas"
    cobertura_sheet_name = "Qualidade Cart 4966"
    if custo_sheet_name not in wb.sheetnames:
        raise ValueError(f"Aba não encontrada: {custo_sheet_name!r}. Disponíveis: {wb.sheetnames}")
    if cobertura_sheet_name not in wb.sheetnames:
        raise ValueError(f"Aba não encontrada: {cobertura_sheet_name!r}. Disponíveis: {wb.sheetnames}")
    ws_custo = wb[custo_sheet_name]
    ws_cobertura = wb[cobertura_sheet_name]

    series_names = ["PDD Expandida", "Recuperação de Crédito"]
    palette = ["#0B2E6B", "#5B8FF9"]

    generated: list[Path] = []

    # 09) Custo de crédito - Trimestres
    tri_labels = [("" if v is None else str(v)).strip() for v in _read_range_row(ws_custo, "G2:H2")]
    tri_pdd = _as_positive(to_float_list(_read_range_row(ws_custo, "G13:H13")))
    tri_rec = _as_negative(to_float_list(_read_range_row(ws_custo, "G5:H5")))
    tri_values = np.column_stack((np.asarray(tri_pdd, dtype=float), np.asarray(tri_rec, dtype=float)))
    out19 = output_dir / "09_custo_credito_trimestres.png"
    _plot_stacked_bars_with_total(
        xlabels=tri_labels,
        series_names=series_names,
        values=tri_values,
        output_path=out19,
        colors=palette,
    )
    generated.append(out19)

    # 09) Custo de crédito - 9M
    ytd_labels = [("" if v is None else str(v)).strip() for v in _read_range_row(ws_custo, "D2:F2")]
    ytd_pdd = _as_positive(to_float_list(_read_range_row(ws_custo, "D13:F13")))
    ytd_rec = _as_negative(to_float_list(_read_range_row(ws_custo, "D5:F5")))
    ytd_values = np.column_stack((np.asarray(ytd_pdd, dtype=float), np.asarray(ytd_rec, dtype=float)))
    out20 = output_dir / "09_custo_credito_9m.png"
    _plot_stacked_bars_with_total(
        xlabels=ytd_labels,
        series_names=series_names,
        values=ytd_values,
        output_path=out20,
        colors=palette,
    )
    generated.append(out20)

    # 09) Índice de cobertura
    cov_labels = [("" if v is None else str(v)).strip() for v in _read_range_row(ws_cobertura, "D2:F2")]
    cov_values = to_float_list(_read_range_row(ws_cobertura, "D17:F17"))
    _plot_indice_cobertura_percent(
        xlabels=cov_labels,
        values=cov_values,
        output_path=output_dir / "09_indice_cobertura.png",
        highlight_last_count=2,
    )
    generated.append(output_dir / "09_indice_cobertura.png")

    # 09) Variação do custo de crédito - Trimestres
    fig, _ax = plot_line_from_excel(
        file_path=xlsx_path,
        sheet_name=custo_sheet_name,
        values_range="D10:F10",
        xlabels_range="D2:F2",
        output_path=output_dir / "09_custo_variacao_custo_credito.png",
        fmt_as_percent=True,
        y_baseline=0.0,
        y_expand=0.10,
        smooth=True,
        line_width=6.0,
        label_fontsize=28.0,
        marker_size=160.0,
        label_offset_pts=20.0,
    )
    close_figure(fig)
    generated.append(output_dir / "09_custo_variacao_custo_credito.png")

    # 09) Variação do custo de crédito - 9M
    fig, _ax = plot_line_from_excel(
        file_path=xlsx_path,
        sheet_name=custo_sheet_name,
        values_range="G10:H10",
        xlabels_range="G2:H2",
        output_path=output_dir / "09_custo_variacao_custo_credito_9m.png",
        fmt_as_percent=True,
        smooth=True,
        line_width=6.0,
        label_fontsize=28.0,
        marker_size=160.0,
        label_offset_pts=20.0,
    )
    close_figure(fig)
    generated.append(output_dir / "09_custo_variacao_custo_credito_9m.png")

    return generated


if __name__ == "__main__":
    xlsx = Path("testing.xlsx")
    out = Path(".")
    if xlsx.exists():
        files = generate_slide9_charts(xlsx_path=xlsx, output_dir=out)
        print(f"Gerados: {files}")
    else:
        print(f"Arquivo {xlsx} não encontrado")
