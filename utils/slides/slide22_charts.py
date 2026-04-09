from __future__ import annotations

from dataclasses import dataclass
from decimal import Decimal, ROUND_HALF_UP
from pathlib import Path

import numpy as np
from openpyxl import load_workbook
from openpyxl.utils.cell import range_boundaries

from utils.charts_common import close_figure


SLIDE22_SHEET_CANDIDATES = ("Qualidade Cart 4966",)
SLIDE22_STACKED_COLORS = ("#123A7A", "#64D6D0", "#2CA44F")
SLIDE22_FIGSIZE = (4.8, 1.6)
SLIDE22_BAR_WIDTH = 0.66
SLIDE22_BAR_SLOT = 1.18
SLIDE22_CALLOUT_DX = 0.18
SLIDE22_CALLOUT_MIN_DY = 9.0
SLIDE22_LINE_COLORS = ("#123A7A", "#64D6D0")
SLIDE22_LINE_FIGSIZE = (4.8, 2.6)
SLIDE22_LINE_FONT_SCALE = 1.4
SLIDE22_SIMPLE_BAR_FIGSIZE = (4.8, 1.8)
SLIDE22_SIMPLE_BAR_COLOR = "#123A7A"
SLIDE22_SIMPLE_BAR_WIDTH = 0.58
SLIDE22_SIMPLE_BAR_SLOT = 1.0


@dataclass(frozen=True)
class Slide22ChartSpec:
    output_name: str
    labels_range: str
    values_range: str


SLIDE22_CHARTS: tuple[Slide22ChartSpec, ...] = (
    Slide22ChartSpec(
        output_name="22_qualidade_4966_bloco1.png",
        labels_range="E30:F30",
        values_range="E32:F34",
    ),
    Slide22ChartSpec(
        output_name="22_qualidade_4966_bloco2.png",
        labels_range="E30:F30",
        values_range="E42:F44",
    ),
    Slide22ChartSpec(
        output_name="22_qualidade_4966_bloco3.png",
        labels_range="E30:F30",
        values_range="E52:F54",
    ),
)

SLIDE22_LINE_CHARTS: tuple[Slide22ChartSpec, ...] = (
    Slide22ChartSpec(
        output_name="22_qualidade_4966_linha1.png",
        labels_range="E30:F30",
        values_range="E37:F38",
    ),
    Slide22ChartSpec(
        output_name="22_qualidade_4966_linha2.png",
        labels_range="E30:F30",
        values_range="E47:F48",
    ),
    Slide22ChartSpec(
        output_name="22_qualidade_4966_linha3.png",
        labels_range="E30:F30",
        values_range="E57:F58",
    ),
)

SLIDE22_REESTRUTURADA_BAR_CHART = Slide22ChartSpec(
    output_name="22_carteira_reestruturada_barras.png",
    labels_range="E61:F61",
    values_range="E64:F64",
)

SLIDE22_REESTRUTURADA_LINE_CHART = Slide22ChartSpec(
    output_name="22_carteira_reestruturada_linha.png",
    labels_range="E61:F61",
    values_range="E65:F65",
)

SLIDE22_COBERTURA_REESTRUTURADA_LINE_CHART = Slide22ChartSpec(
    output_name="22_cobertura_reestruturada_linha.png",
    labels_range="E61:F61",
    values_range="E67:F67",
)

SLIDE22_NPL_BAR_CHART = Slide22ChartSpec(
    output_name="22_npl_barras.png",
    labels_range="E2:F2",
    values_range="E19:F19",
)

SLIDE22_NPL_LINE_CHART = Slide22ChartSpec(
    output_name="22_npl_linha.png",
    labels_range="E2:F2",
    values_range="E20:F20",
)


def _read_range_row(ws, cell_range: str) -> list[object]:
    min_col, min_row, max_col, max_row = range_boundaries(cell_range)
    out: list[object] = []
    for r in range(min_row, max_row + 1):
        for c in range(min_col, max_col + 1):
            out.append(ws.cell(row=r, column=c).value)
    return out


def _read_range_matrix(ws, cell_range: str) -> np.ndarray:
    min_col, min_row, max_col, max_row = range_boundaries(cell_range)
    rows: list[list[float]] = []
    for r in range(min_row, max_row + 1):
        row: list[float] = []
        for c in range(min_col, max_col + 1):
            row.append(_to_float_or_nan(ws.cell(row=r, column=c).value))
        rows.append(row)
    return np.asarray(rows, dtype=float)


def _to_float_or_nan(v: object) -> float:
    if v is None:
        return float("nan")
    if isinstance(v, str):
        s = v.strip()
        if not s:
            return float("nan")
        s = s.replace("%", "").replace(" ", "")
        s = s.replace(".", "").replace(",", ".") if ("," in s and "." in s) else s.replace(",", ".")
        try:
            return float(s)
        except Exception:
            return float("nan")
    try:
        return float(v)
    except Exception:
        return float("nan")


def _fmt_pct(v: float) -> str:
    return f"{float(v):.1f}%".replace(".", ",")


def _fmt_num_int(v: float) -> str:
    rounded = int(Decimal(str(float(v))).quantize(Decimal("1"), rounding=ROUND_HALF_UP))
    return f"{rounded}"


def _fmt_pct_int(v: float) -> str:
    rounded = int(Decimal(str(float(v))).quantize(Decimal("1"), rounding=ROUND_HALF_UP))
    return f"{rounded}%"


def _fmt_bracket_pct(curr: float, prev: float) -> str:
    return f"{((curr / prev) - 1.0) * 100.0:+.1f}%".replace(".", ",")


def _text_color_for_bg_rgba(rgba) -> str:
    r, g, b = float(rgba[0]), float(rgba[1]), float(rgba[2])
    lum = 0.2126 * r + 0.7152 * g + 0.0722 * b
    return "#ffffff" if lum < 0.50 else "#2f2f2f"


def _resolve_sheet_name(wb, candidates: tuple[str, ...], label: str) -> str:
    for candidate in candidates:
        if candidate in wb.sheetnames:
            return candidate
    raise ValueError(f"Aba não encontrada: {label!r}. Disponíveis: {wb.sheetnames}")


def _stack_order_for_bar(values_row: np.ndarray) -> tuple[int, ...]:
    row = np.asarray(values_row, dtype=float)
    return tuple(
        sorted(
            range(len(row)),
            key=lambda idx: (
                1 if not np.isfinite(float(row[idx])) else 0,
                -float(row[idx]) if np.isfinite(float(row[idx])) else 0.0,
                idx,
            ),
        )
    )


def _callout_indices_for_bar(
    values_row: np.ndarray,
    *,
    overlap_threshold: float = SLIDE22_CALLOUT_MIN_DY,
) -> set[int]:
    row = np.asarray(values_row, dtype=float)
    order = _stack_order_for_bar(row)
    centers: list[tuple[int, float, float]] = []
    bottom = 0.0
    for idx in order:
        value = float(row[idx])
        if not np.isfinite(value) or value <= 0:
            continue
        centers.append((idx, bottom + value / 2.0, value))
        bottom += value

    external_indices: set[int] = set()
    for prev, curr in zip(centers, centers[1:]):
        prev_idx, prev_y, prev_value = prev
        curr_idx, curr_y, curr_value = curr
        if abs(curr_y - prev_y) >= float(overlap_threshold):
            continue

        candidate_idx, _candidate_y, _candidate_value = max(
            (prev, curr),
            key=lambda item: (item[1], -item[2], item[0]),
        )
        external_indices.add(candidate_idx)

    return external_indices


def _plot_stacked_pct_block(
    *,
    xlabels: list[str],
    values: np.ndarray,  # [n_bars, n_series]
    output_path: Path,
) -> None:
    import matplotlib.pyplot as plt
    from matplotlib.colors import to_rgba

    values_arr = np.asarray(values, dtype=float)
    if values_arr.ndim != 2:
        raise ValueError("Formato de valores inválido para slide 22")
    n, m = values_arr.shape
    if n == 0 or m == 0:
        raise ValueError("Slide 22 sem dados para plotar")

    fig, ax = plt.subplots(figsize=SLIDE22_FIGSIZE, dpi=240)
    fig.patch.set_alpha(0)
    ax.set_facecolor("none")

    x = np.arange(n, dtype=float) * SLIDE22_BAR_SLOT
    width = SLIDE22_BAR_WIDTH
    totals = np.zeros(n, dtype=float)
    label_top_max = 0.0

    for i in range(n):
        bottom = 0.0
        order = _stack_order_for_bar(values_arr[i, :])
        callout_indices = _callout_indices_for_bar(values_arr[i, :])
        callout_targets: list[tuple[float, float, str, str, str, str]] = []
        for idx in order:
            value = float(values_arr[i, idx])
            if not np.isfinite(value) or value <= 0:
                continue
            color = SLIDE22_STACKED_COLORS[idx % len(SLIDE22_STACKED_COLORS)]
            ax.bar(
                float(x[i]),
                value,
                width=width,
                bottom=bottom,
                color=color,
                edgecolor="none",
                zorder=2,
            )

            txt_color = _text_color_for_bg_rgba(to_rgba(color))
            y = bottom + value / 2.0
            use_badge = value < 8.0
            if idx in callout_indices:
                callout_targets.append(
                    (
                        y,
                        float(x[i]) + width / 2.0,
                        _fmt_pct(value),
                        color,
                        "bold" if idx == order[-1] else "normal",
                        txt_color,
                    )
                )
            else:
                ax.text(
                    float(x[i]),
                    y,
                    _fmt_pct(value),
                    ha="center",
                    va="center",
                    fontsize=9.2 if not use_badge else 8.4,
                    fontweight="bold" if idx == order[-1] else "normal",
                    color=txt_color,
                    zorder=5,
                    clip_on=not use_badge,
                    bbox=(
                        {
                            "facecolor": color,
                            "edgecolor": "none",
                            "boxstyle": "round,pad=0.16",
                        }
                        if use_badge
                        else None
                    ),
                )
            bottom += value

        totals[i] = bottom
        callout_y_positions: list[float] = []
        for y_anchor, x_anchor, label, color, fontweight, txt_color in sorted(callout_targets, key=lambda item: item[0]):
            y_label = y_anchor
            if callout_y_positions and y_label - callout_y_positions[-1] < SLIDE22_CALLOUT_MIN_DY:
                y_label = callout_y_positions[-1] + SLIDE22_CALLOUT_MIN_DY
            callout_y_positions.append(y_label)
            label_top_max = max(label_top_max, y_label)
            x_label = x_anchor + SLIDE22_CALLOUT_DX
            ax.plot(
                [x_anchor + 0.01, x_label - 0.02],
                [y_anchor, y_label],
                color=color,
                linewidth=1.0,
                solid_capstyle="round",
                zorder=4,
                clip_on=False,
            )
            ax.text(
                x_label,
                y_label,
                label,
                ha="left",
                va="center",
                fontsize=8.6,
                fontweight=fontweight,
                color="#2f2f2f",
                zorder=6,
                clip_on=False,
            )

    ax.set_xlim(float(x.min()) - 0.45, float(x.max()) + 0.75)
    ax.set_ylim(0.0, max(float(np.nanmax(totals)) * 1.03, 100.0, label_top_max + 1.2))
    ax.set_xticks(x)
    ax.set_xticklabels(xlabels, fontsize=11.0)
    ax.tick_params(axis="x", bottom=False, pad=8)
    ax.set_yticks([])
    for spine in ("left", "right", "top", "bottom"):
        ax.spines[spine].set_visible(False)
    ax.grid(False)
    ax.margins(x=0.05, y=0.04)

    fig.tight_layout(pad=0.25)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    fig.savefig(output_path, dpi=450, transparent=True, bbox_inches="tight", pad_inches=0.06)
    close_figure(fig)


def _plot_two_point_lines(
    *,
    xlabels: list[str],
    values: np.ndarray,  # [n_series, n_points]
    output_path: Path,
    label_decimals: int = 0,
    label_font_scale: float = 1.0,
    x_tick_font_scale: float = 1.0,
    label_offset_factor: float = 0.06,
    label_offset_min: float = 1.6,
    y_pad_factor: float = 0.35,
    y_pad_min: float = 3.5,
    show_xlabels: bool = True,
) -> None:
    import matplotlib.pyplot as plt

    values_arr = np.asarray(values, dtype=float)
    if values_arr.ndim != 2:
        raise ValueError("Formato de valores inválido para linhas do slide 22")
    n_series, n_points = values_arr.shape
    if n_series == 0 or n_points == 0:
        raise ValueError("Slide 22 sem dados de linha para plotar")

    fig, ax = plt.subplots(figsize=SLIDE22_LINE_FIGSIZE, dpi=240)
    fig.patch.set_alpha(0)
    ax.set_facecolor("none")

    x = np.arange(n_points, dtype=float)
    finite_values = values_arr[np.isfinite(values_arr)]
    y_min = float(np.nanmin(finite_values)) if finite_values.size else 0.0
    y_max = float(np.nanmax(finite_values)) if finite_values.size else 0.0
    y_pad = max((y_max - y_min) * float(y_pad_factor), float(y_pad_min))

    for idx in range(n_series):
        series = np.asarray(values_arr[idx, :], dtype=float)
        color = SLIDE22_LINE_COLORS[idx % len(SLIDE22_LINE_COLORS)]
        ax.plot(
            x,
            series,
            color=color,
            linewidth=3.0,
            marker="o",
            markersize=6.5,
            markerfacecolor=color,
            markeredgewidth=0.0,
            zorder=2,
        )
        for point_idx, (xi, yi) in enumerate(zip(x, series)):
            if not np.isfinite(yi):
                continue
            label_offset = max((y_max - y_min) * float(label_offset_factor), float(label_offset_min))
            direction = -1.0 if idx > 0 else 1.0
            ax.text(
                float(xi),
                float(yi) + label_offset * direction,
                _fmt_pct_int(yi) if label_decimals == 0 else _fmt_pct(yi),
                ha="center",
                va="bottom" if direction > 0 else "top",
                fontsize=9.4 * SLIDE22_LINE_FONT_SCALE * float(label_font_scale),
                fontweight="normal",
                color=color,
                zorder=3,
                clip_on=False,
            )

    ax.set_xlim(float(x.min()) - 0.08, float(x.max()) + 0.08)
    ax.set_ylim(max(0.0, y_min - y_pad), min(100.0, y_max + y_pad))
    if show_xlabels:
        ax.set_xticks(x)
        ax.set_xticklabels(xlabels, fontsize=11.0 * SLIDE22_LINE_FONT_SCALE * float(x_tick_font_scale))
        ax.tick_params(axis="x", bottom=False, pad=8)
    else:
        ax.set_xticks([])
        ax.tick_params(axis="x", bottom=False, labelbottom=False)
    ax.set_yticks([])
    for spine in ("left", "right", "top", "bottom"):
        ax.spines[spine].set_visible(False)
    ax.grid(False)
    ax.margins(x=0.03, y=0.10)

    fig.tight_layout(pad=0.25)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    fig.savefig(output_path, dpi=450, transparent=True, bbox_inches="tight", pad_inches=0.06)
    close_figure(fig)


def _plot_two_bar_values(
    *,
    xlabels: list[str],
    values: np.ndarray,  # [n_points]
    output_path: Path,
    bar_width: float = SLIDE22_SIMPLE_BAR_WIDTH,
    bar_slot: float = SLIDE22_SIMPLE_BAR_SLOT,
    font_scale: float = 1.0,
    bracket_anchor: str = "center",
    bracket_top_gap_scale: float = 0.08,
    bracket_top_gap_min: float = 18.0,
    bracket_label_clearance: float = 0.0,
    x_margin: float | None = None,
) -> None:
    import matplotlib.pyplot as plt

    vals = np.asarray(values, dtype=float)
    n = len(vals)
    if n == 0:
        raise ValueError("Sem dados para bar chart do slide 22")

    fig, ax = plt.subplots(figsize=SLIDE22_SIMPLE_BAR_FIGSIZE, dpi=240)
    fig.patch.set_alpha(0)
    ax.set_facecolor("none")

    width = float(bar_width)
    x = np.arange(n, dtype=float) * float(bar_slot)
    bars = ax.bar(
        x,
        vals,
        width=width,
        color=SLIDE22_SIMPLE_BAR_COLOR,
        edgecolor="none",
        zorder=2,
    )

    label_tops: list[float] = []
    for i, (rect, value) in enumerate(zip(bars, vals)):
        if not np.isfinite(value):
            continue
        y_label = float(rect.get_height()) + max(abs(float(value)) * 0.03, 10.0)
        label_tops.append(y_label)
        ax.text(
            rect.get_x() + rect.get_width() / 2.0,
            y_label,
            _fmt_num_int(value),
            ha="center",
            va="bottom",
            fontsize=10.0 * float(font_scale),
            fontweight="bold" if i == n - 1 else "normal",
            color="#2f2f2f",
            zorder=4,
            clip_on=False,
        )

    if n >= 2:
        abs_max = float(np.nanmax(np.abs(vals))) if np.isfinite(np.nanmax(np.abs(vals))) else 0.0
        offset_y = max(abs_max * 0.08, 16.0)
        bracket_h = max(abs_max * 0.03, 10.0)
        top_labels_max = max(label_tops) if label_tops else float(np.nanmax(vals))
        top_base = (
            float(top_labels_max)
            + max(abs_max * float(bracket_top_gap_scale), float(bracket_top_gap_min))
            + float(bracket_label_clearance)
        )
        prev = float(vals[0])
        curr = float(vals[1])
        if np.isfinite(prev) and np.isfinite(curr) and abs(prev) > 1e-12:
            if bracket_anchor == "outer_edges":
                edge_pad = width * 0.08
                x1 = float(x[0]) - width / 2.0 - edge_pad
                x2 = float(x[1]) + width / 2.0 + edge_pad
            elif bracket_anchor == "inner_edges":
                x1 = float(x[0]) + width / 2.0
                x2 = float(x[1]) - width / 2.0
            else:
                x1 = float(x[0])
                x2 = float(x[1])
            text_y = top_base + bracket_h + offset_y * 0.25
            ax.plot(
                [x1, x1, x2, x2],
                [top_base, top_base + bracket_h, top_base + bracket_h, top_base],
                color="#2f2f2f",
                linewidth=1.2,
                solid_capstyle="round",
                zorder=4,
            )
            ax.text(
                (x1 + x2) / 2.0,
                text_y,
                _fmt_bracket_pct(curr, prev),
                ha="center",
                va="bottom",
                fontsize=9.2 * float(font_scale),
                color="#2f2f2f",
                zorder=5,
            )
            ymin, ymax = ax.get_ylim()
            ax.set_ylim(ymin, max(ymax, text_y + offset_y * 0.8))

    side_margin = float(x_margin) if x_margin is not None else max(width * 0.7, 0.22)
    ax.set_xlim(float(x.min()) - side_margin, float(x.max()) + side_margin)
    ax.set_xticks(x)
    ax.set_xticklabels(xlabels, fontsize=11.0 * float(font_scale))
    ax.tick_params(axis="x", bottom=False, pad=8)
    ax.set_yticks([])
    for spine in ("left", "right", "top", "bottom"):
        ax.spines[spine].set_visible(False)
    ax.grid(False)
    ax.margins(x=0.05, y=0.06)

    fig.tight_layout(pad=0.25)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    fig.savefig(output_path, dpi=450, transparent=True, bbox_inches="tight", pad_inches=0.06)
    close_figure(fig)


def generate_slide22_charts(*, xlsx_path: Path, output_dir: Path) -> list[Path]:
    """Slide 22: primeiros três gráficos empilhados da aba Qualidade Cart 4966."""

    output_dir.mkdir(parents=True, exist_ok=True)

    wb = load_workbook(filename=xlsx_path, data_only=True)
    ws = wb[_resolve_sheet_name(wb, SLIDE22_SHEET_CANDIDATES, "Qualidade Cart 4966")]

    generated: list[Path] = []
    for chart_spec in SLIDE22_CHARTS:
        xlabels = [
            ("" if value is None else str(value)).strip()
            for value in _read_range_row(ws, chart_spec.labels_range)
        ]
        raw_values = _read_range_matrix(ws, chart_spec.values_range)
        values = raw_values.T.astype(float)
        if np.nanmax(np.abs(values)) <= 1.5:
            values = values * 100.0

        output_path = output_dir / chart_spec.output_name
        _plot_stacked_pct_block(
            xlabels=xlabels,
            values=values,
            output_path=output_path,
        )
        generated.append(output_path)

    for chart_spec in SLIDE22_LINE_CHARTS:
        xlabels = [
            ("" if value is None else str(value)).strip()
            for value in _read_range_row(ws, chart_spec.labels_range)
        ]
        values = _read_range_matrix(ws, chart_spec.values_range).astype(float)
        if np.nanmax(np.abs(values)) <= 1.5:
            values = values * 100.0

        output_path = output_dir / chart_spec.output_name
        _plot_two_point_lines(
            xlabels=xlabels,
            values=values,
            output_path=output_path,
        )
        generated.append(output_path)

    xlabels_reestruturada = [
        ("" if value is None else str(value)).strip()
        for value in _read_range_row(ws, SLIDE22_REESTRUTURADA_BAR_CHART.labels_range)
    ]

    reestruturada_bar_values = np.asarray(
        [_to_float_or_nan(v) for v in _read_range_row(ws, SLIDE22_REESTRUTURADA_BAR_CHART.values_range)],
        dtype=float,
    ) / 1000.0
    reestruturada_bar_output = output_dir / SLIDE22_REESTRUTURADA_BAR_CHART.output_name
    _plot_two_bar_values(
        xlabels=xlabels_reestruturada,
        values=reestruturada_bar_values,
        output_path=reestruturada_bar_output,
        bar_width=0.066,
        bar_slot=0.096,
        font_scale=1.5,
        bracket_anchor="center",
        bracket_top_gap_scale=0.18,
        bracket_top_gap_min=55.0,
        bracket_label_clearance=24.0,
        x_margin=0.12,
    )
    generated.append(reestruturada_bar_output)

    reestruturada_line_values = np.asarray(
        [_to_float_or_nan(v) for v in _read_range_row(ws, SLIDE22_REESTRUTURADA_LINE_CHART.values_range)],
        dtype=float,
    )
    if np.nanmax(np.abs(reestruturada_line_values)) <= 1.5:
        reestruturada_line_values = reestruturada_line_values * 100.0
    reestruturada_line_output = output_dir / SLIDE22_REESTRUTURADA_LINE_CHART.output_name
    _plot_two_point_lines(
        xlabels=xlabels_reestruturada,
        values=reestruturada_line_values.reshape(1, -1),
        output_path=reestruturada_line_output,
        label_decimals=1,
        label_font_scale=2.24,
        label_offset_factor=0.008,
        label_offset_min=0.05,
        y_pad_factor=0.18,
        y_pad_min=0.28,
        show_xlabels=False,
    )
    generated.append(reestruturada_line_output)

    cobertura_line_values = np.asarray(
        [_to_float_or_nan(v) for v in _read_range_row(ws, SLIDE22_COBERTURA_REESTRUTURADA_LINE_CHART.values_range)],
        dtype=float,
    )
    if np.nanmax(np.abs(cobertura_line_values)) <= 1.5:
        cobertura_line_values = cobertura_line_values * 100.0
    cobertura_line_output = output_dir / SLIDE22_COBERTURA_REESTRUTURADA_LINE_CHART.output_name
    _plot_two_point_lines(
        xlabels=xlabels_reestruturada,
        values=cobertura_line_values.reshape(1, -1),
        output_path=cobertura_line_output,
        label_decimals=1,
    )
    generated.append(cobertura_line_output)

    xlabels_npl = [
        ("" if value is None else str(value)).strip()
        for value in _read_range_row(ws, SLIDE22_NPL_BAR_CHART.labels_range)
    ]

    npl_bar_values = np.asarray(
        [_to_float_or_nan(v) for v in _read_range_row(ws, SLIDE22_NPL_BAR_CHART.values_range)],
        dtype=float,
    )
    npl_bar_output = output_dir / SLIDE22_NPL_BAR_CHART.output_name
    _plot_two_bar_values(
        xlabels=xlabels_npl,
        values=npl_bar_values,
        output_path=npl_bar_output,
        bracket_top_gap_scale=0.12,
        bracket_top_gap_min=36.0,
        bracket_label_clearance=18.0,
    )
    generated.append(npl_bar_output)

    npl_line_values = np.asarray(
        [_to_float_or_nan(v) for v in _read_range_row(ws, SLIDE22_NPL_LINE_CHART.values_range)],
        dtype=float,
    )
    if np.nanmax(np.abs(npl_line_values)) <= 1.5:
        npl_line_values = npl_line_values * 100.0
    npl_line_output = output_dir / SLIDE22_NPL_LINE_CHART.output_name
    _plot_two_point_lines(
        xlabels=xlabels_npl,
        values=npl_line_values.reshape(1, -1),
        output_path=npl_line_output,
        label_decimals=1,
        label_font_scale=1.34,
        label_offset_factor=0.008,
        label_offset_min=0.05,
        show_xlabels=False,
    )
    generated.append(npl_line_output)

    return generated


if __name__ == "__main__":
    xlsx = Path("testing.xlsx")
    out = Path(".")
    if xlsx.exists():
        files = generate_slide22_charts(xlsx_path=xlsx, output_dir=out)
        print(f"Gerados: {files}")
    else:
        print(f"Arquivo {xlsx} não encontrado")
