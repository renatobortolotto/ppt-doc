from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook
from openpyxl.utils.cell import range_boundaries

from utils.charts_common import (
    ExcelBarChartSpec,
    close_figure,
    plot_bar_from_excel,
    plot_donut_chart,
    to_float_list,
)

SLIDE4_DONUT_ROW_SPECS: tuple[tuple[str, str, tuple[int, ...]], ...] = (
    ("Veiculos Leves", "Veiculos Leves Usados", (16,)),
    ("Growth", "Outros Veiculos", (17, 18)),
    ("Growth", "Paineis Solares", (19,)),
    ("Growth", "EGV", (23,)),
    ("Growth", "Cartões", (24, 25)),
    ("Atacado", "Corporate", (28, 35)),
    ("Atacado", "Large Corporate + instituicoes financeiras", (29, 36)),
    ("Atacado", "Pequenas e Medias Empresas (PME)", (30,)),
)


def _extract_slide4_donut_series(ws, *, source_range: str) -> tuple[list[str], list[str], list[float]]:
    min_col, min_row, max_col, max_row = range_boundaries(source_range)
    if min_col == max_col:
        raise ValueError(f"Range do donut precisa ter ao menos 2 colunas: {source_range}")

    value_col = max_col

    categories: list[str] = []
    labels: list[str] = []
    values: list[float] = []
    for category, display_label, rows in SLIDE4_DONUT_ROW_SPECS:
        row_values: list[object] = []
        for row in rows:
            if row < min_row or row > max_row:
                continue
            row_values.append(ws.cell(row=row, column=value_col).value)
        if not row_values:
            continue
        try:
            value = float(sum(to_float_list(row_values)))
        except (TypeError, ValueError) as exc:
            raise ValueError(
                f"Valor nao numerico para o donut em {ws.title}!{display_label}: {row_values!r}"
            ) from exc
        if abs(value) <= 1e-12:
            continue
        categories.append(category)
        labels.append(display_label)
        values.append(value)

    if not values:
        raise ValueError(f"Nenhum item mapeado para o donut no range {source_range}")

    return categories, labels, values


def generate_slide4_charts(*, xlsx_path: Path, output_dir: Path) -> list[Path]:
    """Slide 4: gera 10, 11, 12."""

    output_dir.mkdir(parents=True, exist_ok=True)

    generated: list[Path] = []

    # 10) Grafico de Carteira de Credito Ampliada (donut)
    wb = load_workbook(filename=xlsx_path, data_only=True)
    donut_sheet_name = "Carteira"
    if donut_sheet_name not in wb.sheetnames:
        raise ValueError(f"Aba nao encontrada: {donut_sheet_name!r}. Disponiveis: {wb.sheetnames}")
    donut_ws = wb[donut_sheet_name]
    categories, labels, values = _extract_slide4_donut_series(donut_ws, source_range="C12:F36")
    fig, _ax = plot_donut_chart(
        categories=categories,
        labels=labels,
        values=values,
        center_text="",
        title=None,
        output_path=output_dir / "10_pizza_carteira.png",
        figsize=(16, 12),
        font_scale=1.5,
        mirror_horizontal=True,
    )
    close_figure(fig)
    generated.append(output_dir / "10_pizza_carteira.png")

    # 11) Barras - Trimestres (H3:J3 labels, H4:J4 valores)
    fig, _ax = plot_bar_from_excel(
        ExcelBarChartSpec(
            file_path=xlsx_path,
            sheet_name="Pizza Teste",
            values_range="H4:J4",
            xlabels_range="H3:J3",
            title=None,
            highlight_last=True,
            bar_color="#123a7a",
            show_delta_pct=True,
            show_delta_bracket=True,
            delta_pairs=((0, 1), (1, 2)),
            font_scale=1.5,
            output_path=output_dir / "11_pizza_trimestres.png",
        )
    )
    close_figure(fig)
    generated.append(output_dir / "11_pizza_trimestres.png")

    # 12) Barras - 9M (K3:L3 labels, K4:L4 valores)
    fig, _ax = plot_bar_from_excel(
        ExcelBarChartSpec(
            file_path=xlsx_path,
            sheet_name="Pizza Teste",
            values_range="K4:L4",
            xlabels_range="K3:L3",
            title=None,
            highlight_last=True,
            bar_color="#123a7a",
            show_delta_pct=True,
            show_delta_bracket=True,
            fixed_slot_count=9,
            font_scale=1.5,
            output_path=output_dir / "12_pizza_9m.png",
        )
    )
    close_figure(fig)
    generated.append(output_dir / "12_pizza_9m.png")

    return generated


if __name__ == "__main__":
    xlsx = Path("testing.xlsx")
    out = Path(".")
    if xlsx.exists():
        files = generate_slide4_charts(xlsx_path=xlsx, output_dir=out)
        print(f"Gerados: {files}")
    else:
        print(f"Arquivo {xlsx} nao encontrado")
